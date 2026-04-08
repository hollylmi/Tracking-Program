from datetime import date, datetime, timedelta

from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, jsonify
from blueprints.auth import require_role

from models import (db, Project, Machine, MachineGroup, DailyEntry, HiredMachine, PlannedData,
                    ProjectNonWorkDate, ProjectBudgetedRole, ProjectMachine,
                    ProjectWorkedSunday, ProjectDocument, Role, PublicHoliday,
                    CFMEUDate, ProjectEquipmentRequirement, ProjectEquipmentAssignment)
from utils.progress import compute_project_progress, compute_delay_summary, compute_material_productivity
from utils.gantt import compute_gantt_data
from utils.reports import generate_project_report_pdf, generate_weekly_report_pdf
from utils.settings import load_settings, get_airports, get_locations
from utils.helpers import _natural_key

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

projects_bp = Blueprint('projects', __name__)


# ---------------------------------------------------------------------------
# Progress API — running totals for entry form
# ---------------------------------------------------------------------------

@projects_bp.route('/api/progress')
@require_role('admin', 'supervisor', 'site')
def api_progress():
    project_id = request.args.get('project_id', '')
    lot = request.args.get('lot', '').strip()
    material = request.args.get('material', '').strip()
    exclude_entry_id = request.args.get('exclude_id', '')

    if not project_id:
        return jsonify({'error': 'project_id required'}), 400

    # Planned SQM for this lot+material (case-insensitive matching)
    planned_q = PlannedData.query.filter_by(project_id=int(project_id))
    if lot:
        planned_q = planned_q.filter(db.func.upper(db.func.trim(PlannedData.lot)) == lot.strip().upper())
    if material:
        planned_q = planned_q.filter(db.func.upper(db.func.trim(PlannedData.material)) == material.strip().upper())
    planned_sqm = sum(p.planned_sqm or 0 for p in planned_q.all())

    # Installed SQM so far — query production lines first, then legacy entries
    from models import EntryProductionLine
    pl_q = (db.session.query(db.func.sum(EntryProductionLine.install_sqm))
            .join(DailyEntry, EntryProductionLine.entry_id == DailyEntry.id)
            .filter(DailyEntry.project_id == int(project_id)))
    if lot:
        pl_q = pl_q.filter(db.func.upper(db.func.trim(EntryProductionLine.lot_number)) == lot.strip().upper())
    if material:
        pl_q = pl_q.filter(db.func.upper(db.func.trim(EntryProductionLine.material)) == material.strip().upper())
    if exclude_entry_id:
        try:
            pl_q = pl_q.filter(DailyEntry.id != int(exclude_entry_id))
        except ValueError:
            pass
    pl_total = pl_q.scalar() or 0

    # Legacy entries (no production lines)
    legacy_q = (DailyEntry.query
                .filter_by(project_id=int(project_id))
                .filter(DailyEntry.install_sqm > 0)
                .filter(~DailyEntry.id.in_(
                    db.session.query(EntryProductionLine.entry_id).distinct()
                )))
    if lot:
        legacy_q = legacy_q.filter(db.func.upper(db.func.trim(DailyEntry.lot_number)) == lot.strip().upper())
    if material:
        legacy_q = legacy_q.filter(db.func.upper(db.func.trim(DailyEntry.material)) == material.strip().upper())
    if exclude_entry_id:
        try:
            legacy_q = legacy_q.filter(DailyEntry.id != int(exclude_entry_id))
        except ValueError:
            pass
    legacy_total = sum(e.install_sqm or 0 for e in legacy_q.all())

    installed_sqm = pl_total + legacy_total

    remaining = max(0, planned_sqm - installed_sqm)
    pct = round(installed_sqm / planned_sqm * 100, 1) if planned_sqm > 0 else None

    return jsonify({
        'planned_sqm': round(planned_sqm, 2),
        'installed_sqm': round(installed_sqm, 2),
        'remaining': round(remaining, 2),
        'pct_complete': pct,
        'has_planned': planned_sqm > 0,
    })


@projects_bp.route('/api/progress-debug')
@require_role('admin')
def api_progress_debug():
    """Debug: show every entry contributing to a material's installed SQM."""
    project_id = request.args.get('project_id', type=int)
    material = request.args.get('material', '').strip()
    if not project_id or not material:
        return jsonify({'error': 'project_id and material required'}), 400

    from models import EntryProductionLine
    entries = DailyEntry.query.filter_by(project_id=project_id).order_by(DailyEntry.entry_date).all()
    rows = []
    total = 0
    for e in entries:
        if e.production_lines:
            for pl in e.production_lines:
                if (pl.material or '').strip().upper() == material.strip().upper():
                    sqm = pl.install_sqm or 0
                    total += sqm
                    rows.append({
                        'entry_id': e.id,
                        'date': e.entry_date.strftime('%Y-%m-%d'),
                        'source': 'production_line',
                        'lot': pl.lot_number,
                        'material': pl.material,
                        'sqm': sqm,
                        'hours': pl.install_hours,
                        'running_total': round(total, 2),
                    })
        else:
            if (e.material or '').strip().upper() == material.strip().upper() and (e.install_sqm or 0) > 0:
                sqm = e.install_sqm or 0
                total += sqm
                rows.append({
                    'entry_id': e.id,
                    'date': e.entry_date.strftime('%Y-%m-%d'),
                    'source': 'legacy',
                    'lot': e.lot_number,
                    'material': e.material,
                    'sqm': sqm,
                    'hours': e.install_hours,
                    'running_total': round(total, 2),
                })

    # Planned
    planned_sqm = sum(p.planned_sqm or 0 for p in PlannedData.query.filter_by(
        project_id=project_id, material=material).all())

    return jsonify({
        'material': material,
        'planned_sqm': planned_sqm,
        'total_actual_sqm': round(total, 2),
        'pct': round(total / planned_sqm * 100, 1) if planned_sqm > 0 else None,
        'entry_count': len(rows),
        'entries': rows,
    })


# ---------------------------------------------------------------------------
# Project Dashboard
# ---------------------------------------------------------------------------

@projects_bp.route('/api/project/<int:project_id>/machines')
@require_role('admin', 'supervisor', 'site')
def api_project_machines(project_id):
    """Return owned machines assigned to this project."""
    assignments = (ProjectMachine.query
                   .filter_by(project_id=project_id)
                   .join(Machine)
                   .filter(Machine.active == True)
                   .all())
    machines = [{'id': pm.machine.id, 'name': pm.machine.name,
                 'machine_type': pm.machine.machine_type or ''}
                for pm in assignments]
    return jsonify({'machines': machines})


@projects_bp.route('/api/planned-options')
@require_role('admin', 'supervisor', 'site')
def api_planned_options():
    """Return distinct lots (and materials for a given lot) from planned data."""
    project_id = request.args.get('project_id', '').strip()
    lot = request.args.get('lot', '').strip()
    if not project_id:
        return jsonify({'lots': [], 'materials': []})
    q = PlannedData.query.filter_by(project_id=int(project_id))
    rows = q.all()
    lots = sorted({r.lot for r in rows if r.lot}, key=_natural_key)
    if lot:
        materials = sorted({r.material for r in rows if r.material and r.lot == lot},
                           key=_natural_key)
    else:
        materials = sorted({r.material for r in rows if r.material}, key=_natural_key)
    return jsonify({'lots': lots, 'materials': materials})


@projects_bp.route('/project/<int:project_id>/dashboard')
@require_role('admin', 'supervisor')
def project_dashboard(project_id):
    project = Project.query.get_or_404(project_id)
    progress = compute_project_progress(project_id)
    gantt_data = compute_gantt_data(project_id)
    delay_summary = compute_delay_summary(project_id)
    productivity = compute_material_productivity(project_id)
    non_work_dates = ProjectNonWorkDate.query.filter_by(project_id=project_id).order_by(ProjectNonWorkDate.date).all()
    has_openpyxl = HAS_OPENPYXL

    # Budgeted crew vs actual — actual counts employees by role from all entries
    budgeted_roles = ProjectBudgetedRole.query.filter_by(project_id=project_id).order_by(ProjectBudgetedRole.role_name).all()
    all_roles = Role.query.order_by(Role.name).all()

    # Build actual crew counts by role from entries (last 30 days or all time if sparse)
    actual_role_counts = {}
    all_entries_for_crew = DailyEntry.query.filter_by(project_id=project_id).all()
    for e in all_entries_for_crew:
        for emp in e.employees:
            role_name = emp.role or 'Unassigned'
            actual_role_counts[role_name] = actual_role_counts.get(role_name, 0) + 1
    # Convert to per-entry average
    entry_count = len(all_entries_for_crew) or 1
    actual_role_avg = {k: round(v / entry_count, 1) for k, v in actual_role_counts.items()}

    # Hired machines at this project
    hired_machines_list = HiredMachine.query.filter_by(project_id=project_id, active=True).all()

    # Owned machines assigned to this project
    owned_machine_assignments = (ProjectMachine.query
                                  .filter_by(project_id=project_id)
                                  .join(Machine)
                                  .filter(Machine.active == True)
                                  .all())
    owned_machines_list = owned_machine_assignments  # list of ProjectMachine rows

    # Worked Sundays for this project
    worked_sundays_list = (ProjectWorkedSunday.query
                           .filter_by(project_id=project_id)
                           .order_by(ProjectWorkedSunday.date)
                           .all())

    # Project documents
    project_documents = (ProjectDocument.query
                         .filter_by(project_id=project_id)
                         .order_by(ProjectDocument.uploaded_at.desc())
                         .all())

    # All owned machines (for "assign" dropdown)
    all_machines = Machine.query.filter_by(active=True).order_by(Machine.name).all()
    # Already-assigned machine ids
    assigned_machine_ids = {pm.machine_id for pm in owned_machine_assignments}

    # Cost estimation: labour (budgeted roles × hourly rate) + owned machines + hired machines
    role_rate_map = {r.name: (r.delay_rate or 0) for r in all_roles}
    hours_pd = project.hours_per_day or 8
    labour_daily = sum(
        br.budgeted_count * role_rate_map.get(br.role_name, 0) * hours_pd
        for br in budgeted_roles
    )
    # Hired machine daily cost — always derived from weekly rate
    hired_machine_daily = 0.0
    for hm in hired_machines_list:
        days_pw = 6 if hm.count_saturdays else 5
        hired_machine_daily += hm.cost_per_week / days_pw if hm.cost_per_week else 0
    # Owned machine daily cost (delay_rate × hours_per_day)
    owned_machine_daily = 0.0
    for pm in owned_machine_assignments:
        if pm.machine and pm.machine.delay_rate:
            owned_machine_daily += pm.machine.delay_rate * hours_pd
    machine_daily = hired_machine_daily + owned_machine_daily
    daily_cost = labour_daily + machine_daily
    target_cost = (daily_cost * project.quoted_days
                   if project.quoted_days and daily_cost > 0 else None)
    # Public holidays and CFMEU dates for this project's state
    state_holidays = []
    all_public = PublicHoliday.query.order_by(PublicHoliday.date).all()
    if project.state:
        state_holidays = [h for h in all_public
                          if 'ALL' in h.states_list() or project.state in h.states_list()]
        if project.is_cfmeu:
            cfmeu = [c for c in CFMEUDate.query.order_by(CFMEUDate.date).all()
                     if 'ALL' in c.states_list() or project.state in c.states_list()]
            state_holidays = sorted(state_holidays + cfmeu, key=lambda x: x.date)
    else:
        # No state set — still show national (ALL) holidays
        state_holidays = [h for h in all_public if 'ALL' in h.states_list()]
    # Build non-work date set including state holidays
    holiday_dates = {h.date for h in state_holidays}
    non_work_set_dates = {nwd.date for nwd in non_work_dates} | holiday_dates
    est_finish_date = None
    if gantt_data and gantt_data.get('est_finish'):
        try:
            from datetime import datetime as _dt
            est_finish_date = _dt.strptime(gantt_data['est_finish'], '%d/%m/%Y').date()
        except Exception:
            pass
    forecast_working_days = None
    forecast_cost = None
    if est_finish_date and project.start_date:
        forecast_working_days = sum(
            1 for i in range((est_finish_date - project.start_date).days + 1)
            if (project.start_date + timedelta(days=i)).weekday() != 6
            and (project.start_date + timedelta(days=i)) not in non_work_set_dates
        )
        forecast_cost = daily_cost * forecast_working_days if daily_cost > 0 else None
    cost_variance = ((forecast_cost - target_cost)
                     if forecast_cost is not None and target_cost is not None else None)
    cost_estimate = {
        'labour_daily': round(labour_daily, 2),
        'hired_machine_daily': round(hired_machine_daily, 2),
        'owned_machine_daily': round(owned_machine_daily, 2),
        'machine_daily': round(machine_daily, 2),
        'daily_total': round(daily_cost, 2),
        'target_cost': round(target_cost, 2) if target_cost is not None else None,
        'forecast_cost': round(forecast_cost, 2) if forecast_cost is not None else None,
        'cost_variance': round(cost_variance, 2) if cost_variance is not None else None,
        'forecast_working_days': forecast_working_days,
        'has_rates': daily_cost > 0,
    }

    # All distinct scheduling groups for the scheduling tab
    _all_roles = Role.query.filter(Role.group_name.isnot(None)).all()
    _seen = set()
    role_groups_all = []
    for r in _all_roles:
        if r.group_name and r.group_name not in _seen:
            _seen.add(r.group_name)
            role_groups_all.append(r.group_name)
    role_groups_all.sort()
    # Build a dict of current budgeted counts by role_name for pre-filling
    budgeted_by_group = {br.role_name: br.budgeted_count for br in budgeted_roles}

    # Equipment requirements for scheduling tab
    equip_reqs = (ProjectEquipmentRequirement.query
                  .filter_by(project_id=project_id)
                  .order_by(ProjectEquipmentRequirement.label)
                  .all())
    # Build coverage rows — assignments are explicit (not type-matched)
    equip_coverage = []
    # Track which machine/hired_machine ids are already assigned to any requirement on this project
    assigned_own_ids = set()
    assigned_hired_ids = set()
    for er in equip_reqs:
        for a in er.assignments:
            if a.machine_id:
                assigned_own_ids.add(a.machine_id)
            if a.hired_machine_id:
                assigned_hired_ids.add(a.hired_machine_id)
    for er in equip_reqs:
        total = len(er.assignments)
        equip_coverage.append({
            'req': er,
            'label': er.label,
            'required': er.required_count,
            'assignments': er.assignments,
            'total': total,
            'gap': max(0, er.required_count - total),
        })
    # All own fleet machines available to assign (active, not yet assigned to any requirement on this project)
    assignable_own = [m for m in Machine.query.filter_by(active=True).order_by(Machine.name).all()
                      if m.id not in assigned_own_ids]
    # All hired machines for this project not yet assigned to a requirement
    assignable_hired = [hm for hm in hired_machines_list if hm.id not in assigned_hired_ids]
    machine_groups = MachineGroup.query.order_by(MachineGroup.name).all()

    return render_template('project_dashboard.html',
                           project=project, progress=progress,
                           gantt_data=gantt_data,
                           delay_summary=delay_summary,
                           productivity=productivity,
                           non_work_dates=non_work_dates,
                           has_openpyxl=has_openpyxl,
                           budgeted_roles=budgeted_roles,
                           all_roles=all_roles,
                           actual_role_avg=actual_role_avg,
                           hired_machines_list=hired_machines_list,
                           owned_machines_list=owned_machines_list,
                           all_machines=all_machines,
                           assigned_machine_ids=assigned_machine_ids,
                           worked_sundays_list=worked_sundays_list,
                           project_documents=project_documents,
                           cost_estimate=cost_estimate,
                           role_groups_all=role_groups_all,
                           budgeted_by_group=budgeted_by_group,
                           equip_reqs=equip_reqs,
                           equip_coverage=equip_coverage,
                           assignable_own=assignable_own,
                           assignable_hired=assignable_hired,
                           machine_groups=machine_groups,
                           state_holidays=state_holidays,
                           today=date.today(),
                           setting_airports=get_airports(),
                           setting_locations=get_locations())


@projects_bp.route('/project/<int:project_id>/non-work-dates/add', methods=['POST'])
@require_role('admin')
def non_work_date_add(project_id):
    project = Project.query.get_or_404(project_id)
    date_str = request.form.get('date', '').strip()
    reason = request.form.get('reason', '').strip()
    if not date_str:
        flash('Date is required.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    try:
        nw_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))

    existing = ProjectNonWorkDate.query.filter_by(project_id=project_id, date=nw_date).first()
    if existing:
        flash(f'{nw_date.strftime("%d/%m/%Y")} already added.', 'warning')
    else:
        db.session.add(ProjectNonWorkDate(project_id=project_id, date=nw_date, reason=reason or None))
        db.session.commit()
        flash(f'Non-work date {nw_date.strftime("%d/%m/%Y")} added.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/non-work-dates/<int:nwd_id>/delete', methods=['POST'])
@require_role('admin')
def non_work_date_delete(project_id, nwd_id):
    nwd = ProjectNonWorkDate.query.get_or_404(nwd_id)
    db.session.delete(nwd)
    db.session.commit()
    flash('Non-work date removed.', 'info')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/budgeted-crew/add', methods=['POST'])
@require_role('admin')
def budgeted_crew_add(project_id):
    Project.query.get_or_404(project_id)
    role_name = request.form.get('role_name', '').strip()
    count_raw = request.form.get('budgeted_count', '').strip()
    if not role_name or not count_raw:
        flash('Role name and count are required.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    try:
        count = int(count_raw)
        if count < 1:
            raise ValueError
    except ValueError:
        flash('Count must be a positive integer.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    existing = ProjectBudgetedRole.query.filter_by(project_id=project_id, role_name=role_name).first()
    if existing:
        existing.budgeted_count = count
        flash(f'Updated {role_name} budget to {count}.', 'info')
    else:
        db.session.add(ProjectBudgetedRole(project_id=project_id, role_name=role_name, budgeted_count=count))
        flash(f'Added {count}x {role_name} to budget.', 'success')
    db.session.commit()
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/budgeted-crew/<int:br_id>/delete', methods=['POST'])
@require_role('admin')
def budgeted_crew_delete(project_id, br_id):
    br = ProjectBudgetedRole.query.get_or_404(br_id)
    db.session.delete(br)
    db.session.commit()
    flash('Budgeted role removed.', 'info')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/planned-upload', methods=['GET', 'POST'])
@require_role('admin')
def planned_upload(project_id):
    project = Project.query.get_or_404(project_id)

    if request.method == 'POST':
        if not HAS_OPENPYXL:
            flash('openpyxl is not installed. Run: pip install openpyxl', 'danger')
            return redirect(url_for('projects.project_dashboard', project_id=project_id))

        file = request.files.get('planned_file')
        if not file or not file.filename:
            flash('No file selected.', 'danger')
            return redirect(url_for('projects.project_dashboard', project_id=project_id))

        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            flash('Please upload an Excel (.xlsx/.xls) or CSV file.', 'danger')
            return redirect(url_for('projects.project_dashboard', project_id=project_id))

        replace_existing = request.form.get('replace_existing') == 'yes'

        try:
            if ext == 'csv':
                import csv, io
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})

            # Map columns (case-insensitive, flexible names)
            def find_col(row_dict, *names):
                for name in names:
                    for k in row_dict:
                        if k and k.strip().lower() == name.lower():
                            return k
                return None

            if not rows:
                flash('File appears empty.', 'danger')
                return redirect(url_for('projects.project_dashboard', project_id=project_id))

            sample = rows[0]
            col_day = find_col(sample, 'Day', 'day', 'DAY', 'Day Number')
            col_lot = find_col(sample, 'Lot', 'lot', 'LOT', 'Lot Number')
            col_loc = find_col(sample, 'Location', 'location', 'LOCATION', 'Loc')
            col_mat = find_col(sample, 'Material', 'material', 'MATERIAL')
            col_sqm = find_col(sample, 'Planned Sqm', 'Planned SQM', 'planned sqm', 'planned_sqm', 'SQM', 'sqm', 'Sqm')

            if not col_day or not col_sqm:
                flash('Could not find required columns (Day, Planned Sqm). Check column headers.', 'danger')
                return redirect(url_for('projects.project_dashboard', project_id=project_id))

            if replace_existing:
                PlannedData.query.filter_by(project_id=project_id).delete()

            count = 0
            skipped = 0
            for row in rows:
                try:
                    day_val = row.get(col_day)
                    sqm_val = row.get(col_sqm)
                    if day_val is None or sqm_val is None:
                        skipped += 1
                        continue
                    day_num = int(float(str(day_val)))
                    sqm = float(str(sqm_val))
                    lot = str(row.get(col_lot) or '').strip() or None
                    location = str(row.get(col_loc) or '').strip() or None if col_loc else None
                    material = str(row.get(col_mat) or '').strip() or None
                    db.session.add(PlannedData(
                        project_id=project_id,
                        lot=lot, location=location, material=material,
                        day_number=day_num, planned_sqm=sqm,
                    ))
                    count += 1
                except (ValueError, TypeError):
                    skipped += 1

            db.session.commit()
            flash(f'Imported {count} planned rows.{f" Skipped {skipped} invalid rows." if skipped else ""}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {e}', 'danger')

        return redirect(url_for('projects.project_dashboard', project_id=project_id))

    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/planned/clear', methods=['POST'])
@require_role('admin')
def planned_clear(project_id):
    Project.query.get_or_404(project_id)
    deleted = PlannedData.query.filter_by(project_id=project_id).delete()
    db.session.commit()
    flash(f'Cleared {deleted} planned rows.', 'info')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Own Equipment (ProjectMachine) — add / remove
# ---------------------------------------------------------------------------

@projects_bp.route('/project/<int:project_id>/own-equipment/add', methods=['POST'])
@require_role('admin')
def own_equipment_add(project_id):
    Project.query.get_or_404(project_id)
    machine_id_raw = request.form.get('machine_id', '').strip()
    if not machine_id_raw:
        flash('Please select a machine.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    machine_id = int(machine_id_raw)
    # A machine can only be on one project — move if already assigned elsewhere
    existing = ProjectMachine.query.filter_by(machine_id=machine_id).first()
    if existing:
        if existing.project_id == project_id:
            flash('That machine is already assigned to this project.', 'warning')
            return redirect(url_for('projects.project_dashboard', project_id=project_id))
        # Move from old project
        old_project = Project.query.get(existing.project_id)
        existing.project_id = project_id
        db.session.commit()
        machine = Machine.query.get(machine_id)
        flash(f'"{machine.name}" moved from {old_project.name if old_project else "another project"} to this project.', 'info')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    assigned_date_str = request.form.get('assigned_date', '').strip()
    assigned_date = None
    if assigned_date_str:
        try:
            assigned_date = datetime.strptime(assigned_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    notes = request.form.get('notes', '').strip() or None
    db.session.add(ProjectMachine(project_id=project_id, machine_id=machine_id,
                                   assigned_date=assigned_date, notes=notes))
    db.session.commit()
    machine = Machine.query.get(machine_id)
    flash(f'Own equipment "{machine.name}" assigned to project.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/equipment-bulk', methods=['POST'])
@require_role('admin', 'supervisor')
def equipment_bulk_assign(project_id):
    """Bulk update which machines are assigned to this project."""
    Project.query.get_or_404(project_id)
    selected_ids = set(int(x) for x in request.form.getlist('machine_ids') if x)

    # Remove machines no longer selected (only ones currently on THIS project)
    current = ProjectMachine.query.filter_by(project_id=project_id).all()
    for pm in current:
        if pm.machine_id not in selected_ids:
            db.session.delete(pm)

    # Add newly selected machines (move from other projects if needed)
    for mid in selected_ids:
        existing = ProjectMachine.query.filter_by(machine_id=mid).first()
        if existing:
            if existing.project_id != project_id:
                existing.project_id = project_id  # move to this project
        else:
            db.session.add(ProjectMachine(project_id=project_id, machine_id=mid))

    db.session.commit()
    flash(f'Equipment updated — {len(selected_ids)} item{"s" if len(selected_ids) != 1 else ""} on project.', 'success')
    return redirect(url_for('scheduling.scheduling_project', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/assign-group', methods=['POST'])
@require_role('admin', 'supervisor')
def assign_group_to_project(project_id):
    """Assign all machines in a group to the project (ProjectMachine entries).
    A machine can only be on one project at a time — existing assignments are moved."""
    group_id = request.form.get('group_id', type=int)
    if not group_id:
        flash('Select a group.', 'danger')
        return redirect(url_for('scheduling.scheduling_project', project_id=project_id))
    grp = MachineGroup.query.get_or_404(group_id)
    count = 0
    moved = 0
    for m in grp.machines:
        if not m.active:
            continue
        existing = ProjectMachine.query.filter_by(machine_id=m.id).first()
        if existing:
            if existing.project_id == project_id:
                continue  # already on this project
            # Move from old project to this one
            existing.project_id = project_id
            moved += 1
        else:
            db.session.add(ProjectMachine(project_id=project_id, machine_id=m.id))
        count += 1
    db.session.commit()
    msg = f'Group "{grp.name}" assigned — {count} item{"s" if count != 1 else ""} on project.'
    if moved:
        msg += f' ({moved} moved from other project{"s" if moved != 1 else ""})'
    flash(msg, 'success')
    return redirect(url_for('scheduling.scheduling_project', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/own-equipment/<int:pm_id>/remove', methods=['POST'])
@require_role('admin')
def own_equipment_remove(project_id, pm_id):
    pm = ProjectMachine.query.get_or_404(pm_id)
    db.session.delete(pm)
    db.session.commit()
    flash('Own equipment removed from project.', 'info')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Sunday Work Exceptions — add / remove
# ---------------------------------------------------------------------------

@projects_bp.route('/project/<int:project_id>/worked-sunday/add', methods=['POST'])
@require_role('admin')
def worked_sunday_add(project_id):
    Project.query.get_or_404(project_id)
    date_str = request.form.get('date', '').strip()
    reason = request.form.get('reason', '').strip()
    if not date_str:
        flash('Date is required.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    try:
        ws_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    if ws_date.weekday() != 6:
        flash('That date is not a Sunday.', 'warning')
        return redirect(url_for('projects.project_dashboard', project_id=project_id))
    existing = ProjectWorkedSunday.query.filter_by(project_id=project_id, date=ws_date).first()
    if existing:
        flash(f'{ws_date.strftime("%d/%m/%Y")} already added.', 'warning')
    else:
        db.session.add(ProjectWorkedSunday(project_id=project_id, date=ws_date,
                                           reason=reason or None))
        db.session.commit()
        flash(f'Worked Sunday {ws_date.strftime("%d/%m/%Y")} added.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


@projects_bp.route('/project/<int:project_id>/worked-sunday/<int:ws_id>/delete', methods=['POST'])
@require_role('admin')
def worked_sunday_delete(project_id, ws_id):
    ws = ProjectWorkedSunday.query.get_or_404(ws_id)
    db.session.delete(ws)
    db.session.commit()
    flash('Worked Sunday removed.', 'info')
    return redirect(url_for('projects.project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Project Progress PDF Report
# ---------------------------------------------------------------------------

@projects_bp.route('/project/<int:project_id>/report/pdf')
@require_role('admin', 'supervisor')
def project_report_pdf(project_id):
    project = Project.query.get_or_404(project_id)

    # Date range for filtering entries
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except ValueError:
        date_from = None
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        date_to = None

    progress = compute_project_progress(project_id)
    delay_summary = compute_delay_summary(project_id)
    gantt_data = compute_gantt_data(project_id)
    settings = load_settings()

    progress_dict = None
    if progress:
        progress_dict = {
            'tasks': progress['tasks'],
            'total_planned': progress['total_planned'],
            'total_actual': progress['total_actual'],
            'total_remaining': progress['total_remaining'],
            'overall_pct': progress['overall_pct'],
            'install_rate': progress.get('install_rate'),
        }

    pdf_bytes = generate_project_report_pdf(
        project, progress_dict, delay_summary, {}, settings,
        date_from=date_from, date_to=date_to, gantt_data=gantt_data
    )
    safe_name = project.name.replace(' ', '_').replace('/', '-')
    from_str = date_from.strftime('%Y%m%d') if date_from else 'all'
    to_str = date_to.strftime('%Y%m%d') if date_to else date.today().strftime('%Y%m%d')
    filename = f"Progress_Report_{safe_name}_{from_str}_to_{to_str}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Client Delay & Variation Report PDF
# ---------------------------------------------------------------------------

@projects_bp.route('/project/<int:project_id>/delay-report/pdf')
@require_role('admin', 'supervisor')
def project_delay_report_pdf(project_id):
    from utils.reports import generate_client_delay_report_pdf
    project = Project.query.get_or_404(project_id)
    settings = load_settings()
    pdf_bytes = generate_client_delay_report_pdf(project, settings)
    safe_name = project.name.replace(' ', '_').replace('/', '-')
    filename = f"Delay_Variation_Report_{safe_name}_{date.today().strftime('%Y%m%d')}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Weekly Progress Report PDF (client distribution)
# ---------------------------------------------------------------------------

@projects_bp.route('/project/<int:project_id>/weekly-report/pdf')
@require_role('admin', 'supervisor')
def project_weekly_report_pdf(project_id):
    project = Project.query.get_or_404(project_id)
    week_start_str = request.args.get('week_start', '').strip()
    today = date.today()
    if week_start_str:
        try:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        except ValueError:
            week_start = today - timedelta(days=today.weekday())
    else:
        week_start = today - timedelta(days=today.weekday())  # Monday of current week
    week_end = week_start + timedelta(days=6)

    entries = (DailyEntry.query
               .filter_by(project_id=project_id)
               .filter(DailyEntry.entry_date >= week_start)
               .filter(DailyEntry.entry_date <= week_end)
               .order_by(DailyEntry.entry_date)
               .all())

    settings = load_settings()
    pdf_bytes = generate_weekly_report_pdf(project, week_start, week_end, entries, settings)
    safe_name = project.name.replace(' ', '_').replace('/', '-')
    filename = f"Weekly_Report_{safe_name}_{week_start.strftime('%Y%m%d')}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


@projects_bp.route('/project/<int:project_id>/budgeted-crew/save-all', methods=['POST'])
@require_role('admin')
def budgeted_crew_save_all(project_id):
    project = Project.query.get_or_404(project_id)
    # Delete all existing budgeted roles for this project
    ProjectBudgetedRole.query.filter_by(project_id=project_id).delete()
    # Re-add from form (group_name -> count)
    for key, val in request.form.items():
        if key.startswith('group_'):
            group_name = key[6:]  # strip 'group_' prefix
            try:
                count = int(val)
            except (ValueError, TypeError):
                count = 0
            if count > 0:
                db.session.add(ProjectBudgetedRole(
                    project_id=project_id,
                    role_name=group_name,
                    budgeted_count=count
                ))
    db.session.commit()
    flash('Crew requirements saved.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')


@projects_bp.route('/project/<int:project_id>/equipment-requirements/add', methods=['POST'])
@require_role('admin')
def equipment_req_add(project_id):
    Project.query.get_or_404(project_id)
    label = request.form.get('label', '').strip()
    try:
        required_count = max(1, int(request.form.get('required_count', 1)))
    except (ValueError, TypeError):
        required_count = 1
    if not label:
        flash('Equipment name is required.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')
    db.session.add(ProjectEquipmentRequirement(
        project_id=project_id, label=label, required_count=required_count))
    db.session.commit()
    flash(f'Added requirement: {label}.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')


@projects_bp.route('/project/<int:project_id>/equipment-requirements/<int:req_id>/update', methods=['POST'])
@require_role('admin')
def equipment_req_update(project_id, req_id):
    req = ProjectEquipmentRequirement.query.filter_by(id=req_id, project_id=project_id).first_or_404()
    label = request.form.get('label', '').strip()
    try:
        required_count = max(1, int(request.form.get('required_count', 1)))
    except (ValueError, TypeError):
        required_count = 1
    if label:
        req.label = label
    req.required_count = required_count
    db.session.commit()
    flash('Requirement updated.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')


@projects_bp.route('/project/<int:project_id>/equipment-requirements/<int:req_id>/delete', methods=['POST'])
@require_role('admin')
def equipment_req_delete(project_id, req_id):
    req = ProjectEquipmentRequirement.query.filter_by(id=req_id, project_id=project_id).first_or_404()
    # cascade delete assignments
    ProjectEquipmentAssignment.query.filter_by(requirement_id=req_id).delete()
    db.session.delete(req)
    db.session.commit()
    flash('Requirement removed.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')


@projects_bp.route('/project/<int:project_id>/equipment-requirements/<int:req_id>/assign', methods=['POST'])
@require_role('admin')
def equipment_req_assign(project_id, req_id):
    req = ProjectEquipmentRequirement.query.filter_by(id=req_id, project_id=project_id).first_or_404()
    machine_id = request.form.get('machine_id', '').strip()
    hired_machine_id = request.form.get('hired_machine_id', '').strip()
    group_id = request.form.get('group_id', '').strip()

    # Assign entire group — add all active machines from the group
    if group_id:
        grp = MachineGroup.query.get(int(group_id))
        if grp:
            count = 0
            for m in grp.machines:
                if not m.active:
                    continue
                # Skip if already assigned to this requirement
                existing = ProjectEquipmentAssignment.query.filter_by(
                    requirement_id=req.id, machine_id=m.id).first()
                if not existing:
                    db.session.add(ProjectEquipmentAssignment(
                        requirement_id=req.id, machine_id=m.id))
                    count += 1
            # Also add hired machines in the group
            for hm in grp.hired_machines:
                if not hm.active:
                    continue
                existing = ProjectEquipmentAssignment.query.filter_by(
                    requirement_id=req.id, hired_machine_id=hm.id).first()
                if not existing:
                    db.session.add(ProjectEquipmentAssignment(
                        requirement_id=req.id, hired_machine_id=hm.id))
                    count += 1
            db.session.commit()
            flash(f'Group "{grp.name}" assigned — {count} item{"s" if count != 1 else ""} added.', 'success')
        return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')

    if not machine_id and not hired_machine_id:
        flash('Select a machine or group to assign.', 'danger')
        return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')
    assignment = ProjectEquipmentAssignment(requirement_id=req.id)
    if machine_id:
        assignment.machine_id = int(machine_id)
    else:
        assignment.hired_machine_id = int(hired_machine_id)
    db.session.add(assignment)
    db.session.commit()
    flash('Machine assigned.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')


@projects_bp.route('/project/<int:project_id>/equipment-assignments/<int:assign_id>/remove', methods=['POST'])
@require_role('admin')
def equipment_req_unassign(project_id, assign_id):
    a = ProjectEquipmentAssignment.query.filter_by(id=assign_id).first_or_404()
    # verify it belongs to this project
    if a.requirement.project_id != project_id:
        return 'Forbidden', 403
    db.session.delete(a)
    db.session.commit()
    flash('Machine removed from requirement.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-scheduling')


@projects_bp.route('/project/<int:project_id>/settings/save', methods=['POST'])
@require_role('admin')
def project_settings_save(project_id):
    project = Project.query.get_or_404(project_id)
    project.name = request.form.get('name', '').strip() or project.name
    project.description = request.form.get('description', '').strip() or None
    start_date_str = request.form.get('start_date', '').strip()
    project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
    planned_crew = request.form.get('planned_crew', '').strip()
    project.planned_crew = int(planned_crew) if planned_crew else None
    hours_per_day = request.form.get('hours_per_day', '').strip()
    project.hours_per_day = float(hours_per_day) if hours_per_day else None
    quoted_days = request.form.get('quoted_days', '').strip()
    project.quoted_days = int(quoted_days) if quoted_days else None
    project.state = request.form.get('state', '').strip() or None
    project.is_cfmeu = bool(request.form.get('is_cfmeu'))
    project.track_by_lot = bool(request.form.get('track_by_lot'))
    project.city = request.form.get('city', '').strip() or None
    project.nearest_airport = request.form.get('nearest_airport', '').strip().upper() or None
    project.site_address = request.form.get('site_address', '').strip() or project.site_address
    end_date_str = request.form.get('planned_end_date', '').strip()
    project.planned_end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    db.session.commit()
    flash('Project settings saved.', 'success')
    return redirect(url_for('projects.project_dashboard', project_id=project_id) + '#tab-settings')
