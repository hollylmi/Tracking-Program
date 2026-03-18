from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_from_directory, send_file, Response, jsonify)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import (db, User, Project, Employee, Machine, DailyEntry, HiredMachine,
                    StandDown, Role, EntryPhoto, PlannedData, ProjectNonWorkDate,
                    ProjectBudgetedRole, ProjectMachine, ProjectWorkedSunday,
                    ProjectDocument, SwingPattern, EmployeeSwing, EmployeeLeave,
                    ProjectAssignment, ProjectEquipmentRequirement, ProjectEquipmentAssignment,
                    MachineBreakdown, BreakdownPhoto, PublicHoliday, CFMEUDate, AUSTRALIAN_STATES,
                    ScheduleDayOverride, DiagramLayer, PanelInstallRecord)
from datetime import date, datetime, timedelta
from fpdf import FPDF, XPos, YPos
import os, json, uuid, smtplib, math, re
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from werkzeug.utils import secure_filename
import storage
try:
    from dxf_to_svg import dxf_to_svg as _dxf_to_svg
    _DXF_AVAILABLE = True
except ImportError:
    _DXF_AVAILABLE = False

try:
    import fitz as _fitz          # PyMuPDF — PDF → image
    _PYMUPDF_AVAILABLE = True
except ImportError:
    _PYMUPDF_AVAILABLE = False

try:
    import cv2 as _cv2
    import numpy as _np
    _CV2_AVAILABLE = True
except ImportError:
    _CV2_AVAILABLE = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Per-project colour palette (bg, text) — cycled by project.id order
SCHED_PALETTE = [
    ('#cfe2ff', '#084298'),  # blue
    ('#d1e7dd', '#0a3622'),  # green
    ('#f8d7da', '#842029'),  # red
    ('#fff3cd', '#664d03'),  # yellow
    ('#d2f4ea', '#0b4c34'),  # teal
    ('#fde8d8', '#6c3a00'),  # orange
    ('#e2d9f3', '#3d1a78'),  # purple
    ('#dee2e6', '#343a40'),  # grey
]

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///tracking.db')
# Railway provides postgres:// — SQLAlchemy requires postgresql://
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32 MB upload limit

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx', 'dwg', 'dxf'}
DOC_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'dwg', 'dxf', 'doc', 'docx', 'xls', 'xlsx'}
PHOTO_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'instance', 'uploads')
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), 'instance', 'settings.json')

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'warning'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


with app.app_context():
    db.create_all()
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    # Fix entries with delay_reason set but delay_hours = 0
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            conn.execute(text("UPDATE daily_entry SET delay_reason = NULL WHERE delay_hours = 0 OR delay_hours IS NULL"))
            conn.commit()
    except Exception:
        pass
    # Add group_name column to role table if it doesn't exist yet
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE role ADD COLUMN group_name VARCHAR(100)"))
            conn.commit()
    except Exception:
        pass  # Column already exists
    for stmt in [
        "ALTER TABLE project ADD COLUMN state VARCHAR(10)",
        "ALTER TABLE project ADD COLUMN is_cfmeu BOOLEAN DEFAULT 0",
        "ALTER TABLE swing_pattern RENAME COLUMN work_days TO work_weeks",
        "ALTER TABLE employee_swing ADD COLUMN day_offset INTEGER DEFAULT 0",
        "ALTER TABLE machine ADD COLUMN plant_id VARCHAR(100)",
        "ALTER TABLE machine ADD COLUMN description TEXT",
        "ALTER TABLE hired_machine ADD COLUMN plant_id VARCHAR(100)",
        "ALTER TABLE hired_machine ADD COLUMN description TEXT",
        "ALTER TABLE user ADD COLUMN email VARCHAR(200)",
        "ALTER TABLE daily_entry ADD COLUMN weather VARCHAR(200)",
        "ALTER TABLE diagram_layer ADD COLUMN canvas_bg_filename VARCHAR(500)",
        "ALTER TABLE diagram_layer ADD COLUMN canvas_bg_original_name VARCHAR(500)",
        "ALTER TABLE panel_install_record ADD COLUMN source VARCHAR(20)",
    ]:
        try:
            db.session.execute(db.text(stmt))
            db.session.commit()
        except Exception:
            db.session.rollback()
    # Seed default admin user if none exist (guard against race condition with multiple workers)
    try:
        if User.query.count() == 0:
            from werkzeug.security import generate_password_hash
            admin = User(
                username='admin',
                display_name='Administrator',
                password_hash=generate_password_hash('admin123'),
                is_admin=True,
                active=True,
            )
            db.session.add(admin)
            db.session.commit()
            print("Created default admin user: admin / admin123 — please change the password!")
    except Exception:
        db.session.rollback()


@app.context_processor
def inject_active_projects():
    """Make active projects available in all templates for the Progress nav dropdown."""
    try:
        projects = Project.query.filter_by(active=True).order_by(Project.name).all()
    except Exception:
        projects = []
    return {'_active_projects': projects}


@app.before_request
def require_login():
    """Redirect unauthenticated users to login for all routes except login/static."""
    public_endpoints = {'login', 'logout', 'static'}
    if request.endpoint not in public_endpoints and not current_user.is_authenticated:
        return redirect(url_for('login', next=request.url))


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        from werkzeug.security import check_password_hash
        if user and user.active and check_password_hash(user.password_hash, password):
            login_user(user, remember=request.form.get('remember') == 'on')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_photo(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in PHOTO_EXTENSIONS


def allowed_doc(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in DOC_EXTENSIONS


def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, 'r') as f:
            return json.load(f)
    return {
        'company_name': '',
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,
        'smtp_username': '',
        'smtp_password': '',
        'from_name': '',
        'from_email': '',
    }


def save_settings(data):
    os.makedirs(os.path.dirname(SETTINGS_FILE), exist_ok=True)
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(data, f, indent=2)


def safe(s):
    """Sanitize text for fpdf2 core fonts (Latin-1 only)."""
    if s is None:
        return ''
    s = str(s)
    for old, new in {
        '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"',
        '\u2013': '-', '\u2014': '--',
        '\u2026': '...',
        '\u00a0': ' ',
        '\u2022': '*',
    }.items():
        s = s.replace(old, new)
    return s.encode('latin-1', errors='replace').decode('latin-1')


def build_day_summary(hm, date_from, date_to):
    """Return a list of dicts for each day in range, plus summary counts."""
    sd_map = {sd.stand_down_date: sd.reason for sd in hm.stand_downs}
    count_saturdays = hm.count_saturdays if hm.count_saturdays is not None else True
    days = []
    current = date_from
    while current <= date_to:
        weekday = current.weekday()
        is_sunday = weekday == 6
        is_saturday = weekday == 5

        if is_sunday or (is_saturday and not count_saturdays):
            status = 'non_working'
        elif hm.delivery_date and current < hm.delivery_date:
            status = 'not_delivered'
        elif hm.return_date and current > hm.return_date:
            status = 'returned'
        elif current in sd_map:
            status = 'stood_down'
        else:
            status = 'on_site'

        days.append({
            'date': current,
            'day_name': current.strftime('%A'),
            'status': status,
            'reason': sd_map.get(current, ''),
        })
        current += timedelta(days=1)

    on_site    = sum(1 for d in days if d['status'] == 'on_site')
    stood_down = sum(1 for d in days if d['status'] == 'stood_down')
    working_days = on_site + stood_down
    days_pw = 6 if count_saturdays else 5
    cost_per_day_derived = hm.cost_per_week / days_pw if hm.cost_per_week else None
    summary = {
        'on_site': on_site,
        'stood_down': stood_down,
        'working_days': working_days,
        'total_days': len(days),
        'cost_day': round(on_site * cost_per_day_derived, 2) if cost_per_day_derived else None,
        'cost_per_day_derived': round(cost_per_day_derived, 2) if cost_per_day_derived else None,
        'cost_week': hm.cost_per_week,
        'count_saturdays': count_saturdays,
    }
    return days, summary


def _natural_key(s):
    """Sort key for natural ordering: LOT 1, LOT 2, LOT 10 (not LOT 1, LOT 10, LOT 2)."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s or '')]


def compute_project_progress(project_id):
    """Return planned vs actual progress data for a project."""
    planned = PlannedData.query.filter_by(project_id=project_id).all()
    if not planned:
        return None

    entries = (DailyEntry.query
               .filter_by(project_id=project_id)
               .filter(DailyEntry.install_sqm > 0)
               .all())

    # Aggregate planned by (lot, material)
    planned_by_task = {}
    for p in planned:
        key = (p.lot or '', p.material or '')
        if key not in planned_by_task:
            planned_by_task[key] = {
                'lot': p.lot, 'location': p.location, 'material': p.material,
                'planned_sqm': 0, 'min_day': float('inf'), 'max_day': 0,
            }
        planned_by_task[key]['planned_sqm'] += p.planned_sqm or 0
        planned_by_task[key]['min_day'] = min(planned_by_task[key]['min_day'], p.day_number or 0)
        planned_by_task[key]['max_day'] = max(planned_by_task[key]['max_day'], p.day_number or 0)

    # Aggregate actuals by (lot_number, material)
    actual_by_task = {}
    total_install_hours = 0.0
    for e in entries:
        key = (e.lot_number or '', e.material or '')
        actual_by_task[key] = actual_by_task.get(key, 0.0) + (e.install_sqm or 0)
        total_install_hours += (e.install_hours or 0)

    tasks = []
    total_planned = 0.0
    total_actual = 0.0
    # Natural sort by lot then material
    sorted_items = sorted(planned_by_task.items(),
                          key=lambda x: (_natural_key(x[1]['lot']), _natural_key(x[1]['material'])))
    for key, data in sorted_items:
        actual = actual_by_task.get(key, 0.0)
        planned_sqm = data['planned_sqm']
        pct = round(actual / planned_sqm * 100, 1) if planned_sqm > 0 else 0
        total_planned += planned_sqm
        total_actual += actual
        tasks.append({
            'lot': data['lot'],
            'location': data['location'],
            'material': data['material'],
            'planned_sqm': round(planned_sqm, 2),
            'actual_sqm': round(actual, 2),
            'remaining': round(max(0, planned_sqm - actual), 2),
            'pct_complete': min(pct, 100),
            'min_day': data['min_day'],
            'max_day': data['max_day'],
        })

    overall_pct = round(total_actual / total_planned * 100, 1) if total_planned > 0 else 0
    install_rate = round(total_actual / total_install_hours, 2) if total_install_hours > 0 else None

    # Latest crew count
    all_entries = DailyEntry.query.filter_by(project_id=project_id).order_by(DailyEntry.entry_date.desc()).first()
    current_crew = all_entries.num_people if all_entries and all_entries.num_people else None

    project = Project.query.get(project_id)
    return {
        'tasks': tasks,
        'total_planned': round(total_planned, 2),
        'total_actual': round(total_actual, 2),
        'total_remaining': round(max(0, total_planned - total_actual), 2),
        'overall_pct': min(overall_pct, 100),
        'install_rate': install_rate,
        'planned_crew': project.planned_crew,
        'current_crew': current_crew,
    }


def compute_gantt_data(project_id):
    """Build day-by-day Gantt data matching the matplotlib approach.

    Each planned/actual/forecast entry produces individual 1-day-wide bars
    (not spans), shading is applied per-day for Sundays and non-work dates,
    and per-task variance labels are computed.
    """
    project = Project.query.get(project_id)
    if not project or not project.start_date:
        return None

    planned = PlannedData.query.filter_by(project_id=project_id).order_by(PlannedData.day_number).all()
    if not planned:
        return None

    entries = (DailyEntry.query
               .filter_by(project_id=project_id)
               .filter(DailyEntry.install_sqm > 0)
               .order_by(DailyEntry.entry_date)
               .all())

    non_work_set = {nwd.date for nwd in project.non_work_dates}
    worked_sundays_set = {ws.date for ws in ProjectWorkedSunday.query.filter_by(project_id=project_id).all()}

    # Map day number → calendar date (skip Sundays unless worked, and non-work dates)
    max_day = max((p.day_number or 0) for p in planned)
    day_to_date = {}
    current = project.start_date
    for d in range(1, max_day + 1):
        while (current.weekday() == 6 and current not in worked_sundays_set) or current in non_work_set:
            current += timedelta(days=1)
        day_to_date[d] = current
        current += timedelta(days=1)

    # Task info: planned_dates is a set of calendar dates (one per planned row)
    task_info = {}
    task_order_keys = []
    for p in planned:
        key = (p.lot or '', p.material or '')
        if key not in task_info:
            label = (f"{p.lot} — {p.material}" if p.lot and p.material
                     else (p.lot or p.material or 'Unknown'))
            task_info[key] = {
                'label': label,
                'planned_dates': set(),
                'planned_sqm': 0,
            }
            task_order_keys.append(key)
        d_date = day_to_date.get(p.day_number)
        if d_date:
            task_info[key]['planned_dates'].add(d_date)
        task_info[key]['planned_sqm'] += p.planned_sqm or 0

    task_order_keys.sort(key=lambda k: (_natural_key(k[0]), _natural_key(k[1])))

    # Actuals by (lot_number, material) — deduplicate dates per task
    actuals_by_task = {}
    for e in entries:
        key = (e.lot_number or '', e.material or '')
        if key not in actuals_by_task:
            actuals_by_task[key] = {'sqm': 0.0, 'hrs': 0.0, 'dates': set()}
        actuals_by_task[key]['sqm'] += e.install_sqm or 0
        actuals_by_task[key]['hrs'] += e.install_hours or 0
        actuals_by_task[key]['dates'].add(e.entry_date)

    # Install rates per task + global fallback
    task_rates = {}
    for key, data in actuals_by_task.items():
        if data['hrs'] > 0 and data['sqm'] > 0:
            task_rates[key] = data['sqm'] / data['hrs']
    total_sqm = sum(d['sqm'] for d in actuals_by_task.values())
    total_hrs = sum(d['hrs'] for d in actuals_by_task.values())
    global_rate = (total_sqm / total_hrs) if total_hrs > 0 else 87.5

    # Current crew for forecast
    latest_entry = (DailyEntry.query.filter_by(project_id=project_id)
                    .order_by(DailyEntry.entry_date.desc()).first())
    hours_per_day = project.hours_per_day or 8
    curr_crew = (latest_entry.num_people if latest_entry and latest_entry.num_people
                 else project.planned_crew or 10)
    daily_cap_hrs = curr_crew * hours_per_day

    today = date.today()
    target_finish = max(day_to_date.values()) if day_to_date else None

    # Standdown dates — split by cause for Gantt shading, combined for forecast skip
    _delay_entries = (DailyEntry.query.filter_by(project_id=project_id)
                      .filter(DailyEntry.delay_hours > 0).all())
    delay_entry_dates = {e.entry_date for e in _delay_entries}
    weather_delay_dates = {
        e.entry_date for e in _delay_entries
        if e.delay_reason and 'weather' in e.delay_reason.lower()
    }
    client_delay_dates = delay_entry_dates - weather_delay_dates

    def next_work_day(d):
        # Advance d past Sundays (unless worked), non-work dates, and standdown dates
        while (d.weekday() == 6 and d not in worked_sundays_set) or d in non_work_set or d in delay_entry_dates:
            d += timedelta(days=1)
        return d

    # Sort tasks: by lot (natural), then by planned start date within each lot
    task_order_keys.sort(key=lambda k: (
        _natural_key(k[0]),
        min(task_info[k]['planned_dates']) if task_info[k]['planned_dates'] else date.max,
    ))

    # Compute forecast — all tasks are globally sequential
    # (sorted by lot then planned start, each task starts after the previous finishes)
    forecast_by_task = {}
    est_finish_by_task = {}

    # Initialize global_last_date to the latest actual work done on ANY task.
    # This ensures out-of-order actual work pushes all subsequent forecasts forward.
    max_global_actual = None
    for act_data in actuals_by_task.values():
        if act_data['dates']:
            t_max = max(act_data['dates'])
            if max_global_actual is None or t_max > max_global_actual:
                max_global_actual = t_max
    global_last_date = max_global_actual  # replaces "global_last_date = None"

    for key in task_order_keys:
        info = task_info[key]
        act = actuals_by_task.get(key)
        planned_sqm = info['planned_sqm']
        actual_sqm = act['sqm'] if act else 0.0
        remaining = planned_sqm - actual_sqm

        t_last_act = max(act['dates']) if act and act['dates'] else None

        # Base forecast start
        if t_last_act:
            candidate = max(t_last_act + timedelta(days=1), today)
        else:
            planned_start = min(info['planned_dates']) if info['planned_dates'] else today
            candidate = max(planned_start, today)

        # Don't start before the previous task has finished
        if global_last_date is not None:
            candidate = max(candidate, global_last_date + timedelta(days=1))

        start_f = next_work_day(candidate)

        if remaining > planned_sqm * 0.005 and daily_cap_hrs > 0:
            actual_work_days = len(act['dates']) if act and act['dates'] else 0
            if actual_work_days > 0 and act and act['sqm'] > 0:
                # Use effective daily sqm — naturally accounts for partial standdowns
                effective_daily_sqm = act['sqm'] / actual_work_days
                days_req = max(1, math.ceil(remaining / effective_daily_sqm - 0.001))
            else:
                # No actuals yet — use planned number of days
                days_req = max(1, len(info['planned_dates']))
            forecast_dates = []
            temp = start_f
            for _ in range(days_req):
                temp = next_work_day(temp)
                forecast_dates.append(temp)
                temp += timedelta(days=1)
            forecast_by_task[key] = forecast_dates
            est_finish = forecast_dates[-1] if forecast_dates else start_f
            est_finish_by_task[key] = est_finish
            global_last_date = est_finish
        else:
            forecast_by_task[key] = []
            planned_finish = max(info['planned_dates']) if info['planned_dates'] else None
            est_end = t_last_act if t_last_act else planned_finish
            est_finish_by_task[key] = est_end
            if est_end:
                if global_last_date is None or est_end > global_last_date:
                    global_last_date = est_end

    est_finish_overall = max((v for v in est_finish_by_task.values() if v), default=target_finish)

    # Gantt date range
    gantt_start = project.start_date
    candidates = [d for d in [target_finish, est_finish_overall, today] if d]
    gantt_end = max(candidates) + timedelta(days=7) if candidates else gantt_start + timedelta(days=30)
    total_span = max((gantt_end - gantt_start).days, 1)
    day_width_pct = round(100.0 / total_span, 4)

    def pct(d):
        if d is None:
            return None
        return round((d - gantt_start).days / total_span * 100, 3)

    # Shade stripes: Sunday=pink (unless worked Sunday), non-work=blue, standdown=orange
    shade_stripes = []
    d = gantt_start
    while d <= gantt_end:
        if d.weekday() == 6 and d not in worked_sundays_set:
            shade_stripes.append({'left': pct(d), 'w': day_width_pct, 'type': 'sun'})
        elif d in non_work_set:
            shade_stripes.append({'left': pct(d), 'w': day_width_pct, 'type': 'nwd'})
        if d in weather_delay_dates:
            shade_stripes.append({'left': pct(d), 'w': day_width_pct, 'type': 'wwd'})
        if d in client_delay_dates:
            shade_stripes.append({'left': pct(d), 'w': day_width_pct, 'type': 'cld'})
        d += timedelta(days=1)

    # Week markers (Mondays)
    week_markers = []
    wk = gantt_start - timedelta(days=gantt_start.weekday())
    while wk <= gantt_end:
        p = pct(wk)
        if p is not None and -1 <= p <= 101:
            week_markers.append({'date': wk.strftime('%d/%m'), 'left_pct': p})
        wk += timedelta(days=7)

    # Month markers
    month_markers = []
    mo = gantt_start.replace(day=1)
    while mo <= gantt_end:
        p = pct(mo)
        if p is not None and 0 <= p <= 100:
            month_markers.append({'label': mo.strftime('%b %Y'), 'left_pct': p})
        mo = mo.replace(month=mo.month + 1) if mo.month < 12 else mo.replace(year=mo.year + 1, month=1)

    # Build rows: individual day pct positions
    rows = []
    for key in task_order_keys:
        info = task_info[key]
        act = actuals_by_task.get(key)
        planned_finish = max(info['planned_dates']) if info['planned_dates'] else None
        est_finish = est_finish_by_task.get(key)
        variance = (est_finish - planned_finish).days if planned_finish and est_finish else None

        rows.append({
            'label': info['label'],
            'planned_days': sorted([pct(d) for d in info['planned_dates']]),
            'actual_days': sorted([pct(d) for d in (act['dates'] if act else set())]),
            'forecast_days': [pct(d) for d in forecast_by_task.get(key, [])],
            'variance_days': variance,
        })

    variance_overall = (est_finish_overall - target_finish).days if target_finish and est_finish_overall else None

    return {
        'rows': rows,
        'shade_stripes': shade_stripes,
        'day_width_pct': day_width_pct,
        'today_pct': pct(today),
        'target_finish_pct': pct(target_finish),
        'target_finish': target_finish.strftime('%d/%m/%Y') if target_finish else None,
        'est_finish': est_finish_overall.strftime('%d/%m/%Y') if est_finish_overall else None,
        'variance_days': variance_overall,
        'week_markers': week_markers,
        'month_markers': month_markers,
    }


def compute_delay_summary(project_id):
    """Return a delay summary grouped by (reason, billable) for the project dashboard."""
    entries = (DailyEntry.query
               .filter_by(project_id=project_id)
               .filter(DailyEntry.delay_hours > 0)
               .all())
    if not entries:
        return []

    categories = {}
    for e in entries:
        reason = e.delay_reason or 'Unspecified'
        is_billable = e.delay_billable if e.delay_billable is not None else True
        key = (reason, is_billable)
        if key not in categories:
            categories[key] = {'reason': reason, 'billable': is_billable, 'hours': 0.0, 'events': 0}
        categories[key]['hours'] += e.delay_hours or 0
        categories[key]['events'] += 1

    result = sorted(categories.values(), key=lambda x: (-x['hours'], x['reason']))
    for cat in result:
        cat['hours'] = round(cat['hours'], 1)
    return result


def build_delay_report(project_id, date_from, date_to, billable_filter='all'):
    """Return per-entry delay breakdown grouped by role, with billable split."""
    query = (DailyEntry.query
             .filter(DailyEntry.delay_hours > 0)
             .filter(DailyEntry.entry_date >= date_from)
             .filter(DailyEntry.entry_date <= date_to)
             .order_by(DailyEntry.entry_date))
    if project_id:
        query = query.filter_by(project_id=int(project_id))
    if billable_filter == 'billable':
        query = query.filter(DailyEntry.delay_billable == True)
    elif billable_filter == 'non_billable':
        query = query.filter(DailyEntry.delay_billable == False)

    rows = []
    total_cost = 0.0
    total_hours_billable = 0.0
    total_hours_non_billable = 0.0

    for entry in query.all():
        # Group employees by (role_title, rate) — show role not individual name
        role_counts = {}
        for emp in entry.employees:
            if emp.delay_rate:
                label = emp.role if emp.role else emp.name
                rate = emp.delay_rate
                key = (label, rate)
                role_counts[key] = role_counts.get(key, 0) + 1

        emp_lines = []
        for (label, rate), count in sorted(role_counts.items()):
            cost = round(entry.delay_hours * rate * count, 2)
            people_str = f"{count} {'person' if count == 1 else 'people'}"
            emp_lines.append({
                'name': f"{label} ({people_str})",
                'rate': rate, 'count': count,
                'hours': entry.delay_hours, 'cost': cost,
            })

        machine_lines = []
        for m in entry.machines:
            if m.delay_rate:
                cost = round(entry.delay_hours * m.delay_rate, 2)
                machine_lines.append({'name': m.name, 'rate': m.delay_rate,
                                      'hours': entry.delay_hours, 'cost': cost})

        is_billable = entry.delay_billable if entry.delay_billable is not None else True
        entry_cost = sum(r['cost'] for r in emp_lines) + sum(r['cost'] for r in machine_lines)

        if is_billable:
            total_cost += entry_cost
            total_hours_billable += entry.delay_hours
        else:
            total_hours_non_billable += entry.delay_hours

        rows.append({
            'entry': entry,
            'emp_lines': emp_lines,
            'machine_lines': machine_lines,
            'entry_cost': entry_cost,
            'billable': is_billable,
        })

    billable_count = sum(1 for r in rows if r['billable'])
    summary = {
        'total_cost': round(total_cost, 2),
        'total_hours': total_hours_billable + total_hours_non_billable,
        'total_hours_billable': total_hours_billable,
        'total_hours_non_billable': total_hours_non_billable,
        'entry_count': len(rows),
        'billable_count': billable_count,
        'non_billable_count': len(rows) - billable_count,
    }
    return rows, summary


def generate_pdf(hm, date_from, date_to, days, summary, settings):
    """Build a stand-down report PDF and return bytes."""
    pdf = FPDF()
    pdf.set_margins(15, 15, 15)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    company = safe(settings.get('company_name', '') or 'Project Tracker')

    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 10, company, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', 'B', 13)
    pdf.cell(0, 8, 'MACHINE HIRE STAND-DOWN REPORT', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 5, f'Generated: {date.today().strftime("%d/%m/%Y")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(6)

    pdf.set_fill_color(240, 244, 255)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'MACHINE DETAILS', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
    pdf.set_font('Helvetica', '', 10)

    def detail_row(label, value):
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(45, 6, label + ':')
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, safe(value) if value else '-', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    detail_row('Machine', hm.machine_name)
    detail_row('Type', hm.machine_type)
    detail_row('Hire Company', hm.hire_company)
    detail_row('Company Email', hm.hire_company_email)
    detail_row('Delivery Date', hm.delivery_date.strftime('%d/%m/%Y') if hm.delivery_date else None)
    detail_row('Return Date', hm.return_date.strftime('%d/%m/%Y') if hm.return_date else None)
    detail_row('Project', hm.project.name)
    sat_billing = 'Yes (Saturdays billable)' if (hm.count_saturdays is not False) else 'No (Saturdays excluded)'
    detail_row('Saturday Billing', sat_billing)
    pdf.ln(4)

    pdf.set_fill_color(240, 244, 255)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'REPORT PERIOD', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, f'{date_from.strftime("%d/%m/%Y")}  to  {date_to.strftime("%d/%m/%Y")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    pdf.set_fill_color(240, 244, 255)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'DAILY SUMMARY', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)

    col_w = [28, 26, 30, 96]
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(220, 230, 255)
    for i, (header, w) in enumerate(zip(['Date', 'Day', 'Status', 'Notes / Reason'], col_w)):
        is_last = (i == 3)
        pdf.cell(w, 6, header, border=1, fill=True,
                 new_x=XPos.LMARGIN if is_last else XPos.RIGHT,
                 new_y=YPos.NEXT if is_last else YPos.TOP)

    status_labels = {
        'on_site': 'On Site', 'stood_down': 'Stood Down',
        'not_delivered': 'Not Yet Delivered', 'returned': 'Returned',
        'non_working': 'Non-Working Day',
    }
    for d in days:
        is_sd = d['status'] == 'stood_down'
        is_nw = d['status'] == 'non_working'
        pdf.set_text_color(180, 30, 30) if is_sd else (
            pdf.set_text_color(160, 160, 160) if is_nw else pdf.set_text_color(0, 0, 0))
        pdf.set_font('Helvetica', 'B' if is_sd else '', 8)
        pdf.cell(col_w[0], 6, d['date'].strftime('%d/%m/%Y'), border=1)
        pdf.cell(col_w[1], 6, d['day_name'], border=1)
        pdf.cell(col_w[2], 6, status_labels.get(d['status'], d['status']), border=1)
        pdf.set_font('Helvetica', '', 8)
        pdf.cell(col_w[3], 6, safe(d['reason'])[:60], border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)

    pdf.set_fill_color(240, 244, 255)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'SUMMARY', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
    pdf.set_font('Helvetica', '', 9)
    detail_row('Total Calendar Days', summary['total_days'])
    detail_row('Working Days in Period', summary['working_days'])
    detail_row('Days On Site (Billable)', summary['on_site'])
    detail_row('Days Stood Down', summary['stood_down'])
    detail_row('Saturday Billing', 'Included' if summary.get('count_saturdays', True) else 'Excluded')
    if summary['cost_week']:
        detail_row('Rate (per week)', f"${hm.cost_per_week:,.2f}")
    if summary.get('cost_per_day_derived') is not None:
        detail_row('Rate (per day, derived)', f"${summary['cost_per_day_derived']:,.2f}")
    if summary['cost_day'] is not None:
        detail_row('Estimated Cost (on-site days)', f"${summary['cost_day']:,.2f}")
    pdf.ln(4)

    sd_days = [d for d in days if d['status'] == 'stood_down']
    if sd_days:
        pdf.set_fill_color(255, 240, 240)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, 7, 'STAND-DOWN DAYS (NOT CHARGED)', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
        pdf.set_font('Helvetica', '', 9)
        for d in sd_days:
            pdf.set_font('Helvetica', 'B', 9)
            pdf.cell(35, 6, d['date'].strftime('%d/%m/%Y') + ' ' + d['day_name'][:3] + ':')
            pdf.set_font('Helvetica', '', 9)
            pdf.cell(0, 6, safe(d['reason']) or '-', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(4)

    pdf.ln(6)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(90, 6, 'Prepared by: ________________________________')
    pdf.cell(0, 6, 'Date: ____________________', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(3)
    pdf.cell(90, 6, 'Signature: ___________________________________', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


def generate_delay_pdf(rows, summary, date_from, date_to, project_name, settings):
    """Build a client delay charge report PDF and return bytes."""
    pdf = FPDF()
    pdf.set_margins(15, 15, 15)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    company = safe(settings.get('company_name', '') or 'Project Tracker')

    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 10, company, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', 'B', 13)
    pdf.cell(0, 8, 'DELAY CHARGE REPORT', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 5, f'Generated: {date.today().strftime("%d/%m/%Y")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(5)

    pdf.set_fill_color(240, 244, 255)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'REPORT DETAILS', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)

    def detail_row(label, value):
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(45, 6, label + ':')
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, safe(value) if value else '-', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    detail_row('Period', f'{date_from.strftime("%d/%m/%Y")} to {date_to.strftime("%d/%m/%Y")}')
    detail_row('Project', project_name or 'All Projects')
    detail_row('Billable Delay Events', str(summary['billable_count']))
    detail_row('Total Billable Hours', f'{summary["total_hours_billable"]} hrs')
    detail_row('Non-Billable Events', str(summary['non_billable_count']))
    pdf.ln(4)

    billable_rows = [r for r in rows if r['billable']]
    non_billable_rows = [r for r in rows if not r['billable']]

    def render_rows(rows_list, title, fill_color):
        if not rows_list:
            return
        pdf.set_fill_color(*fill_color)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
        pdf.ln(1)

        col_w = [70, 30, 30, 50]
        for row in rows_list:
            entry = row['entry']
            pdf.set_fill_color(230, 235, 255)
            pdf.set_font('Helvetica', 'B', 10)
            label = safe(f'{entry.entry_date.strftime("%d/%m/%Y")} ({entry.day_name})  -  '
                         f'{entry.project.name}  -  {entry.delay_hours} hrs delay')
            pdf.cell(0, 7, label, new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)

            pdf.set_font('Helvetica', 'I', 9)
            pdf.cell(0, 5, safe(f'Reason: {entry.delay_reason or "Not specified"}'), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            if entry.delay_description:
                pdf.multi_cell(0, 5, safe(entry.delay_description))
            pdf.ln(1)

            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_fill_color(210, 218, 255)
            for header, w in zip(['Role / Equipment', 'Rate ($/hr)', 'Hours', 'Cost ($)'], col_w):
                pdf.cell(w, 6, header, border=1, fill=True)
            pdf.ln()

            pdf.set_font('Helvetica', '', 8)
            if row['emp_lines']:
                for line in row['emp_lines']:
                    pdf.cell(col_w[0], 5, safe('  ' + line['name']), border=1)
                    pdf.cell(col_w[1], 5, f'${line["rate"]:.2f}', border=1)
                    pdf.cell(col_w[2], 5, str(line['hours']), border=1)
                    pdf.cell(col_w[3], 5, f'${line["cost"]:.2f}', border=1)
                    pdf.ln()
            if row['machine_lines']:
                for line in row['machine_lines']:
                    pdf.cell(col_w[0], 5, safe('  ' + line['name']), border=1)
                    pdf.cell(col_w[1], 5, f'${line["rate"]:.2f}', border=1)
                    pdf.cell(col_w[2], 5, str(line['hours']), border=1)
                    pdf.cell(col_w[3], 5, f'${line["cost"]:.2f}', border=1)
                    pdf.ln()

            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_text_color(30, 80, 180)
            pdf.cell(0, 6, f'  Event Total: ${row["entry_cost"]:.2f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(3)

    render_rows(billable_rows, 'BILLABLE DELAYS (CHARGED TO CLIENT)', (255, 235, 235))
    render_rows(non_billable_rows, 'NON-BILLABLE DELAYS (OWN COST)', (235, 245, 255))

    pdf.ln(3)
    pdf.set_fill_color(30, 80, 180)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 10, f'  TOTAL BILLABLE DELAY CHARGES:  ${summary["total_cost"]:,.2f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
    pdf.set_text_color(0, 0, 0)

    pdf.ln(8)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(90, 6, 'Authorised by: ________________________________')
    pdf.cell(0, 6, 'Date: ____________________', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(3)
    pdf.cell(90, 6, 'Signature: ___________________________________', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    today = date.today()
    recent_entries = (
        DailyEntry.query
        .order_by(DailyEntry.entry_date.desc(), DailyEntry.created_at.desc())
        .limit(10).all()
    )
    total_entries = DailyEntry.query.count()
    entries_today = DailyEntry.query.filter_by(entry_date=today).count()
    active_projects = Project.query.filter_by(active=True).count()
    active_hired = HiredMachine.query.filter_by(active=True).count()
    return render_template('index.html', recent_entries=recent_entries,
                           total_entries=total_entries, entries_today=entries_today,
                           active_projects=active_projects, active_hired=active_hired,
                           today=today)


# ---------------------------------------------------------------------------
# Daily Entry — New / Edit / Delete / List
# ---------------------------------------------------------------------------

@app.route('/entry/new', methods=['GET', 'POST'])
def new_entry():
    projects = Project.query.filter_by(active=True).order_by(Project.name).all()
    employees = Employee.query.filter_by(active=True).order_by(Employee.name).all()
    machines = Machine.query.filter_by(active=True).order_by(Machine.name).all()
    hired_machines = HiredMachine.query.filter_by(active=True).order_by(HiredMachine.machine_name).all()

    if request.method == 'POST':
        project_id = request.form.get('project_id')
        entry_date_str = request.form.get('entry_date')
        if not project_id or not entry_date_str:
            flash('Project and date are required.', 'danger')
            return render_template('entry_form.html', projects=projects, employees=employees,
                                   machines=machines, hired_machines=hired_machines,
                                   standdown_machine_ids=[], today=date.today())
        try:
            entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format.', 'danger')
            return render_template('entry_form.html', projects=projects, employees=employees,
                                   machines=machines, hired_machines=hired_machines,
                                   standdown_machine_ids=[], today=date.today())

        delay_hours = float(request.form.get('delay_hours') or 0)
        delay_billable = request.form.get('delay_billable', 'true') == 'true'
        machines_stood_down = bool(request.form.get('machines_stood_down'))

        entry = DailyEntry(
            project_id=int(project_id),
            entry_date=entry_date,
            lot_number=request.form.get('lot_number', '').strip() or None,
            location=request.form.get('location', '').strip() or None,
            material=request.form.get('material', '').strip() or None,
            num_people=int(request.form.get('num_people')) if request.form.get('num_people') else None,
            install_hours=float(request.form.get('install_hours') or 0),
            install_sqm=float(request.form.get('install_sqm') or 0),
            delay_hours=delay_hours,
            delay_billable=delay_billable,
            delay_reason=(request.form.get('delay_reason', '').strip() or None) if delay_hours > 0 else None,
            delay_description=request.form.get('delay_description', '').strip() or None,
            machines_stood_down=machines_stood_down,
            weather=request.form.get('weather', '').strip() or None,
            notes=request.form.get('notes', '').strip() or None,
            other_work_description=request.form.get('other_work_description', '').strip() or None,
            user_id=current_user.id,
        )
        employee_ids = request.form.getlist('employee_ids')
        machine_ids = request.form.getlist('machine_ids')
        if employee_ids:
            entry.employees = Employee.query.filter(Employee.id.in_(employee_ids)).all()
        if machine_ids:
            entry.machines = Machine.query.filter(Machine.id.in_(machine_ids)).all()
        db.session.add(entry)
        db.session.flush()  # get entry.id before photos

        # Photo uploads
        photos = request.files.getlist('photos')
        for photo in photos:
            if photo and photo.filename and allowed_photo(photo.filename):
                ext = photo.filename.rsplit('.', 1)[1].lower()
                stored_name = f"photo_{uuid.uuid4().hex}.{ext}"
                storage.upload_file(photo, f'photos/{stored_name}', os.path.join(UPLOAD_FOLDER, stored_name))
                caption = request.form.get(f'caption_{photo.filename}', '').strip() or None
                db.session.add(EntryPhoto(entry_id=entry.id, filename=stored_name,
                                          original_name=secure_filename(photo.filename),
                                          caption=caption))

        # Auto-create stand-downs for selected hired machines
        standdown_ids = request.form.getlist('standdown_machine_ids')
        if standdown_ids and delay_hours > 0:
            sd_reason = (entry.delay_description or entry.delay_reason or 'Delay')
            sd_count = 0
            for hm_id in standdown_ids:
                hm_obj = HiredMachine.query.get(int(hm_id))
                if hm_obj:
                    existing = StandDown.query.filter_by(
                        hired_machine_id=hm_obj.id, stand_down_date=entry_date).first()
                    if not existing:
                        db.session.add(StandDown(
                            hired_machine_id=hm_obj.id,
                            entry_id=entry.id,
                            stand_down_date=entry_date,
                            reason=sd_reason))
                        sd_count += 1
            if sd_count:
                flash(f'Stand-down recorded for {sd_count} hired machine{"s" if sd_count != 1 else ""}.', 'info')

        db.session.commit()
        flash('Entry saved successfully!', 'success')
        return redirect(url_for('entries'))

    # Build machine_project_map: {machine_id: [project_id, ...]} for JS filtering
    all_pm = ProjectMachine.query.all()
    machine_project_map = {}
    for pm in all_pm:
        machine_project_map.setdefault(pm.machine_id, [])
        machine_project_map[pm.machine_id].append(pm.project_id)

    return render_template('entry_form.html', projects=projects, employees=employees,
                           machines=machines, hired_machines=hired_machines,
                           machine_project_map=machine_project_map,
                           standdown_machine_ids=[], today=date.today())


@app.route('/entry/<int:entry_id>/edit', methods=['GET', 'POST'])
def edit_entry(entry_id):
    entry = DailyEntry.query.get_or_404(entry_id)
    projects = Project.query.filter_by(active=True).order_by(Project.name).all()
    employees = Employee.query.filter_by(active=True).order_by(Employee.name).all()
    machines = Machine.query.filter_by(active=True).order_by(Machine.name).all()
    hired_machines = HiredMachine.query.filter_by(active=True).order_by(HiredMachine.machine_name).all()

    if request.method == 'POST':
        entry.project_id = int(request.form.get('project_id'))
        entry.lot_number = request.form.get('lot_number', '').strip() or None
        entry.location = request.form.get('location', '').strip() or None
        entry.material = request.form.get('material', '').strip() or None
        num_people = request.form.get('num_people')
        entry.num_people = int(num_people) if num_people else None
        entry.install_hours = float(request.form.get('install_hours') or 0)
        entry.install_sqm = float(request.form.get('install_sqm') or 0)
        entry.delay_hours = float(request.form.get('delay_hours') or 0)
        entry.delay_billable = request.form.get('delay_billable', 'true') == 'true'
        entry.delay_reason = (request.form.get('delay_reason', '').strip() or None) if entry.delay_hours > 0 else None
        entry.delay_description = request.form.get('delay_description', '').strip() or None
        entry.machines_stood_down = bool(request.form.get('machines_stood_down'))
        entry.weather = request.form.get('weather', '').strip() or None
        entry.notes = request.form.get('notes', '').strip() or None
        entry.other_work_description = request.form.get('other_work_description', '').strip() or None
        try:
            entry.entry_date = datetime.strptime(request.form.get('entry_date'), '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format.', 'danger')

        employee_ids = request.form.getlist('employee_ids')
        machine_ids = request.form.getlist('machine_ids')
        entry.employees = Employee.query.filter(Employee.id.in_(employee_ids)).all() if employee_ids else []
        entry.machines = Machine.query.filter(Machine.id.in_(machine_ids)).all() if machine_ids else []
        entry.updated_at = datetime.utcnow()

        # New photo uploads
        photos = request.files.getlist('photos')
        for photo in photos:
            if photo and photo.filename and allowed_photo(photo.filename):
                ext = photo.filename.rsplit('.', 1)[1].lower()
                stored_name = f"photo_{uuid.uuid4().hex}.{ext}"
                storage.upload_file(photo, f'photos/{stored_name}', os.path.join(UPLOAD_FOLDER, stored_name))
                db.session.add(EntryPhoto(entry_id=entry.id, filename=stored_name,
                                          original_name=secure_filename(photo.filename)))

        # Stand-downs for newly selected hired machines
        standdown_ids = request.form.getlist('standdown_machine_ids')
        if standdown_ids and entry.delay_hours > 0:
            sd_reason = (entry.delay_description or entry.delay_reason or 'Delay')
            sd_count = 0
            for hm_id in standdown_ids:
                hm_obj = HiredMachine.query.get(int(hm_id))
                if hm_obj:
                    existing = StandDown.query.filter_by(
                        hired_machine_id=hm_obj.id, stand_down_date=entry.entry_date).first()
                    if not existing:
                        db.session.add(StandDown(
                            hired_machine_id=hm_obj.id,
                            entry_id=entry.id,
                            stand_down_date=entry.entry_date,
                            reason=sd_reason))
                        sd_count += 1
            if sd_count:
                flash(f'Stand-down recorded for {sd_count} hired machine{"s" if sd_count != 1 else ""}.', 'info')

        db.session.commit()
        flash('Entry updated successfully!', 'success')
        return redirect(url_for('entries'))

    # Pre-check machines that already have a stand-down on this entry's date
    existing_sd_ids = {
        sd.hired_machine_id
        for hm in hired_machines
        for sd in hm.stand_downs
        if sd.stand_down_date == entry.entry_date
    }

    # Build machine_project_map: {machine_id: [project_id, ...]} for JS filtering
    all_pm = ProjectMachine.query.all()
    machine_project_map = {}
    for pm in all_pm:
        machine_project_map.setdefault(pm.machine_id, [])
        machine_project_map[pm.machine_id].append(pm.project_id)

    return render_template('entry_form.html', entry=entry, projects=projects,
                           employees=employees, machines=machines,
                           hired_machines=hired_machines,
                           machine_project_map=machine_project_map,
                           standdown_machine_ids=existing_sd_ids,
                           selected_employee_ids=[e.id for e in entry.employees],
                           selected_machine_ids=[m.id for m in entry.machines],
                           today=date.today())


@app.route('/entry/<int:entry_id>/delete', methods=['POST'])
def delete_entry(entry_id):
    entry = DailyEntry.query.get_or_404(entry_id)
    # Remove photo files
    for photo in entry.photos:
        storage.delete_file(f'photos/{photo.filename}', os.path.join(UPLOAD_FOLDER, photo.filename))
    db.session.delete(entry)
    db.session.commit()
    flash('Entry deleted.', 'info')
    return redirect(url_for('entries'))


@app.route('/entry/<int:entry_id>/photo/<int:photo_id>/delete', methods=['POST'])
def delete_photo(entry_id, photo_id):
    photo = EntryPhoto.query.get_or_404(photo_id)
    if photo.entry_id != entry_id:
        flash('Invalid request.', 'danger')
        return redirect(url_for('edit_entry', entry_id=entry_id))
    storage.delete_file(f'photos/{photo.filename}', os.path.join(UPLOAD_FOLDER, photo.filename))
    db.session.delete(photo)
    db.session.commit()
    flash('Photo deleted.', 'info')
    return redirect(url_for('edit_entry', entry_id=entry_id))


@app.route('/entry-photo/<filename>')
def serve_entry_photo(filename):
    return storage.serve_file(f'photos/{filename}', os.path.join(UPLOAD_FOLDER, filename))


@app.route('/entries')
def entries():
    project_filter = request.args.get('project_id', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    query = DailyEntry.query.order_by(DailyEntry.entry_date.desc(), DailyEntry.created_at.desc())
    if project_filter:
        query = query.filter_by(project_id=int(project_filter))
    if date_from:
        try:
            query = query.filter(DailyEntry.entry_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(DailyEntry.entry_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    all_entries = query.all()
    projects = Project.query.order_by(Project.name).all()
    return render_template('entries_list.html', entries=all_entries, projects=projects,
                           project_filter=project_filter, date_from=date_from, date_to=date_to)


# ---------------------------------------------------------------------------
# Progress API — running totals for entry form
# ---------------------------------------------------------------------------

@app.route('/api/progress')
def api_progress():
    project_id = request.args.get('project_id', '')
    lot = request.args.get('lot', '').strip()
    material = request.args.get('material', '').strip()
    exclude_entry_id = request.args.get('exclude_id', '')

    if not project_id:
        return jsonify({'error': 'project_id required'}), 400

    # Planned SQM for this lot+material
    planned_q = PlannedData.query.filter_by(project_id=int(project_id))
    if lot:
        planned_q = planned_q.filter_by(lot=lot)
    if material:
        planned_q = planned_q.filter_by(material=material)
    planned_sqm = sum(p.planned_sqm or 0 for p in planned_q.all())

    # Installed SQM so far
    actual_q = (DailyEntry.query
                .filter_by(project_id=int(project_id))
                .filter(DailyEntry.install_sqm > 0))
    if lot:
        actual_q = actual_q.filter(DailyEntry.lot_number == lot)
    if material:
        actual_q = actual_q.filter(DailyEntry.material == material)
    if exclude_entry_id:
        try:
            actual_q = actual_q.filter(DailyEntry.id != int(exclude_entry_id))
        except ValueError:
            pass
    installed_sqm = sum(e.install_sqm or 0 for e in actual_q.all())

    remaining = max(0, planned_sqm - installed_sqm)
    pct = round(installed_sqm / planned_sqm * 100, 1) if planned_sqm > 0 else None

    return jsonify({
        'planned_sqm': round(planned_sqm, 2),
        'installed_sqm': round(installed_sqm, 2),
        'remaining': round(remaining, 2),
        'pct_complete': pct,
        'has_planned': planned_sqm > 0,
    })


# ---------------------------------------------------------------------------
# Project Dashboard
# ---------------------------------------------------------------------------



@app.route('/api/project/<int:project_id>/machines')
def api_project_machines(project_id):
    """Return owned machines assigned to this project."""
    assignments = (ProjectMachine.query
                   .filter_by(project_id=project_id)
                   .join(Machine)
                   .filter(Machine.active == True)
                   .all())
    machines = [{'id': pm.machine.id, 'name': pm.machine.name,
                 'machine_type': pm.machine.machine_type or ''}
                for pm in assignments]
    return jsonify({'machines': machines})


@app.route('/api/planned-options')
def api_planned_options():
    """Return distinct lots (and materials for a given lot) from planned data."""
    project_id = request.args.get('project_id', '').strip()
    lot = request.args.get('lot', '').strip()
    if not project_id:
        return jsonify({'lots': [], 'materials': []})
    q = PlannedData.query.filter_by(project_id=int(project_id))
    rows = q.all()
    lots = sorted({r.lot for r in rows if r.lot}, key=_natural_key)
    if lot:
        materials = sorted({r.material for r in rows if r.material and r.lot == lot},
                           key=_natural_key)
    else:
        materials = sorted({r.material for r in rows if r.material}, key=_natural_key)
    return jsonify({'lots': lots, 'materials': materials})


@app.route('/project/<int:project_id>/dashboard')
def project_dashboard(project_id):
    project = Project.query.get_or_404(project_id)
    progress = compute_project_progress(project_id)
    gantt_data = compute_gantt_data(project_id)
    delay_summary = compute_delay_summary(project_id)
    non_work_dates = ProjectNonWorkDate.query.filter_by(project_id=project_id).order_by(ProjectNonWorkDate.date).all()
    has_openpyxl = HAS_OPENPYXL

    # Budgeted crew vs actual — actual counts employees by role from all entries
    budgeted_roles = ProjectBudgetedRole.query.filter_by(project_id=project_id).order_by(ProjectBudgetedRole.role_name).all()
    all_roles = Role.query.order_by(Role.name).all()

    # Build actual crew counts by role from entries (last 30 days or all time if sparse)
    actual_role_counts = {}
    all_entries_for_crew = DailyEntry.query.filter_by(project_id=project_id).all()
    for e in all_entries_for_crew:
        for emp in e.employees:
            role_name = emp.role or 'Unassigned'
            actual_role_counts[role_name] = actual_role_counts.get(role_name, 0) + 1
    # Convert to per-entry average
    entry_count = len(all_entries_for_crew) or 1
    actual_role_avg = {k: round(v / entry_count, 1) for k, v in actual_role_counts.items()}

    # Hired machines at this project
    hired_machines_list = HiredMachine.query.filter_by(project_id=project_id, active=True).all()

    # Owned machines assigned to this project
    owned_machine_assignments = (ProjectMachine.query
                                  .filter_by(project_id=project_id)
                                  .join(Machine)
                                  .filter(Machine.active == True)
                                  .all())
    owned_machines_list = owned_machine_assignments  # list of ProjectMachine rows

    # Worked Sundays for this project
    worked_sundays_list = (ProjectWorkedSunday.query
                           .filter_by(project_id=project_id)
                           .order_by(ProjectWorkedSunday.date)
                           .all())

    # Project documents
    project_documents = (ProjectDocument.query
                         .filter_by(project_id=project_id)
                         .order_by(ProjectDocument.uploaded_at.desc())
                         .all())

    # All owned machines (for "assign" dropdown)
    all_machines = Machine.query.filter_by(active=True).order_by(Machine.name).all()
    # Already-assigned machine ids
    assigned_machine_ids = {pm.machine_id for pm in owned_machine_assignments}

    # Cost estimation: labour (budgeted roles × hourly rate) + owned machines + hired machines
    role_rate_map = {r.name: (r.delay_rate or 0) for r in all_roles}
    hours_pd = project.hours_per_day or 8
    labour_daily = sum(
        br.budgeted_count * role_rate_map.get(br.role_name, 0) * hours_pd
        for br in budgeted_roles
    )
    # Hired machine daily cost — always derived from weekly rate
    hired_machine_daily = 0.0
    for hm in hired_machines_list:
        days_pw = 6 if hm.count_saturdays else 5
        hired_machine_daily += hm.cost_per_week / days_pw if hm.cost_per_week else 0
    # Owned machine daily cost (delay_rate × hours_per_day)
    owned_machine_daily = 0.0
    for pm in owned_machine_assignments:
        if pm.machine and pm.machine.delay_rate:
            owned_machine_daily += pm.machine.delay_rate * hours_pd
    machine_daily = hired_machine_daily + owned_machine_daily
    daily_cost = labour_daily + machine_daily
    target_cost = (daily_cost * project.quoted_days
                   if project.quoted_days and daily_cost > 0 else None)
    # Public holidays and CFMEU dates for this project's state
    state_holidays = []
    if project.state:
        state_holidays = [h for h in PublicHoliday.query.order_by(PublicHoliday.date).all()
                          if project.state in h.states_list()]
        if project.is_cfmeu:
            cfmeu = [c for c in CFMEUDate.query.order_by(CFMEUDate.date).all()
                     if 'ALL' in c.states_list() or project.state in c.states_list()]
            state_holidays = sorted(state_holidays + cfmeu, key=lambda x: x.date)
    # Build non-work date set including state holidays
    holiday_dates = {h.date for h in state_holidays}
    non_work_set_dates = {nwd.date for nwd in non_work_dates} | holiday_dates
    est_finish_date = None
    if gantt_data and gantt_data.get('est_finish'):
        try:
            from datetime import datetime as _dt
            est_finish_date = _dt.strptime(gantt_data['est_finish'], '%d/%m/%Y').date()
        except Exception:
            pass
    forecast_working_days = None
    forecast_cost = None
    if est_finish_date and project.start_date:
        forecast_working_days = sum(
            1 for i in range((est_finish_date - project.start_date).days + 1)
            if (project.start_date + timedelta(days=i)).weekday() != 6
            and (project.start_date + timedelta(days=i)) not in non_work_set_dates
        )
        forecast_cost = daily_cost * forecast_working_days if daily_cost > 0 else None
    cost_variance = ((forecast_cost - target_cost)
                     if forecast_cost is not None and target_cost is not None else None)
    cost_estimate = {
        'labour_daily': round(labour_daily, 2),
        'hired_machine_daily': round(hired_machine_daily, 2),
        'owned_machine_daily': round(owned_machine_daily, 2),
        'machine_daily': round(machine_daily, 2),
        'daily_total': round(daily_cost, 2),
        'target_cost': round(target_cost, 2) if target_cost is not None else None,
        'forecast_cost': round(forecast_cost, 2) if forecast_cost is not None else None,
        'cost_variance': round(cost_variance, 2) if cost_variance is not None else None,
        'forecast_working_days': forecast_working_days,
        'has_rates': daily_cost > 0,
    }

    # All distinct scheduling groups for the scheduling tab
    _all_roles = Role.query.filter(Role.group_name.isnot(None)).all()
    _seen = set()
    role_groups_all = []
    for r in _all_roles:
        if r.group_name and r.group_name not in _seen:
            _seen.add(r.group_name)
            role_groups_all.append(r.group_name)
    role_groups_all.sort()
    # Build a dict of current budgeted counts by role_name for pre-filling
    budgeted_by_group = {br.role_name: br.budgeted_count for br in budgeted_roles}

    # Equipment requirements for scheduling tab
    equip_reqs = (ProjectEquipmentRequirement.query
                  .filter_by(project_id=project_id)
                  .order_by(ProjectEquipmentRequirement.label)
                  .all())
    # Build coverage rows — assignments are explicit (not type-matched)
    equip_coverage = []
    # Track which machine/hired_machine ids are already assigned to any requirement on this project
    assigned_own_ids = set()
    assigned_hired_ids = set()
    for er in equip_reqs:
        for a in er.assignments:
            if a.machine_id:
                assigned_own_ids.add(a.machine_id)
            if a.hired_machine_id:
                assigned_hired_ids.add(a.hired_machine_id)
    for er in equip_reqs:
        total = len(er.assignments)
        equip_coverage.append({
            'req': er,
            'label': er.label,
            'required': er.required_count,
            'assignments': er.assignments,
            'total': total,
            'gap': max(0, er.required_count - total),
        })
    # All own fleet machines available to assign (active, not yet assigned to any requirement on this project)
    assignable_own = [m for m in Machine.query.filter_by(active=True).order_by(Machine.name).all()
                      if m.id not in assigned_own_ids]
    # All hired machines for this project not yet assigned to a requirement
    assignable_hired = [hm for hm in hired_machines_list if hm.id not in assigned_hired_ids]

    return render_template('project_dashboard.html',
                           project=project, progress=progress,
                           gantt_data=gantt_data,
                           delay_summary=delay_summary,
                           non_work_dates=non_work_dates,
                           has_openpyxl=has_openpyxl,
                           budgeted_roles=budgeted_roles,
                           all_roles=all_roles,
                           actual_role_avg=actual_role_avg,
                           hired_machines_list=hired_machines_list,
                           owned_machines_list=owned_machines_list,
                           all_machines=all_machines,
                           assigned_machine_ids=assigned_machine_ids,
                           worked_sundays_list=worked_sundays_list,
                           project_documents=project_documents,
                           cost_estimate=cost_estimate,
                           role_groups_all=role_groups_all,
                           budgeted_by_group=budgeted_by_group,
                           equip_reqs=equip_reqs,
                           equip_coverage=equip_coverage,
                           assignable_own=assignable_own,
                           assignable_hired=assignable_hired,
                           state_holidays=state_holidays,
                           today=date.today())


# ---------------------------------------------------------------------------
# Panel diagram routes
# ---------------------------------------------------------------------------

SVG_FOLDER = os.path.join(os.path.dirname(__file__), 'instance', 'uploads', 'panels')


@app.route('/project/<int:project_id>/panels')
@login_required
def panel_overview(project_id):
    project = Project.query.get_or_404(project_id)
    layers = (DiagramLayer.query
              .filter_by(project_id=project_id)
              .order_by(DiagramLayer.sort_order, DiagramLayer.layer_name)
              .all())
    return render_template('panels/overview.html', project=project, layers=layers)


@app.route('/project/<int:project_id>/panels/layer/add', methods=['POST'])
@login_required
def panel_layer_add(project_id):
    project = Project.query.get_or_404(project_id)
    layer_name = request.form.get('layer_name', '').strip()
    description = request.form.get('description', '').strip() or None
    sort_order = int(request.form.get('sort_order', 0) or 0)
    if not layer_name:
        flash('Layer name is required.', 'warning')
        return redirect(url_for('panel_overview', project_id=project_id))
    layer = DiagramLayer(
        project_id=project_id,
        layer_name=layer_name,
        description=description,
        sort_order=sort_order,
    )
    svg_file = request.files.get('svg_file')
    if svg_file and svg_file.filename:
        os.makedirs(SVG_FOLDER, exist_ok=True)
        ext = svg_file.filename.rsplit('.', 1)[-1].lower()
        if ext == 'dxf':
            if not _DXF_AVAILABLE:
                flash('DXF support not installed. Run: pip install ezdxf', 'danger')
                return redirect(url_for('panel_overview', project_id=project_id))
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.dxf', delete=False) as tmp:
                svg_file.save(tmp.name)
                tmp_path = tmp.name
            try:
                svg_content = _dxf_to_svg(tmp_path)
            except Exception as exc:
                os.unlink(tmp_path)
                db.session.add(layer)
                db.session.commit()
                flash(f'Layer added but DXF conversion failed: {exc}', 'warning')
                return redirect(url_for('panel_overview', project_id=project_id))
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
            stored_name = f'{uuid.uuid4().hex}.svg'
            storage.upload_text(svg_content, f'panels/{stored_name}', os.path.join(SVG_FOLDER, stored_name))
            layer.svg_filename = stored_name
            layer.svg_original_name = secure_filename(svg_file.filename).replace('.dxf', '.svg')
        elif ext == 'svg':
            stored_name = f'{uuid.uuid4().hex}.svg'
            storage.upload_file(svg_file, f'panels/{stored_name}', os.path.join(SVG_FOLDER, stored_name))
            layer.svg_filename = stored_name
            layer.svg_original_name = secure_filename(svg_file.filename)
        else:
            flash('Only SVG or DXF files are accepted for diagrams.', 'warning')
            return redirect(url_for('panel_overview', project_id=project_id))
    db.session.add(layer)
    db.session.commit()
    flash(f'Layer "{layer_name}" added.', 'success')
    return redirect(url_for('panel_overview', project_id=project_id))


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/upload-svg', methods=['POST'])
@login_required
def panel_layer_upload_svg(project_id, layer_id):
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return 'Not found', 404
    svg_file = request.files.get('svg_file')
    if not svg_file or not svg_file.filename:
        flash('No file selected.', 'warning')
        return redirect(url_for('panel_overview', project_id=project_id))
    ext = svg_file.filename.rsplit('.', 1)[-1].lower()
    if ext != 'svg':
        flash('Only SVG files are accepted.', 'warning')
        return redirect(url_for('panel_overview', project_id=project_id))
    os.makedirs(SVG_FOLDER, exist_ok=True)
    if layer.svg_filename:
        storage.delete_file(f'panels/{layer.svg_filename}', os.path.join(SVG_FOLDER, layer.svg_filename))
    stored_name = f'{uuid.uuid4().hex}.svg'
    storage.upload_file(svg_file, f'panels/{stored_name}', os.path.join(SVG_FOLDER, stored_name))
    layer.svg_filename = stored_name
    layer.svg_original_name = secure_filename(svg_file.filename)
    db.session.commit()
    flash('SVG diagram uploaded.', 'success')
    return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/upload-dxf', methods=['POST'])
@login_required
def panel_layer_upload_dxf(project_id, layer_id):
    """Accept a DXF file, convert it to SVG with panel IDs, and save it."""
    if not _DXF_AVAILABLE:
        flash('DXF support is not installed. Run: pip install ezdxf', 'danger')
        return redirect(url_for('panel_overview', project_id=project_id))
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return 'Not found', 404
    dxf_file = request.files.get('dxf_file')
    if not dxf_file or not dxf_file.filename:
        flash('No DXF file selected.', 'warning')
        return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))
    ext = dxf_file.filename.rsplit('.', 1)[-1].lower()
    if ext != 'dxf':
        flash('Only DXF files are accepted here.', 'warning')
        return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))
    # Save the DXF to a temp path, convert, then discard
    import tempfile
    os.makedirs(SVG_FOLDER, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix='.dxf', delete=False) as tmp:
        dxf_file.save(tmp.name)
        tmp_path = tmp.name
    try:
        svg_content = _dxf_to_svg(tmp_path)
    except Exception as exc:
        os.unlink(tmp_path)
        flash(f'DXF conversion failed: {exc}', 'danger')
        return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    # Remove old SVG if present
    if layer.svg_filename:
        storage.delete_file(f'panels/{layer.svg_filename}', os.path.join(SVG_FOLDER, layer.svg_filename))
    stored_name = f'{uuid.uuid4().hex}.svg'
    storage.upload_text(svg_content, f'panels/{stored_name}', os.path.join(SVG_FOLDER, stored_name))
    layer.svg_filename = stored_name
    layer.svg_original_name = secure_filename(dxf_file.filename).replace('.dxf', '.svg')
    db.session.commit()
    flash('DXF converted and diagram saved. Click a panel to start recording.', 'success')
    return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/upload-bg', methods=['POST'])
@login_required
def panel_layer_upload_bg(project_id, layer_id):
    """Upload a background image (JPG/PNG) or PDF (first page) for a diagram layer."""
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return 'Not found', 404
    bg_file = request.files.get('bg_file')
    if not bg_file or not bg_file.filename:
        flash('No file selected.', 'warning')
        return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))
    ext = bg_file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('jpg', 'jpeg', 'png', 'pdf'):
        flash('Background must be JPG, PNG, or PDF.', 'warning')
        return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))
    os.makedirs(SVG_FOLDER, exist_ok=True)
    # Remove old background
    if layer.bg_filename:
        storage.delete_file(f'panels/{layer.bg_filename}', os.path.join(SVG_FOLDER, layer.bg_filename))
    stored_name = f'{uuid.uuid4().hex}_bg.png'
    dest = os.path.join(SVG_FOLDER, stored_name)
    os.makedirs(SVG_FOLDER, exist_ok=True)
    if ext == 'pdf':
        if not _PYMUPDF_AVAILABLE:
            flash('PDF conversion requires PyMuPDF. Run: pip install PyMuPDF', 'danger')
            return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            bg_file.save(tmp.name)
            tmp_path = tmp.name
        try:
            doc = _fitz.open(tmp_path)
            page = doc[0]
            mat = _fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            pix.save(dest)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        storage.upload_local_file(dest, f'panels/{stored_name}')
    else:
        storage.upload_file(bg_file, f'panels/{stored_name}', dest)
    layer.bg_filename = stored_name
    layer.bg_original_name = secure_filename(bg_file.filename)
    db.session.commit()
    flash('Background image saved.', 'success')
    return redirect(url_for('panel_layer_view', project_id=project_id, layer_id=layer_id))


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/bg-image')
@login_required
def panel_layer_bg_image(project_id, layer_id):
    """Serve the background image for a diagram layer (diagram view)."""
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id or not layer.bg_filename:
        return '', 404
    return storage.serve_file(f'panels/{layer.bg_filename}', os.path.join(SVG_FOLDER, layer.bg_filename))



@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/autodetect', methods=['POST'])
@login_required
def panel_layer_autodetect(project_id, layer_id):
    """Run OpenCV contour detection on the background image and return panel polygons."""
    if not _CV2_AVAILABLE:
        return jsonify({'ok': False, 'error': 'opencv-python-headless not installed'}), 500
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id or not layer.bg_filename:
        return jsonify({'ok': False, 'error': 'No background image uploaded'}), 400
    bg_path = os.path.join(SVG_FOLDER, layer.bg_filename)
    if not os.path.exists(bg_path):
        return jsonify({'ok': False, 'error': 'Background image file not found'}), 404

    img = _cv2.imread(bg_path)
    if img is None:
        return jsonify({'ok': False, 'error': 'Could not read background image'}), 500
    h, w = img.shape[:2]

    # Get detection parameters from request
    data = request.get_json() or {}
    min_area_pct = float(data.get('min_area_pct', 0.002))   # 0.2% of image area
    max_area_pct = float(data.get('max_area_pct', 0.25))    # 25% of image area
    min_area = int(w * h * min_area_pct)
    max_area = int(w * h * max_area_pct)

    gray = _cv2.cvtColor(img, _cv2.COLOR_BGR2GRAY)

    threshold_val = int(data.get('threshold', 180))
    dilate_iter   = int(data.get('dilate', 3))

    _, binary = _cv2.threshold(gray, threshold_val, 255, _cv2.THRESH_BINARY_INV)
    kernel = _np.ones((3, 3), _np.uint8)

    # ── 1. Find the overall outer boundary of the drawing ─────────────
    # Use heavy dilation to ensure the outer outline is fully closed.
    outer_bin = _cv2.dilate(binary, kernel, iterations=dilate_iter * 5)
    cnts_outer, _ = _cv2.findContours(outer_bin, _cv2.RETR_EXTERNAL, _cv2.CHAIN_APPROX_SIMPLE)
    if not cnts_outer:
        return jsonify({'ok': False, 'error': 'No drawing boundary found. Try lowering the threshold.'})
    outer_cnt = max(cnts_outer, key=_cv2.contourArea)
    if _cv2.contourArea(outer_cnt) < w * h * 0.05:
        return jsonify({'ok': False, 'error': 'Drawing boundary too small. Check threshold.'})
    outer_mask = _np.zeros((h, w), _np.uint8)
    _cv2.fillPoly(outer_mask, [outer_cnt], 255)

    # ── 2. Detect panel divider lines using Hough transform ───────────
    # Use minimal dilation so we detect individual lines, not blobs.
    line_bin = _cv2.dilate(binary, kernel, iterations=1)
    lines = _cv2.HoughLinesP(
        line_bin, 1, _np.pi / 180,
        threshold=40,
        minLineLength=int(h * 0.15),   # line must span ≥15% of image height
        maxLineGap=int(h * 0.06),      # allow gaps up to 6% of image height
    )
    if lines is None or len(lines) < 2:
        return jsonify({'ok': False, 'error': 'Not enough panel lines detected. Try lowering threshold or gap-close.'})

    # ── 3. Keep lines that are mostly vertical (panel dividers) ───────
    # Panels can be tilted, so accept angles 25°–90° from horizontal.
    mid_y = h / 2.0
    dividers = []
    for ln in lines:
        x1, y1, x2, y2 = map(int, ln[0])
        dy = y2 - y1
        dx = x2 - x1
        if abs(dy) < 1:
            continue
        angle = abs(_np.degrees(_np.arctan2(dy, dx)))
        if angle < 25:
            continue   # too horizontal — skip boundary/baseline lines
        # Normalise so y1 is always the top point
        if y1 > y2:
            x1, y1, x2, y2 = x2, y2, x1, y1
        # X-position at mid-image (used for clustering / sorting)
        t = (mid_y - y1) / (y2 - y1) if (y2 - y1) != 0 else 0.5
        x_mid = x1 + t * (x2 - x1)
        dividers.append({'x_mid': x_mid, 'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})

    if len(dividers) < 2:
        return jsonify({'ok': False, 'error': 'Too few vertical lines found. Try lowering threshold.'})

    dividers.sort(key=lambda d: d['x_mid'])

    # ── 4. Cluster nearby dividers (merge duplicate Hough detections) ─
    cluster_gap = max(6, w // 100)
    clusters, current = [], [dividers[0]]
    for d in dividers[1:]:
        if d['x_mid'] - current[-1]['x_mid'] < cluster_gap:
            current.append(d)
        else:
            clusters.append(current)
            current = [d]
    clusters.append(current)

    # Average each cluster into one representative line
    merged = []
    for cl in clusters:
        merged.append({
            'x1': int(_np.mean([d['x1'] for d in cl])),
            'y1': int(_np.mean([d['y1'] for d in cl])),
            'x2': int(_np.mean([d['x2'] for d in cl])),
            'y2': int(_np.mean([d['y2'] for d in cl])),
        })

    # ── 5. Build panel quads from each adjacent pair of dividers ──────
    polygons = []
    for i in range(len(merged) - 1):
        L, R = merged[i], merged[i + 1]
        pts = [
            [L['x1'], L['y1']],   # top-left
            [R['x1'], R['y1']],   # top-right
            [R['x2'], R['y2']],   # bottom-right
            [L['x2'], L['y2']],   # bottom-left
        ]
        cnt_arr = _np.array(pts, dtype=_np.int32)
        area = _cv2.contourArea(cnt_arr)
        if area < min_area or area > max_area:
            continue
        # Panel must overlap the drawing boundary by at least 40%
        pmask = _np.zeros((h, w), _np.uint8)
        _cv2.fillPoly(pmask, [cnt_arr], 255)
        overlap = _cv2.countNonZero(_cv2.bitwise_and(pmask, outer_mask))
        total   = _cv2.countNonZero(pmask)
        if total > 0 and overlap / total < 0.4:
            continue
        polygons.append(pts)

    return jsonify({
        'ok': True,
        'img_width': w,
        'img_height': h,
        'polygons': polygons,
        'count': len(polygons),
    })


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/delete', methods=['POST'])
@login_required
def panel_layer_delete(project_id, layer_id):
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return 'Not found', 404
    for fname in (layer.svg_filename, layer.bg_filename):
        if fname:
            storage.delete_file(f'panels/{fname}', os.path.join(SVG_FOLDER, fname))
    db.session.delete(layer)
    db.session.commit()
    flash(f'Layer "{layer.layer_name}" deleted.', 'success')
    return redirect(url_for('panel_overview', project_id=project_id))


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>')
@login_required
def panel_layer_view(project_id, layer_id):
    project = Project.query.get_or_404(project_id)
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return 'Not found', 404
    all_layers = (DiagramLayer.query
                  .filter_by(project_id=project_id)
                  .order_by(DiagramLayer.sort_order, DiagramLayer.layer_name)
                  .all())
    svg_content = None
    if layer.svg_filename:
        svg_content = storage.read_text(f'panels/{layer.svg_filename}', os.path.join(SVG_FOLDER, layer.svg_filename))
    panel_data = {}
    for rec in layer.panels:
        panel_data[rec.panel_id] = {
            'id': rec.id,
            'panel_id': rec.panel_id,
            'panel_label': rec.panel_label or rec.panel_id,
            'status': rec.status,
            'installed_date': rec.installed_date.isoformat() if rec.installed_date else '',
            'employee_id': rec.employee_id or '',
            'notes': rec.notes or '',
            'roll_number': rec.roll_number or '',
            'install_time': rec.install_time or '',
            'width_m': rec.width_m if rec.width_m is not None else '',
            'length_m': rec.length_m if rec.length_m is not None else '',
            'area_sqm': rec.area_sqm if rec.area_sqm is not None else '',
            'panel_type': rec.panel_type or '',
        }
    employees = Employee.query.filter_by(active=True).order_by(Employee.name).all()
    bg_url = url_for('panel_layer_bg_image', project_id=project_id, layer_id=layer_id) \
             if layer.bg_filename else None
    return render_template('panels/layer.html',
                           project=project, layer=layer,
                           all_layers=all_layers,
                           svg_content=svg_content,
                           panel_data_json=json.dumps(panel_data),
                           employees=employees,
                           bg_url=bg_url,
                           cv2_available=_CV2_AVAILABLE)


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/record', methods=['POST'])
@login_required
def panel_record_save(project_id, layer_id):
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    data = request.get_json() or {}
    panel_id = (data.get('panel_id') or '').strip()
    if not panel_id:
        return jsonify({'ok': False, 'error': 'panel_id required'}), 400
    status = data.get('status', 'planned')
    panel_label = (data.get('panel_label') or '').strip() or panel_id
    notes = (data.get('notes') or '').strip() or None
    roll_number = (data.get('roll_number') or '').strip() or None
    install_time = (data.get('install_time') or '').strip() or None
    panel_type = (data.get('panel_type') or '').strip() or None
    def _float_or_none(v):
        try:
            return float(v) if v not in (None, '') else None
        except (ValueError, TypeError):
            return None
    width_m   = _float_or_none(data.get('width_m'))
    length_m  = _float_or_none(data.get('length_m'))
    area_sqm  = _float_or_none(data.get('area_sqm'))
    employee_id = data.get('employee_id') or None
    if employee_id:
        try:
            employee_id = int(employee_id)
        except (ValueError, TypeError):
            employee_id = None
    installed_date = None
    if data.get('installed_date'):
        try:
            from datetime import datetime as _dt
            installed_date = _dt.strptime(data['installed_date'], '%Y-%m-%d').date()
        except ValueError:
            pass
    rec = PanelInstallRecord.query.filter_by(layer_id=layer_id, panel_id=panel_id).first()
    if rec is None:
        rec = PanelInstallRecord(layer_id=layer_id, panel_id=panel_id)
        db.session.add(rec)
    rec.panel_label = panel_label
    rec.status = status
    rec.installed_date = installed_date
    rec.employee_id = employee_id
    rec.notes = notes
    rec.roll_number  = roll_number
    rec.install_time = install_time
    rec.width_m      = width_m
    rec.length_m     = length_m
    rec.area_sqm     = area_sqm
    rec.panel_type   = panel_type
    rec.recorded_by_id = current_user.id
    rec.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({
        'ok': True,
        'status': rec.status,
        'panel_id': rec.panel_id,
        'panel_label': panel_label,
        'roll_number':  roll_number,
        'install_time': install_time,
        'width_m':      width_m,
        'length_m':     length_m,
        'area_sqm':     area_sqm,
        'panel_type':   panel_type,
    })


@app.route('/project/<int:project_id>/panels/layer/<int:layer_id>/data.json')
@login_required
def panel_data_json(project_id, layer_id):
    layer = DiagramLayer.query.get_or_404(layer_id)
    if layer.project_id != project_id:
        return jsonify({}), 404
    data = {}
    for rec in layer.panels:
        data[rec.panel_id] = {
            'panel_label': rec.panel_label or rec.panel_id,
            'status': rec.status,
            'installed_date': rec.installed_date.isoformat() if rec.installed_date else '',
            'employee_id': rec.employee_id or '',
            'notes': rec.notes or '',
        }
    return jsonify(data)


@app.route('/project/<int:project_id>/non-work-dates/add', methods=['POST'])
@login_required
def non_work_date_add(project_id):
    project = Project.query.get_or_404(project_id)
    date_str = request.form.get('date', '').strip()
    reason = request.form.get('reason', '').strip()
    if not date_str:
        flash('Date is required.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    try:
        nw_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))

    existing = ProjectNonWorkDate.query.filter_by(project_id=project_id, date=nw_date).first()
    if existing:
        flash(f'{nw_date.strftime("%d/%m/%Y")} already added.', 'warning')
    else:
        db.session.add(ProjectNonWorkDate(project_id=project_id, date=nw_date, reason=reason or None))
        db.session.commit()
        flash(f'Non-work date {nw_date.strftime("%d/%m/%Y")} added.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/non-work-dates/<int:nwd_id>/delete', methods=['POST'])
@login_required
def non_work_date_delete(project_id, nwd_id):
    nwd = ProjectNonWorkDate.query.get_or_404(nwd_id)
    db.session.delete(nwd)
    db.session.commit()
    flash('Non-work date removed.', 'info')
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/budgeted-crew/add', methods=['POST'])
def budgeted_crew_add(project_id):
    Project.query.get_or_404(project_id)
    role_name = request.form.get('role_name', '').strip()
    count_raw = request.form.get('budgeted_count', '').strip()
    if not role_name or not count_raw:
        flash('Role name and count are required.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    try:
        count = int(count_raw)
        if count < 1:
            raise ValueError
    except ValueError:
        flash('Count must be a positive integer.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    existing = ProjectBudgetedRole.query.filter_by(project_id=project_id, role_name=role_name).first()
    if existing:
        existing.budgeted_count = count
        flash(f'Updated {role_name} budget to {count}.', 'info')
    else:
        db.session.add(ProjectBudgetedRole(project_id=project_id, role_name=role_name, budgeted_count=count))
        flash(f'Added {count}x {role_name} to budget.', 'success')
    db.session.commit()
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/budgeted-crew/<int:br_id>/delete', methods=['POST'])
def budgeted_crew_delete(project_id, br_id):
    br = ProjectBudgetedRole.query.get_or_404(br_id)
    db.session.delete(br)
    db.session.commit()
    flash('Budgeted role removed.', 'info')
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/planned-upload', methods=['GET', 'POST'])
def planned_upload(project_id):
    project = Project.query.get_or_404(project_id)

    if request.method == 'POST':
        if not HAS_OPENPYXL:
            flash('openpyxl is not installed. Run: pip install openpyxl', 'danger')
            return redirect(url_for('project_dashboard', project_id=project_id))

        file = request.files.get('planned_file')
        if not file or not file.filename:
            flash('No file selected.', 'danger')
            return redirect(url_for('project_dashboard', project_id=project_id))

        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            flash('Please upload an Excel (.xlsx/.xls) or CSV file.', 'danger')
            return redirect(url_for('project_dashboard', project_id=project_id))

        replace_existing = request.form.get('replace_existing') == 'yes'

        try:
            if ext == 'csv':
                import csv, io
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})

            # Map columns (case-insensitive, flexible names)
            def find_col(row_dict, *names):
                for name in names:
                    for k in row_dict:
                        if k and k.strip().lower() == name.lower():
                            return k
                return None

            if not rows:
                flash('File appears empty.', 'danger')
                return redirect(url_for('project_dashboard', project_id=project_id))

            sample = rows[0]
            col_day = find_col(sample, 'Day', 'day', 'DAY', 'Day Number')
            col_lot = find_col(sample, 'Lot', 'lot', 'LOT', 'Lot Number')
            col_loc = find_col(sample, 'Location', 'location', 'LOCATION', 'Loc')
            col_mat = find_col(sample, 'Material', 'material', 'MATERIAL')
            col_sqm = find_col(sample, 'Planned Sqm', 'Planned SQM', 'planned sqm', 'planned_sqm', 'SQM', 'sqm', 'Sqm')

            if not col_day or not col_sqm:
                flash('Could not find required columns (Day, Planned Sqm). Check column headers.', 'danger')
                return redirect(url_for('project_dashboard', project_id=project_id))

            if replace_existing:
                PlannedData.query.filter_by(project_id=project_id).delete()

            count = 0
            skipped = 0
            for row in rows:
                try:
                    day_val = row.get(col_day)
                    sqm_val = row.get(col_sqm)
                    if day_val is None or sqm_val is None:
                        skipped += 1
                        continue
                    day_num = int(float(str(day_val)))
                    sqm = float(str(sqm_val))
                    lot = str(row.get(col_lot) or '').strip() or None
                    location = str(row.get(col_loc) or '').strip() or None if col_loc else None
                    material = str(row.get(col_mat) or '').strip() or None
                    db.session.add(PlannedData(
                        project_id=project_id,
                        lot=lot, location=location, material=material,
                        day_number=day_num, planned_sqm=sqm,
                    ))
                    count += 1
                except (ValueError, TypeError):
                    skipped += 1

            db.session.commit()
            flash(f'Imported {count} planned rows.{f" Skipped {skipped} invalid rows." if skipped else ""}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {e}', 'danger')

        return redirect(url_for('project_dashboard', project_id=project_id))

    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/planned/clear', methods=['POST'])
def planned_clear(project_id):
    Project.query.get_or_404(project_id)
    deleted = PlannedData.query.filter_by(project_id=project_id).delete()
    db.session.commit()
    flash(f'Cleared {deleted} planned rows.', 'info')
    return redirect(url_for('project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Own Equipment (ProjectMachine) — add / remove
# ---------------------------------------------------------------------------

@app.route('/project/<int:project_id>/own-equipment/add', methods=['POST'])
def own_equipment_add(project_id):
    Project.query.get_or_404(project_id)
    machine_id_raw = request.form.get('machine_id', '').strip()
    if not machine_id_raw:
        flash('Please select a machine.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    machine_id = int(machine_id_raw)
    existing = ProjectMachine.query.filter_by(project_id=project_id, machine_id=machine_id).first()
    if existing:
        flash('That machine is already assigned to this project.', 'warning')
        return redirect(url_for('project_dashboard', project_id=project_id))
    assigned_date_str = request.form.get('assigned_date', '').strip()
    assigned_date = None
    if assigned_date_str:
        try:
            assigned_date = datetime.strptime(assigned_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    notes = request.form.get('notes', '').strip() or None
    db.session.add(ProjectMachine(project_id=project_id, machine_id=machine_id,
                                   assigned_date=assigned_date, notes=notes))
    db.session.commit()
    machine = Machine.query.get(machine_id)
    flash(f'Own equipment "{machine.name}" assigned to project.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/own-equipment/<int:pm_id>/remove', methods=['POST'])
def own_equipment_remove(project_id, pm_id):
    pm = ProjectMachine.query.get_or_404(pm_id)
    db.session.delete(pm)
    db.session.commit()
    flash('Own equipment removed from project.', 'info')
    return redirect(url_for('project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Sunday Work Exceptions — add / remove
# ---------------------------------------------------------------------------

@app.route('/project/<int:project_id>/worked-sunday/add', methods=['POST'])
def worked_sunday_add(project_id):
    Project.query.get_or_404(project_id)
    date_str = request.form.get('date', '').strip()
    reason = request.form.get('reason', '').strip()
    if not date_str:
        flash('Date is required.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    try:
        ws_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    if ws_date.weekday() != 6:
        flash('That date is not a Sunday.', 'warning')
        return redirect(url_for('project_dashboard', project_id=project_id))
    existing = ProjectWorkedSunday.query.filter_by(project_id=project_id, date=ws_date).first()
    if existing:
        flash(f'{ws_date.strftime("%d/%m/%Y")} already added.', 'warning')
    else:
        db.session.add(ProjectWorkedSunday(project_id=project_id, date=ws_date,
                                           reason=reason or None))
        db.session.commit()
        flash(f'Worked Sunday {ws_date.strftime("%d/%m/%Y")} added.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/worked-sunday/<int:ws_id>/delete', methods=['POST'])
def worked_sunday_delete(project_id, ws_id):
    ws = ProjectWorkedSunday.query.get_or_404(ws_id)
    db.session.delete(ws)
    db.session.commit()
    flash('Worked Sunday removed.', 'info')
    return redirect(url_for('project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Project Documents — upload / download / delete
# ---------------------------------------------------------------------------

@app.route('/project/<int:project_id>/documents/upload', methods=['POST'])
def project_document_upload(project_id):
    Project.query.get_or_404(project_id)
    f = request.files.get('document')
    if not f or not f.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    if not allowed_doc(f.filename):
        flash('File type not allowed. Accepted: pdf, png, jpg, jpeg, dwg, dxf, doc, docx, xls, xlsx', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    doc_type = request.form.get('doc_type', 'other').strip()
    if doc_type not in ('drawing', 'specification', 'other'):
        doc_type = 'other'
    ext = f.filename.rsplit('.', 1)[1].lower()
    stored_name = f"doc_{uuid.uuid4().hex}.{ext}"
    proj_upload_dir = os.path.join(UPLOAD_FOLDER, 'projects', str(project_id))
    storage.upload_file(f, f'docs/{stored_name}', os.path.join(proj_upload_dir, stored_name))
    db.session.add(ProjectDocument(
        project_id=project_id,
        filename=stored_name,
        original_name=secure_filename(f.filename),
        doc_type=doc_type,
    ))
    db.session.commit()
    flash(f'Document "{f.filename}" uploaded.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/documents/<int:doc_id>/download')
def project_document_download(project_id, doc_id):
    doc = ProjectDocument.query.get_or_404(doc_id)
    if doc.project_id != project_id:
        flash('Document not found.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    proj_upload_dir = os.path.join(UPLOAD_FOLDER, 'projects', str(project_id))
    return storage.serve_file(f'docs/{doc.filename}', os.path.join(proj_upload_dir, doc.filename),
                              as_attachment=True, download_name=doc.original_name or doc.filename)


@app.route('/project/<int:project_id>/documents/<int:doc_id>/delete', methods=['POST'])
def project_document_delete(project_id, doc_id):
    doc = ProjectDocument.query.get_or_404(doc_id)
    if doc.project_id != project_id:
        flash('Document not found.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id))
    proj_upload_dir = os.path.join(UPLOAD_FOLDER, 'projects', str(project_id))
    storage.delete_file(f'docs/{doc.filename}', os.path.join(proj_upload_dir, doc.filename))
    db.session.delete(doc)
    db.session.commit()
    flash('Document deleted.', 'info')
    return redirect(url_for('project_dashboard', project_id=project_id))


# ---------------------------------------------------------------------------
# Project Progress PDF Report
# ---------------------------------------------------------------------------

def generate_project_report_pdf(project, progress, delay_summary, cost_estimate, settings,
                                 date_from=None, date_to=None, gantt_data=None):
    """Generate a progress report PDF for a project and return bytes."""
    from datetime import date as _date
    today = _date.today()

    def merge_spans(pct_list, day_w):
        """Merge contiguous left_pct positions into (start_pct, end_pct) spans."""
        if not pct_list:
            return []
        sorted_p = sorted(pct_list)
        spans = []
        s = sorted_p[0]; e = s + day_w
        for p in sorted_p[1:]:
            if p <= e + day_w * 0.6:
                e = p + day_w
            else:
                spans.append((s, min(e, 100))); s = p; e = p + day_w
        spans.append((s, min(e, 100)))
        return spans

    # ── Fetch entries for date range ─────────────────────────────────────
    entries_q = DailyEntry.query.filter_by(project_id=project.id).order_by(DailyEntry.entry_date)
    if date_from:
        entries_q = entries_q.filter(DailyEntry.entry_date >= date_from)
    if date_to:
        entries_q = entries_q.filter(DailyEntry.entry_date <= date_to)
    period_entries = entries_q.all()
    delay_entries = [e for e in period_entries if (e.delay_hours or 0) > 0]

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    company = safe(settings.get('company_name', '') or 'Project Tracker')

    if date_from or date_to:
        from_str = date_from.strftime('%d/%m/%Y') if date_from else 'Start'
        to_str = date_to.strftime('%d/%m/%Y') if date_to else today.strftime('%d/%m/%Y')
        header_period = safe(f'Progress Report -- {from_str} to {to_str}')
    else:
        header_period = safe(f'Progress Report -- {today.strftime("%d/%m/%Y")}')

    # ════════════════════════════════════════════════════════════════════
    # PAGE 1 — Portrait: Summary + Progress + Lot Bars
    # ════════════════════════════════════════════════════════════════════
    pdf.add_page()
    pdf.set_margins(15, 15, 15)

    def section_header(title):
        pdf.set_fill_color(240, 244, 255)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 7, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)

    def detail_row(label, value):
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(55, 6, label + ':')
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, safe(str(value)) if value is not None else '-',
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Header
    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 10, company, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 8, safe(project.name), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, header_period, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(4)

    # Project Summary
    section_header('PROJECT SUMMARY')
    detail_row('Start Date', project.start_date.strftime('%d/%m/%Y') if project.start_date else None)
    detail_row('Quoted Days', project.quoted_days)
    detail_row('Hours per Day', project.hours_per_day)

    if progress:
        all_entries_all = DailyEntry.query.filter_by(project_id=project.id).all()
        worked_dates = {e.entry_date for e in all_entries_all if e.install_hours and e.install_hours > 0}
        detail_row('Days Worked', len(worked_dates))
        detail_row('Overall Progress', f'{progress["overall_pct"]}%')
        detail_row('Total Planned', f'{progress["total_planned"]} m\u00b2')
        detail_row('Total Installed', f'{progress["total_actual"]} m\u00b2')
        detail_row('Remaining', f'{progress["total_remaining"]} m\u00b2')
        if progress.get('install_rate'):
            detail_row('Install Rate', f'{progress["install_rate"]} m\u00b2/hr')

    if gantt_data:
        pdf.ln(2)
        if gantt_data.get('target_finish'):
            detail_row('Target Finish', safe(gantt_data['target_finish']))
        if gantt_data.get('est_finish'):
            detail_row('Est. Finish', safe(gantt_data['est_finish']))
        if gantt_data.get('variance_days') is not None:
            v = gantt_data['variance_days']
            v_str = f'+{v} days (BEHIND)' if v > 0 else (f'{v} days (AHEAD)' if v < 0 else 'On Schedule')
            pdf.set_font('Helvetica', 'B', 9)
            if v > 0:
                pdf.set_text_color(180, 0, 0)
            elif v < 0:
                pdf.set_text_color(0, 140, 60)
            pdf.cell(55, 6, 'Variance:')
            pdf.cell(0, 6, safe(v_str), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # Progress by Lot / Material table
    if progress and progress.get('tasks'):
        section_header('PROGRESS BY LOT / MATERIAL')
        col_w = [30, 50, 28, 28, 24, 20]
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(220, 230, 255)
        for hdr, w in zip(['Lot', 'Material', 'Planned m\u00b2', 'Actual m\u00b2', '% Done', 'Status'], col_w):
            pdf.cell(w, 6, safe(hdr), border=1, fill=True)
        pdf.ln()

        pdf.set_font('Helvetica', '', 8)
        for task in progress['tasks']:
            pct = task['pct_complete']
            status = 'Complete' if pct >= 100 else ('In Progress' if pct > 0 else 'Not Started')
            pdf.cell(col_w[0], 5, safe(task['lot'] or '-'), border=1)
            pdf.cell(col_w[1], 5, safe(task['material'] or '-'), border=1)
            pdf.cell(col_w[2], 5, str(task['planned_sqm']), border=1, align='R')
            pdf.cell(col_w[3], 5, str(task['actual_sqm']), border=1, align='R')
            pdf.cell(col_w[4], 5, f'{pct}%', border=1, align='R')
            pdf.cell(col_w[5], 5, status, border=1)
            pdf.ln()

        pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(col_w[0] + col_w[1], 5, 'TOTAL', border=1)
        pdf.cell(col_w[2], 5, str(progress['total_planned']), border=1, align='R')
        pdf.cell(col_w[3], 5, str(progress['total_actual']), border=1, align='R')
        pdf.cell(col_w[4], 5, f"{progress['overall_pct']}%", border=1, align='R')
        pdf.cell(col_w[5], 5, '', border=1)
        pdf.ln()
        pdf.ln(4)

    # ════════════════════════════════════════════════════════════════════
    # PAGE 2 — Gantt Chart (screenshot via playwright)
    # ════════════════════════════════════════════════════════════════════
    if gantt_data and gantt_data.get('rows'):
        import tempfile, os as _os
        from playwright.sync_api import sync_playwright

        # Build self-contained Gantt HTML — same CSS/markup as dashboard
        n_rows = len(gantt_data['rows'])
        row_h_px = 56
        header_h_px = 48
        summary_h_px = 70
        total_h_px = header_h_px + n_rows * row_h_px + summary_h_px + 20

        def _esc(s):
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')

        rows_html = ''
        for row in gantt_data['rows']:
            var_class = ''
            var_text = ''
            v = row.get('variance_days')
            if v is not None:
                var_class = 'gantt-var-late' if v > 0 else ('gantt-var-early' if v < 0 else 'gantt-var-ontime')
                var_text = (f'+{v}d' if v > 0 else (f'{v}d' if v < 0 else 'On time'))

            stripes_html = ''.join(
                f'<div class="gantt-stripe gantt-stripe-{_esc(s["type"])}" '
                f'style="left:{s["left"]}%;width:{s["w"]}%;"></div>'
                for s in gantt_data.get('shade_stripes', [])
            )
            today_html = ''
            if gantt_data.get('today_pct') is not None and 0 <= gantt_data['today_pct'] <= 100:
                today_html = f'<div class="gantt-today-line" style="left:{gantt_data["today_pct"]}%;"></div>'
            target_html = ''
            if gantt_data.get('target_finish_pct') is not None and 0 <= gantt_data['target_finish_pct'] <= 100:
                target_html = f'<div class="gantt-target-line" style="left:{gantt_data["target_finish_pct"]}%;"></div>'
            planned_html = ''.join(
                f'<div class="gantt-day gantt-planned" style="left:{lft}%;width:{gantt_data["day_width_pct"]}%;"></div>'
                for lft in row.get('planned_days', [])
            )
            actual_html = ''.join(
                f'<div class="gantt-day gantt-actual" style="left:{lft}%;width:{gantt_data["day_width_pct"]}%;"></div>'
                for lft in row.get('actual_days', [])
            )
            forecast_html = ''.join(
                f'<div class="gantt-day gantt-forecast" style="left:{lft}%;width:{gantt_data["day_width_pct"]}%;"></div>'
                for lft in row.get('forecast_days', [])
            )
            rows_html += f'''
            <div style="display:flex;margin-bottom:4px;">
              <div style="min-width:160px;max-width:160px;padding-right:8px;display:flex;align-items:center;">
                <div style="font-size:0.74rem;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{_esc(row["label"])}</div>
              </div>
              <div style="flex:1;position:relative;height:{row_h_px}px;overflow:hidden;">
                {stripes_html}{today_html}{target_html}{planned_html}{actual_html}{forecast_html}
              </div>
              <div class="gantt-var {var_class}">{var_text}</div>
            </div>'''

        month_html = ''.join(
            f'<div class="gantt-month-label" style="left:{m["left_pct"]}%;">{_esc(m["label"])}</div>'
            for m in gantt_data.get('month_markers', [])
        )
        week_html = ''.join(
            f'<div class="gantt-week-tick" style="left:{w["left_pct"]}%;"></div>'
            f'<div class="gantt-week-label" style="left:{w["left_pct"]}%;">{_esc(w["date"])}</div>'
            for w in gantt_data.get('week_markers', [])
        )
        today_footer = ''
        if gantt_data.get('today_pct') is not None and 0 <= gantt_data['today_pct'] <= 100:
            today_footer = f'<div style="position:absolute;left:{gantt_data["today_pct"]}%;transform:translateX(-50%);font-size:0.6rem;font-weight:700;color:#dc3545;white-space:nowrap;">Today</div>'
        target_footer = ''
        if gantt_data.get('target_finish_pct') is not None:
            target_footer = f'<div style="position:absolute;left:{gantt_data["target_finish_pct"]}%;transform:translateX(-50%);font-size:0.6rem;font-weight:700;color:#0d6efd;white-space:nowrap;">Target</div>'

        var_days = gantt_data.get('variance_days')
        if var_days is not None:
            var_color = '#dc3545' if var_days > 0 else ('#198754' if var_days < 0 else '#6c757d')
            var_sign = '+' if var_days > 0 else ''
            var_label = 'BEHIND' if var_days > 0 else ('AHEAD' if var_days < 0 else 'ON SCHEDULE')
            summary_var = f'<div style="color:{var_color};font-weight:700;">VARIANCE: {var_sign}{var_days} DAYS ({var_label})</div>'
        else:
            summary_var = ''

        html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
body {{ margin: 0; padding: 12px; background: #fff; font-family: -apple-system, Arial, sans-serif; }}
.gantt-container {{ min-width: 900px; }}
.gantt-header-row {{ position:relative; height:44px; border-bottom:2px solid #dee2e6; margin-bottom:4px; }}
.gantt-month-label {{ position:absolute; top:2px; font-size:0.7rem; font-weight:600; color:#495057; white-space:nowrap; transform:translateX(-50%); }}
.gantt-week-tick {{ position:absolute; bottom:0; width:1px; height:10px; background:#ced4da; transform:translateX(-50%); }}
.gantt-week-label {{ position:absolute; bottom:12px; font-size:0.6rem; color:#adb5bd; transform:translateX(-50%); white-space:nowrap; }}
.gantt-stripe {{ position:absolute; top:0; height:100%; pointer-events:none; z-index:0; }}
.gantt-stripe-sun {{ background:rgba(255,105,180,0.18); }}
.gantt-stripe-nwd {{ background:rgba(100,180,255,0.28); }}
.gantt-stripe-wwd {{ background:rgba(0,160,210,0.28); }}
.gantt-stripe-cld {{ background:rgba(255,140,0,0.30); }}
.gantt-today-line {{ position:absolute; top:0; height:100%; width:2px; background:#dc3545; z-index:10; pointer-events:none; }}
.gantt-target-line {{ position:absolute; top:0; height:100%; width:2px; background:#0d6efd; z-index:9; pointer-events:none; }}
.gantt-day {{ position:absolute; border-radius:2px; pointer-events:none; z-index:5; }}
.gantt-planned {{ background:deeppink; opacity:0.35; top:2px; height:22px; }}
.gantt-actual {{ background:#00c060; border:1px solid rgba(0,0,0,0.18); opacity:0.9; top:26px; height:22px; }}
.gantt-forecast {{ background:gold; border:1px dashed rgba(0,0,0,0.25); opacity:0.85; top:26px; height:22px; }}
.gantt-var {{ min-width:72px; text-align:left; padding-left:8px; font-size:0.72rem; font-weight:700; line-height:1.2; display:flex; align-items:center; }}
.gantt-var-late {{ color:#dc3545; }}
.gantt-var-early {{ color:#198754; }}
.gantt-var-ontime {{ color:#6c757d; }}
</style></head><body>
<div class="gantt-container">
  <div style="display:flex;">
    <div style="min-width:160px;max-width:160px;"></div>
    <div style="flex:1;position:relative;" class="gantt-header-row">
      {month_html}{week_html}
    </div>
    <div style="min-width:72px;"></div>
  </div>
  {rows_html}
  <div style="display:flex;margin-top:4px;">
    <div style="min-width:160px;max-width:160px;"></div>
    <div style="flex:1;position:relative;height:18px;">{today_footer}{target_footer}</div>
    <div style="min-width:72px;"></div>
  </div>
  <div style="margin-top:10px;padding:10px;background:#f8f9fa;font-family:monospace;border:1px solid #dee2e6;font-size:0.82rem;">
    <div><strong>TARGET FINISH:</strong> {_esc(gantt_data.get("target_finish") or "-")}</div>
    <div><strong>EST. FINISH:</strong> {_esc(gantt_data.get("est_finish") or "-")}</div>
    {summary_var}
  </div>
</div>
</body></html>'''

        tmp_html = tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8')
        tmp_html.write(html)
        tmp_html.close()
        tmp_img = tmp_html.name.replace('.html', '.png')

        try:
            with sync_playwright() as pw:
                browser = pw.chromium.launch()
                page = browser.new_page(viewport={'width': 1300, 'height': total_h_px + 40})
                page.goto(f'file:///{tmp_html.name.replace(chr(92), "/")}')
                page.wait_for_timeout(300)
                page.locator('.gantt-container').screenshot(path=tmp_img)
                browser.close()

            pdf.add_page('L')
            pdf.set_margins(10, 10, 10)
            pdf.set_auto_page_break(auto=False)
            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 8, safe(f'Schedule Gantt -- {project.name}'), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            img_y = pdf.get_y()
            avail_w = pdf.w - pdf.l_margin - pdf.r_margin
            avail_h = pdf.h - img_y - pdf.b_margin - 2
            pdf.image(tmp_img, x=pdf.l_margin, y=img_y, w=avail_w, h=avail_h, keep_aspect_ratio=True)
        finally:
            _os.unlink(tmp_html.name)
            if _os.path.exists(tmp_img):
                _os.unlink(tmp_img)

        pdf.set_margins(15, 15, 15)
        pdf.set_auto_page_break(auto=True, margin=15)

    # ════════════════════════════════════════════════════════════════════
    # PAGE 3+ — Portrait: Daily Activities
    # ════════════════════════════════════════════════════════════════════
    if period_entries:
        pdf.add_page()
        section_header('DAILY ACTIVITIES')
        if date_from or date_to:
            pdf.set_font('Helvetica', 'I', 8)
            pdf.set_text_color(80, 80, 80)
            pdf.cell(0, 5,
                     safe(f'Period: {date_from.strftime("%d/%m/%Y") if date_from else "Start"}'
                          f' to {date_to.strftime("%d/%m/%Y") if date_to else today.strftime("%d/%m/%Y")}'),
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_text_color(0, 0, 0)

        for entry in period_entries:
            date_str = entry.entry_date.strftime('%A, %d %B %Y')
            lot_str = entry.lot_number or ''
            loc_str = entry.location or ''
            mat_str = entry.material or ''

            # Entry date header
            pdf.set_fill_color(228, 238, 255)
            pdf.set_font('Helvetica', 'B', 9)
            hdr_parts = [safe(date_str)]
            if lot_str:
                hdr_parts.append(safe(f'Lot {lot_str}'))
            if loc_str:
                hdr_parts.append(safe(loc_str))
            pdf.cell(0, 6, '  |  '.join(hdr_parts), new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)

            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(0, 0, 0)

            # Work details
            work_parts = []
            if mat_str:
                work_parts.append(safe(f'Material: {mat_str}'))
            if entry.install_hours:
                work_parts.append(safe(f'Install: {entry.install_hours}h'))
            if entry.install_sqm:
                work_parts.append(safe(f'{entry.install_sqm} m\u00b2'))
            if work_parts:
                pdf.cell(0, 5, '   '.join(work_parts), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            # Delay
            if entry.delay_hours and entry.delay_hours > 0:
                delay_type = 'Client (Billable)' if entry.delay_billable else 'Own (Non-billable)'
                pdf.set_text_color(160, 70, 0)
                pdf.cell(0, 5,
                         safe(f'Delay: {entry.delay_hours}h -- {entry.delay_reason or "N/A"} ({delay_type})'),
                         new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                if entry.delay_description:
                    pdf.set_font('Helvetica', 'I', 8)
                    pdf.multi_cell(0, 4, safe(f'  {entry.delay_description}'),
                                   new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.set_font('Helvetica', '', 8)

            # Machines stood down
            if entry.machines_stood_down:
                pdf.set_text_color(0, 110, 130)
                pdf.cell(0, 5, 'Hired machines stood down (wet weather)',
                         new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            # Other work
            if entry.other_work_description:
                pdf.set_text_color(50, 100, 50)
                pdf.multi_cell(0, 4, safe(f'Other work: {entry.other_work_description}'),
                               new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            # Notes
            if entry.notes:
                pdf.set_text_color(80, 80, 80)
                pdf.set_font('Helvetica', 'I', 8)
                pdf.multi_cell(0, 4, safe(f'Notes: {entry.notes}'),
                               new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font('Helvetica', '', 8)

            pdf.set_text_color(0, 0, 0)
            pdf.ln(2)

    # ════════════════════════════════════════════════════════════════════
    # Delay Details in Period
    # ════════════════════════════════════════════════════════════════════
    if delay_entries:
        if not period_entries:
            pdf.add_page()
        pdf.ln(2)
        section_header('DELAY DETAILS IN PERIOD')

        for entry in delay_entries:
            date_str = entry.entry_date.strftime('%d/%m/%Y')
            delay_type = 'Client (Billable)' if entry.delay_billable else 'Own (Non-billable)'
            pdf.set_fill_color(255, 243, 226)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.cell(0, 5,
                     safe(f'{date_str}  --  {entry.delay_hours}h  {entry.delay_reason or ""}  ({delay_type})'),
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
            if entry.delay_description:
                pdf.set_font('Helvetica', '', 8)
                pdf.set_text_color(80, 40, 0)
                pdf.multi_cell(0, 4, safe(entry.delay_description),
                               new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_text_color(0, 0, 0)
            pdf.ln(1)

    # ════════════════════════════════════════════════════════════════════
    # All-time Delay Summary
    # ════════════════════════════════════════════════════════════════════
    if delay_summary:
        pdf.ln(3)
        section_header('ALL-TIME DELAY SUMMARY')
        col_w2 = [52, 38, 22, 22, 28]
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(255, 220, 200)
        for hdr, w in zip(['Reason', 'Type', 'Events', 'Hours', 'Schedule Impact'], col_w2):
            pdf.cell(w, 6, hdr, border=1, fill=True)
        pdf.ln()

        pdf.set_font('Helvetica', '', 8)
        hrs_per_day = project.hours_per_day or 8
        for cat in delay_summary:
            impact = round(cat['hours'] / hrs_per_day, 1)
            pdf.cell(col_w2[0], 5, safe(cat['reason']), border=1)
            pdf.cell(col_w2[1], 5, 'Client (Billable)' if cat['billable'] else 'Own (Non-billable)', border=1)
            pdf.cell(col_w2[2], 5, str(cat['events']), border=1, align='R')
            pdf.cell(col_w2[3], 5, f"{cat['hours']}h", border=1, align='R')
            pdf.cell(col_w2[4], 5, f"~{impact}d", border=1, align='R')
            pdf.ln()
        pdf.ln(4)

    # ── Footer ───────────────────────────────────────────────────────────
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, f'Generated: {today.strftime("%d/%m/%Y")}',
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')

    return bytes(pdf.output())


@app.route('/project/<int:project_id>/report/pdf')
def project_report_pdf(project_id):
    project = Project.query.get_or_404(project_id)

    # Date range for filtering entries
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str = request.args.get('date_to', '').strip()
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except ValueError:
        date_from = None
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        date_to = None

    progress = compute_project_progress(project_id)
    delay_summary = compute_delay_summary(project_id)
    gantt_data = compute_gantt_data(project_id)
    settings = load_settings()

    progress_dict = None
    if progress:
        progress_dict = {
            'tasks': progress['tasks'],
            'total_planned': progress['total_planned'],
            'total_actual': progress['total_actual'],
            'total_remaining': progress['total_remaining'],
            'overall_pct': progress['overall_pct'],
            'install_rate': progress.get('install_rate'),
        }

    pdf_bytes = generate_project_report_pdf(
        project, progress_dict, delay_summary, {}, settings,
        date_from=date_from, date_to=date_to, gantt_data=gantt_data
    )
    safe_name = project.name.replace(' ', '_').replace('/', '-')
    from_str = date_from.strftime('%Y%m%d') if date_from else 'all'
    to_str = date_to.strftime('%Y%m%d') if date_to else date.today().strftime('%Y%m%d')
    filename = f"Progress_Report_{safe_name}_{from_str}_to_{to_str}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Weekly Progress Report PDF (client distribution)
# ---------------------------------------------------------------------------

def generate_weekly_report_pdf(project, week_start, week_end, entries, settings):
    """Generate a weekly progress report PDF for client distribution."""
    from datetime import date as _date

    pdf = FPDF()
    pdf.set_margins(15, 15, 15)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    company = safe(settings.get('company_name', '') or 'Project Tracker')

    # Header
    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 10, company, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 8, safe(project.name), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 10)
    period = safe(f'Weekly Progress Report: {week_start.strftime("%d/%m/%Y")} - {week_end.strftime("%d/%m/%Y")}')
    pdf.cell(0, 6, period, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(4)

    # Week summary stats
    total_sqm = sum(e.install_sqm or 0 for e in entries)
    total_hours = sum(e.install_hours or 0 for e in entries)
    total_delay = sum(e.delay_hours or 0 for e in entries)
    days_with_install = len({e.entry_date for e in entries if (e.install_sqm or 0) > 0})

    pdf.set_fill_color(240, 244, 255)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'WEEK SUMMARY', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
    pdf.set_font('Helvetica', '', 9)

    def kv(label, value):
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(60, 6, label + ':')
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, safe(str(value)), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    kv('Reporting Period', f'{week_start.strftime("%d/%m/%Y")} to {week_end.strftime("%d/%m/%Y")}')
    kv('Days with Installation', str(days_with_install))
    kv('Total Installed', f'{round(total_sqm, 1)} m2')
    kv('Total Install Hours', f'{round(total_hours, 1)} hrs')
    if total_delay > 0:
        kv('Total Delay Hours', f'{round(total_delay, 1)} hrs')
    pdf.ln(4)

    # Daily breakdown table
    if entries:
        pdf.set_fill_color(220, 230, 255)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, 7, 'DAILY BREAKDOWN', new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)

        col_w = [22, 20, 28, 40, 22, 22, 22]
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(200, 215, 255)
        for h, w in zip(['Date', 'Day', 'Lot', 'Material', 'Hrs', 'm2', 'Delay'], col_w):
            pdf.cell(w, 6, h, border=1, fill=True, align='C')
        pdf.ln()

        sorted_entries = sorted(entries, key=lambda e: e.entry_date)
        pdf.set_font('Helvetica', '', 8)
        for e in sorted_entries:
            pdf.cell(col_w[0], 5, e.entry_date.strftime('%d/%m/%y'), border=1, align='C')
            pdf.cell(col_w[1], 5, e.entry_date.strftime('%a'), border=1, align='C')
            pdf.cell(col_w[2], 5, safe(e.lot_number or '-'), border=1)
            pdf.cell(col_w[3], 5, safe(e.material or '-'), border=1)
            pdf.cell(col_w[4], 5, str(round(e.install_hours or 0, 1)), border=1, align='R')
            pdf.cell(col_w[5], 5, str(round(e.install_sqm or 0, 1)), border=1, align='R')
            pdf.cell(col_w[6], 5, str(round(e.delay_hours or 0, 1)) if e.delay_hours else '-', border=1, align='R')
            pdf.ln()

            # Notes / delay description as sub-row if present
            note_parts = []
            if e.delay_reason and e.delay_hours and e.delay_hours > 0:
                note_parts.append(safe(f'Delay: {e.delay_reason}'))
            if e.delay_description:
                note_parts.append(safe(e.delay_description))
            if e.notes:
                note_parts.append(safe(e.notes))
            if e.other_work_description:
                note_parts.append(safe(f'Other work: {e.other_work_description}'))
            if note_parts:
                pdf.set_font('Helvetica', 'I', 7)
                pdf.set_fill_color(248, 248, 248)
                pdf.cell(sum(col_w), 4, safe(' | '.join(note_parts)), border=1, fill=True)
                pdf.ln()
                pdf.set_font('Helvetica', '', 8)
        pdf.ln(4)
    else:
        pdf.set_font('Helvetica', 'I', 9)
        pdf.cell(0, 6, 'No entries recorded for this week.', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(4)

    # Footer
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, safe(f'Generated: {_date.today().strftime("%d/%m/%Y")}'),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')

    return bytes(pdf.output())


@app.route('/project/<int:project_id>/weekly-report/pdf')
def project_weekly_report_pdf(project_id):
    project = Project.query.get_or_404(project_id)
    week_start_str = request.args.get('week_start', '').strip()
    today = date.today()
    if week_start_str:
        try:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        except ValueError:
            week_start = today - timedelta(days=today.weekday())
    else:
        week_start = today - timedelta(days=today.weekday())  # Monday of current week
    week_end = week_start + timedelta(days=6)

    entries = (DailyEntry.query
               .filter_by(project_id=project_id)
               .filter(DailyEntry.entry_date >= week_start)
               .filter(DailyEntry.entry_date <= week_end)
               .order_by(DailyEntry.entry_date)
               .all())

    settings = load_settings()
    pdf_bytes = generate_weekly_report_pdf(project, week_start, week_end, entries, settings)
    safe_name = project.name.replace(' ', '_').replace('/', '-')
    filename = f"Weekly_Report_{safe_name}_{week_start.strftime('%Y%m%d')}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Morning Standdown Email
# ---------------------------------------------------------------------------

@app.route('/morning-standdown')
def morning_standdown():
    date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    # Find all standdowns on this date
    standdowns = (StandDown.query
                  .filter_by(stand_down_date=selected_date)
                  .join(HiredMachine)
                  .filter(HiredMachine.active == True)
                  .all())

    # Group by hire company email
    companies = {}
    for sd in standdowns:
        hm = sd.hired_machine
        email = hm.hire_company_email or 'No email'
        company = hm.hire_company or 'Unknown'
        key = email
        if key not in companies:
            companies[key] = {
                'email': email,
                'company': company,
                'machines': [],
                'reasons': set(),
                'photos': [],
            }
        companies[key]['machines'].append(hm)
        if sd.reason:
            companies[key]['reasons'].add(sd.reason)
        # Attach photos from the linked entry
        if sd.entry_id:
            entry = DailyEntry.query.get(sd.entry_id)
            if entry:
                for photo in entry.photos:
                    companies[key]['photos'].append(photo)

    # Convert sets to lists for template
    for key in companies:
        companies[key]['reasons'] = list(companies[key]['reasons'])

    return render_template('morning_standdown.html',
                           companies=list(companies.values()),
                           selected_date=selected_date,
                           date_str=date_str,
                           standdown_count=len(standdowns))


# ---------------------------------------------------------------------------
# Hire Tracker — List & Add
# ---------------------------------------------------------------------------

@app.route('/hire')
def hire_list():
    project_filter = request.args.get('project_id', '')
    query = HiredMachine.query.order_by(HiredMachine.created_at.desc())
    if project_filter:
        query = query.filter_by(project_id=int(project_filter))
    hired = query.all()
    projects = Project.query.order_by(Project.name).all()
    return render_template('hire/list.html', hired=hired, projects=projects,
                           project_filter=project_filter)


@app.route('/hire/new', methods=['GET', 'POST'])
def hire_new():
    projects = Project.query.filter_by(active=True).order_by(Project.name).all()

    if request.method == 'POST':
        project_id = request.form.get('project_id')
        machine_name = request.form.get('machine_name', '').strip()
        if not project_id or not machine_name:
            flash('Project and machine name are required.', 'danger')
            return render_template('hire/form.html', projects=projects)

        hm = HiredMachine(
            project_id=int(project_id),
            machine_name=machine_name,
            plant_id=request.form.get('plant_id', '').strip() or None,
            machine_type=request.form.get('machine_type', '').strip() or None,
            description=request.form.get('description', '').strip() or None,
            hire_company=request.form.get('hire_company', '').strip() or None,
            hire_company_email=request.form.get('hire_company_email', '').strip() or None,
            hire_company_phone=request.form.get('hire_company_phone', '').strip() or None,
            cost_per_week=float(request.form.get('cost_per_week')) if request.form.get('cost_per_week') else None,
            count_saturdays='count_saturdays' in request.form,
            notes=request.form.get('notes', '').strip() or None,
        )
        delivery_str = request.form.get('delivery_date', '').strip()
        if delivery_str:
            try:
                hm.delivery_date = datetime.strptime(delivery_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        return_str = request.form.get('return_date', '').strip()
        if return_str:
            try:
                hm.return_date = datetime.strptime(return_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        file = request.files.get('invoice_file')
        if file and file.filename and allowed_file(file.filename):
            ext = file.filename.rsplit('.', 1)[1].lower()
            stored_name = f"{uuid.uuid4().hex}.{ext}"
            local_path = os.path.join(UPLOAD_FOLDER, stored_name)
            storage.upload_file(file, f'invoices/{stored_name}', local_path)
            hm.invoice_filename = stored_name
            hm.invoice_original_name = secure_filename(file.filename)
        db.session.add(hm)
        db.session.commit()
        flash(f'Hired machine "{machine_name}" added.', 'success')
        return redirect(url_for('hire_detail', hm_id=hm.id))

    return render_template('hire/form.html', projects=projects)


@app.route('/hire/<int:hm_id>')
def hire_detail(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    return render_template('hire/detail.html', hm=hm, week_start=week_start, week_end=week_end)


@app.route('/hire/<int:hm_id>/edit', methods=['GET', 'POST'])
def hire_edit(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    projects = Project.query.filter_by(active=True).order_by(Project.name).all()
    if request.method == 'POST':
        hm.project_id = int(request.form.get('project_id'))
        hm.machine_name = request.form.get('machine_name', '').strip()
        hm.plant_id = request.form.get('plant_id', '').strip() or None
        hm.machine_type = request.form.get('machine_type', '').strip() or None
        hm.description = request.form.get('description', '').strip() or None
        hm.hire_company = request.form.get('hire_company', '').strip() or None
        hm.hire_company_email = request.form.get('hire_company_email', '').strip() or None
        hm.hire_company_phone = request.form.get('hire_company_phone', '').strip() or None
        hm.notes = request.form.get('notes', '').strip() or None
        hm.cost_per_week = float(request.form.get('cost_per_week')) if request.form.get('cost_per_week') else None
        hm.count_saturdays = 'count_saturdays' in request.form
        delivery_str = request.form.get('delivery_date', '').strip()
        hm.delivery_date = datetime.strptime(delivery_str, '%Y-%m-%d').date() if delivery_str else None
        return_str = request.form.get('return_date', '').strip()
        hm.return_date = datetime.strptime(return_str, '%Y-%m-%d').date() if return_str else None
        file = request.files.get('invoice_file')
        if file and file.filename and allowed_file(file.filename):
            if hm.invoice_filename:
                storage.delete_file(f'invoices/{hm.invoice_filename}',
                                    os.path.join(UPLOAD_FOLDER, hm.invoice_filename))
            ext = file.filename.rsplit('.', 1)[1].lower()
            stored_name = f"{uuid.uuid4().hex}.{ext}"
            local_path = os.path.join(UPLOAD_FOLDER, stored_name)
            storage.upload_file(file, f'invoices/{stored_name}', local_path)
            hm.invoice_filename = stored_name
            hm.invoice_original_name = secure_filename(file.filename)
        db.session.commit()
        flash('Machine hire record updated.', 'success')
        return redirect(url_for('hire_detail', hm_id=hm.id))
    return render_template('hire/form.html', hm=hm, projects=projects)


@app.route('/hire/<int:hm_id>/delete', methods=['POST'])
def hire_delete(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    if hm.invoice_filename:
        storage.delete_file(f'invoices/{hm.invoice_filename}',
                            os.path.join(UPLOAD_FOLDER, hm.invoice_filename))
    db.session.delete(hm)
    db.session.commit()
    flash('Hire record deleted.', 'info')
    return redirect(url_for('hire_list'))


@app.route('/hire/<int:hm_id>/invoice')
def hire_invoice(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    if not hm.invoice_filename:
        flash('No file attached.', 'warning')
        return redirect(url_for('hire_detail', hm_id=hm_id))
    return storage.serve_file(f'invoices/{hm.invoice_filename}',
                              os.path.join(UPLOAD_FOLDER, hm.invoice_filename),
                              download_name=hm.invoice_original_name,
                              as_attachment=False)


@app.route('/hire/<int:hm_id>/standdown/add', methods=['POST'])
def standdown_add(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    date_str = request.form.get('stand_down_date', '').strip()
    reason = request.form.get('reason', '').strip()
    if not date_str or not reason:
        flash('Date and reason are required.', 'danger')
        return redirect(url_for('hire_detail', hm_id=hm_id))
    try:
        sd_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('hire_detail', hm_id=hm_id))
    existing = StandDown.query.filter_by(hired_machine_id=hm_id, stand_down_date=sd_date).first()
    if existing:
        flash(f'{sd_date.strftime("%d/%m/%Y")} is already recorded as a stand-down.', 'warning')
        return redirect(url_for('hire_detail', hm_id=hm_id))
    db.session.add(StandDown(hired_machine_id=hm_id, stand_down_date=sd_date, reason=reason))
    db.session.commit()
    flash(f'Stand-down recorded for {sd_date.strftime("%d/%m/%Y")}.', 'success')
    return redirect(url_for('hire_detail', hm_id=hm_id))


@app.route('/hire/<int:hm_id>/standdown/<int:sd_id>/delete', methods=['POST'])
def standdown_delete(hm_id, sd_id):
    sd = StandDown.query.get_or_404(sd_id)
    db.session.delete(sd)
    db.session.commit()
    flash('Stand-down removed.', 'info')
    return redirect(url_for('hire_detail', hm_id=hm_id))


@app.route('/hire/<int:hm_id>/report')
def hire_report(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    date_from_str = request.args.get('date_from', week_start.strftime('%Y-%m-%d'))
    date_to_str = request.args.get('date_to', week_end.strftime('%Y-%m-%d'))
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        date_from, date_to = week_start, week_end
    days, summary = build_day_summary(hm, date_from, date_to)
    settings = load_settings()
    return render_template('hire/report.html', hm=hm, days=days, summary=summary,
                           date_from=date_from, date_to=date_to,
                           date_from_str=date_from_str, date_to_str=date_to_str,
                           settings=settings, timedelta=timedelta)


@app.route('/hire/<int:hm_id>/report/pdf')
def hire_report_pdf(hm_id):
    hm = HiredMachine.query.get_or_404(hm_id)
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        today = date.today()
        date_from = today - timedelta(days=today.weekday())
        date_to = date_from + timedelta(days=6)
    days, summary = build_day_summary(hm, date_from, date_to)
    settings = load_settings()
    pdf_bytes = generate_pdf(hm, date_from, date_to, days, summary, settings)
    filename = f"standdown_{hm.machine_name.replace(' ', '_')}_{date_from_str}_to_{date_to_str}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})




# ---------------------------------------------------------------------------
# Admin — Users
# ---------------------------------------------------------------------------

@app.route('/admin/users')
def admin_users():
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))
    users = User.query.order_by(User.username).all()
    return render_template('admin/users.html', users=users)


@app.route('/admin/users/add', methods=['POST'])
def admin_users_add():
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))
    from werkzeug.security import generate_password_hash
    username = request.form.get('username', '').strip().lower()
    display_name = request.form.get('display_name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    is_admin = request.form.get('is_admin') == 'on'
    if not username or not password:
        flash('Username and password are required.', 'danger')
        return redirect(url_for('admin_users'))
    if User.query.filter_by(username=username).first():
        flash(f'Username "{username}" is already taken.', 'danger')
        return redirect(url_for('admin_users'))
    user = User(
        username=username,
        display_name=display_name or username,
        email=email or None,
        password_hash=generate_password_hash(password),
        is_admin=is_admin,
        active=True,
    )
    db.session.add(user)
    db.session.commit()
    flash(f'User "{username}" created successfully.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
def admin_users_toggle(user_id):
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'danger')
        return redirect(url_for('admin_users'))
    user.active = not user.active
    db.session.commit()
    status = 'activated' if user.active else 'deactivated'
    flash(f'User "{user.username}" {status}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
def admin_users_reset_password(user_id):
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))
    from werkzeug.security import generate_password_hash
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password', '')
    if not new_password:
        flash('New password cannot be empty.', 'danger')
        return redirect(url_for('admin_users'))
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    flash(f'Password for "{user.username}" reset successfully.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
def admin_users_toggle_admin(user_id):
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot change your own admin status.', 'danger')
        return redirect(url_for('admin_users'))
    user.is_admin = not user.is_admin
    db.session.commit()
    status = 'granted admin' if user.is_admin else 'removed admin from'
    flash(f'Successfully {status} "{user.username}".', 'success')
    return redirect(url_for('admin_users'))


@app.route('/account/change-password', methods=['GET', 'POST'])
def change_password():
    """Allow any logged-in user to change their own password."""
    from werkzeug.security import check_password_hash, generate_password_hash
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        if not check_password_hash(current_user.password_hash, current_pw):
            flash('Current password is incorrect.', 'danger')
        elif len(new_pw) < 6:
            flash('New password must be at least 6 characters.', 'danger')
        elif new_pw != confirm_pw:
            flash('New passwords do not match.', 'danger')
        else:
            current_user.password_hash = generate_password_hash(new_pw)
            db.session.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('index'))
    return render_template('change_password.html')


# ---------------------------------------------------------------------------
# Admin — SQLite → Postgres one-time migration
# ---------------------------------------------------------------------------

@app.route('/admin/migrate-from-sqlite')
@login_required
def admin_migrate_sqlite():
    """One-time migration: read the old SQLite file on the volume and import into Postgres."""
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))

    db_url = app.config['SQLALCHEMY_DATABASE_URI']
    if not db_url.startswith('postgresql'):
        flash('This tool only works when the app is using PostgreSQL. You are currently on SQLite.', 'warning')
        return redirect(url_for('admin_settings'))

    # Find the SQLite file — try common Railway paths
    sqlite_candidates = [
        os.path.join(os.path.dirname(__file__), 'instance', 'tracking.db'),
        '/app/instance/tracking.db',
        '/data/tracking.db',
        '/var/data/tracking.db',
    ]
    sqlite_path = next((p for p in sqlite_candidates if os.path.exists(p)), None)
    if not sqlite_path:
        flash('SQLite file not found. It may have already been cleaned up, or was never on this volume.', 'danger')
        return redirect(url_for('admin_settings'))

    import sqlite3 as _sqlite3
    conn = _sqlite3.connect(sqlite_path)
    conn.row_factory = _sqlite3.Row

    def tbl(name):
        try:
            cur = conn.execute(f"SELECT * FROM {name}")
            return [dict(r) for r in cur.fetchall()]
        except Exception:
            return []

    def assoc(name):
        try:
            cur = conn.execute(f"SELECT * FROM {name}")
            return [list(r) for r in cur.fetchall()]
        except Exception:
            return []

    data = {
        'role':                  tbl('role'),
        'project':               tbl('project'),
        'employee':              tbl('employee'),
        'machine':               tbl('machine'),
        'daily_entry':           tbl('daily_entry'),
        'entry_employees':       assoc('entry_employees'),
        'entry_machines':        assoc('entry_machines'),
        'hired_machine':         tbl('hired_machine'),
        'stand_down':            tbl('stand_down'),
        'planned_data':          tbl('planned_data'),
        'project_non_work_date': tbl('project_non_work_date'),
        'project_budgeted_role': tbl('project_budgeted_role'),
        'project_machine':       tbl('project_machine'),
        'project_worked_sunday': tbl('project_worked_sunday'),
        'public_holiday':        tbl('public_holiday'),
        'cfmeu_date':            tbl('cfmeu_date'),
        'swing_pattern':         tbl('swing_pattern'),
        'employee_swing':        tbl('employee_swing'),
        'user':                  tbl('user'),
    }
    conn.close()

    total = sum(len(v) for v in data.values())
    if total == 0:
        flash('SQLite file exists but appears to be empty — nothing to migrate.', 'warning')
        return redirect(url_for('admin_settings'))

    # Save as JSON and feed through the existing import route logic
    import io
    json_bytes = json.dumps(data, default=str).encode('utf-8')
    json_file = io.BytesIO(json_bytes)
    json_file.filename = 'export.json'

    # Store in session and redirect to import page with auto-trigger
    # Instead, redirect with the data available via a temp file
    tmp_path = os.path.join(os.path.dirname(__file__), 'instance', '_sqlite_migration.json')
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, default=str)

    flash(f'SQLite file found with {total} rows. Use the Import Data page to upload the migration file, '
          f'or visit /admin/migrate-from-sqlite/run to run it automatically.', 'info')
    return redirect(url_for('admin_migrate_sqlite_run'))


@app.route('/admin/migrate-from-sqlite/run')
@login_required
def admin_migrate_sqlite_run():
    """Actually perform the SQLite → Postgres migration using the temp JSON file."""
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))

    tmp_path = os.path.join(os.path.dirname(__file__), 'instance', '_sqlite_migration.json')
    if not os.path.exists(tmp_path):
        flash('No migration file found. Visit /admin/migrate-from-sqlite first.', 'danger')
        return redirect(url_for('admin_settings'))

    with open(tmp_path, encoding='utf-8') as f:
        data = json.load(f)

    from sqlalchemy import text as _text

    def _d(val):
        if not val:
            return None
        try:
            return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
        except Exception:
            return None

    counts = {}
    try:
        # Roles
        role_id_map = {}
        for r in data.get('role', []):
            existing = Role.query.filter_by(name=r['name']).first()
            if not existing:
                obj = Role(name=r['name'], delay_rate=r.get('delay_rate'), group_name=r.get('group_name'))
                db.session.add(obj)
                db.session.flush()
                role_id_map[r['id']] = obj.id
            else:
                role_id_map[r['id']] = existing.id
        counts['roles'] = len(role_id_map)

        # Projects
        proj_id_map = {}
        for p in data.get('project', []):
            existing = Project.query.filter_by(name=p['name']).first()
            if existing:
                proj_id_map[p['id']] = existing.id
                continue
            obj = Project(
                name=p['name'], description=p.get('description'),
                active=bool(p.get('active', True)),
                start_date=_d(p.get('start_date')),
                planned_crew=p.get('planned_crew'),
                hours_per_day=p.get('hours_per_day'),
                quoted_days=p.get('quoted_days'),
                state=p.get('state'), is_cfmeu=bool(p.get('is_cfmeu', False)),
            )
            db.session.add(obj)
            db.session.flush()
            proj_id_map[p['id']] = obj.id
        counts['projects'] = len(proj_id_map)

        # Employees
        emp_id_map = {}
        for e in data.get('employee', []):
            existing = Employee.query.filter_by(name=e['name']).first()
            if existing:
                emp_id_map[e['id']] = existing.id
                continue
            obj = Employee(
                name=e['name'], role=e.get('role'),
                role_id=role_id_map.get(e['role_id']) if e.get('role_id') else None,
                delay_rate=e.get('delay_rate'), active=bool(e.get('active', True)),
            )
            db.session.add(obj)
            db.session.flush()
            emp_id_map[e['id']] = obj.id
        counts['employees'] = len(emp_id_map)

        # Machines
        mach_id_map = {}
        for m in data.get('machine', []):
            existing = Machine.query.filter_by(name=m['name']).first()
            if existing:
                mach_id_map[m['id']] = existing.id
                continue
            obj = Machine(
                name=m['name'], machine_type=m.get('machine_type'),
                delay_rate=m.get('delay_rate'), active=bool(m.get('active', True)),
            )
            db.session.add(obj)
            db.session.flush()
            mach_id_map[m['id']] = obj.id
        counts['machines'] = len(mach_id_map)

        # Daily Entries
        entry_id_map = {}
        for e in data.get('daily_entry', []):
            new_proj_id = proj_id_map.get(e['project_id'])
            if not new_proj_id:
                continue
            obj = DailyEntry(
                project_id=new_proj_id, entry_date=_d(e['entry_date']),
                lot_number=e.get('lot_number'), location=e.get('location'),
                material=e.get('material'), num_people=e.get('num_people'),
                install_hours=e.get('install_hours') or 0,
                install_sqm=e.get('install_sqm') or 0,
                delay_hours=e.get('delay_hours') or 0,
                delay_billable=bool(e.get('delay_billable', True)),
                delay_reason=e.get('delay_reason'),
                delay_description=e.get('delay_description'),
                machines_stood_down=bool(e.get('machines_stood_down', False)),
                notes=e.get('notes'),
                other_work_description=e.get('other_work_description'),
                weather=e.get('weather'),
            )
            db.session.add(obj)
            db.session.flush()
            entry_id_map[e['id']] = obj.id
        counts['entries'] = len(entry_id_map)

        # Entry ↔ Employee
        assoc_e = 0
        for row in data.get('entry_employees', []):
            new_entry_id = entry_id_map.get(row[0])
            new_emp_id = emp_id_map.get(row[1])
            if new_entry_id and new_emp_id:
                db.session.execute(_text(
                    'INSERT INTO entry_employees (entry_id, employee_id) VALUES (:e, :m)'
                ), {'e': new_entry_id, 'm': new_emp_id})
                assoc_e += 1
        counts['entry_employees'] = assoc_e

        # Entry ↔ Machine
        assoc_m = 0
        for row in data.get('entry_machines', []):
            new_entry_id = entry_id_map.get(row[0])
            new_mach_id = mach_id_map.get(row[1])
            if new_entry_id and new_mach_id:
                db.session.execute(_text(
                    'INSERT INTO entry_machines (entry_id, machine_id) VALUES (:e, :m)'
                ), {'e': new_entry_id, 'm': new_mach_id})
                assoc_m += 1
        counts['entry_machines'] = assoc_m

        # Hired Machines
        hm_id_map = {}
        for h in data.get('hired_machine', []):
            new_proj_id = proj_id_map.get(h['project_id'])
            if not new_proj_id:
                continue
            obj = HiredMachine(
                project_id=new_proj_id, machine_name=h['machine_name'],
                machine_type=h.get('machine_type'), hire_company=h.get('hire_company'),
                delivery_date=_d(h.get('delivery_date')), return_date=_d(h.get('return_date')),
                cost_per_day=h.get('cost_per_day'), cost_per_week=h.get('cost_per_week'),
                active=bool(h.get('active', True)), notes=h.get('notes'),
            )
            db.session.add(obj)
            db.session.flush()
            hm_id_map[h['id']] = obj.id
        counts['hired_machines'] = len(hm_id_map)

        # Public Holidays
        ph_count = 0
        for h in data.get('public_holiday', []):
            existing = PublicHoliday.query.filter_by(date=_d(h['date']), name=h['name']).first()
            if not existing:
                db.session.add(PublicHoliday(
                    date=_d(h['date']), name=h['name'],
                    state=h.get('state', 'ALL'), recurring=bool(h.get('recurring', True)),
                ))
                ph_count += 1
        counts['public_holidays'] = ph_count

        # CFMEU Dates
        cfmeu_count = 0
        for c in data.get('cfmeu_date', []):
            existing = CFMEUDate.query.filter_by(date=_d(c['date'])).first()
            if not existing:
                db.session.add(CFMEUDate(
                    date=_d(c['date']), name=c.get('name', ''),
                    state=c.get('state', 'ALL'),
                ))
                cfmeu_count += 1
        counts['cfmeu_dates'] = cfmeu_count

        # Non-work dates, budgeted roles, worked sundays
        for n in data.get('project_non_work_date', []):
            new_proj_id = proj_id_map.get(n['project_id'])
            if new_proj_id:
                db.session.add(ProjectNonWorkDate(project_id=new_proj_id, date=_d(n['date']), reason=n.get('reason')))
        for b in data.get('project_budgeted_role', []):
            new_proj_id = proj_id_map.get(b['project_id'])
            if new_proj_id:
                db.session.add(ProjectBudgetedRole(project_id=new_proj_id, role_name=b['role_name'], budgeted_count=b.get('budgeted_count', 1)))
        for w in data.get('project_worked_sunday', []):
            new_proj_id = proj_id_map.get(w['project_id'])
            if new_proj_id:
                db.session.add(ProjectWorkedSunday(project_id=new_proj_id, date=_d(w['date']), reason=w.get('reason')))

        db.session.commit()
        os.remove(tmp_path)  # Clean up temp file

        summary = ', '.join(f'{v} {k}' for k, v in counts.items() if v)
        flash(f'Migration successful! {summary}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Migration failed: {e}', 'danger')

    return redirect(url_for('admin_settings'))


# ---------------------------------------------------------------------------
# Admin — Data Import (migrate local data to production)
# ---------------------------------------------------------------------------

@app.route('/admin/import-data', methods=['GET', 'POST'])
def admin_import_data():
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        f = request.files.get('export_file')
        if not f or not f.filename.endswith('.json'):
            flash('Please upload a valid export.json file.', 'danger')
            return redirect(url_for('admin_import_data'))

        try:
            data = json.loads(f.read().decode('utf-8'))
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('admin_import_data'))

        from sqlalchemy import text as _text

        def _d(val):
            """Parse date string or return None."""
            if not val:
                return None
            try:
                return datetime.strptime(val[:10], '%Y-%m-%d').date()
            except Exception:
                return None

        def _dt(val):
            if not val:
                return None
            try:
                return datetime.strptime(val[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                return None

        counts = {}
        try:
            # ── Roles ──────────────────────────────────────────────────────
            role_id_map = {}
            for r in data.get('role', []):
                existing = Role.query.filter_by(name=r['name']).first()
                if not existing:
                    obj = Role(name=r['name'], delay_rate=r.get('delay_rate'))
                    db.session.add(obj)
                    db.session.flush()
                    role_id_map[r['id']] = obj.id
                else:
                    role_id_map[r['id']] = existing.id
            counts['roles'] = len(role_id_map)

            # ── Projects ───────────────────────────────────────────────────
            proj_id_map = {}
            for p in data.get('project', []):
                obj = Project(
                    name=p['name'],
                    description=p.get('description'),
                    active=bool(p.get('active', True)),
                    start_date=_d(p.get('start_date')),
                    planned_crew=p.get('planned_crew'),
                    hours_per_day=p.get('hours_per_day'),
                    quoted_days=p.get('quoted_days'),
                )
                db.session.add(obj)
                db.session.flush()
                proj_id_map[p['id']] = obj.id
            counts['projects'] = len(proj_id_map)

            # ── Employees ──────────────────────────────────────────────────
            emp_id_map = {}
            for e in data.get('employee', []):
                old_role_id = e.get('role_id')
                obj = Employee(
                    name=e['name'],
                    role=e.get('role'),
                    role_id=role_id_map.get(old_role_id) if old_role_id else None,
                    delay_rate=e.get('delay_rate'),
                    active=bool(e.get('active', True)),
                )
                db.session.add(obj)
                db.session.flush()
                emp_id_map[e['id']] = obj.id
            counts['employees'] = len(emp_id_map)

            # ── Machines ───────────────────────────────────────────────────
            mach_id_map = {}
            for m in data.get('machine', []):
                obj = Machine(
                    name=m['name'],
                    machine_type=m.get('machine_type'),
                    delay_rate=m.get('delay_rate'),
                    active=bool(m.get('active', True)),
                )
                db.session.add(obj)
                db.session.flush()
                mach_id_map[m['id']] = obj.id
            counts['machines'] = len(mach_id_map)

            # ── Daily Entries ──────────────────────────────────────────────
            entry_id_map = {}
            for e in data.get('daily_entry', []):
                new_proj_id = proj_id_map.get(e['project_id'])
                if not new_proj_id:
                    continue
                obj = DailyEntry(
                    project_id=new_proj_id,
                    entry_date=_d(e['entry_date']),
                    lot_number=e.get('lot_number'),
                    location=e.get('location'),
                    material=e.get('material'),
                    num_people=e.get('num_people'),
                    install_hours=e.get('install_hours') or 0,
                    install_sqm=e.get('install_sqm') or 0,
                    delay_hours=e.get('delay_hours') or 0,
                    delay_billable=bool(e.get('delay_billable', True)),
                    delay_reason=e.get('delay_reason'),
                    delay_description=e.get('delay_description'),
                    machines_stood_down=bool(e.get('machines_stood_down', False)),
                    notes=e.get('notes'),
                    other_work_description=e.get('other_work_description'),
                )
                db.session.add(obj)
                db.session.flush()
                entry_id_map[e['id']] = obj.id
            counts['entries'] = len(entry_id_map)

            # ── Entry ↔ Employee associations ──────────────────────────────
            assoc_e_count = 0
            for row in data.get('entry_employees', []):
                new_entry_id = entry_id_map.get(row[0])
                new_emp_id = emp_id_map.get(row[1])
                if new_entry_id and new_emp_id:
                    db.session.execute(
                        _text('INSERT INTO entry_employees (entry_id, employee_id) VALUES (:e, :m)'),
                        {'e': new_entry_id, 'm': new_emp_id}
                    )
                    assoc_e_count += 1
            counts['entry_employees'] = assoc_e_count

            # ── Entry ↔ Machine associations ───────────────────────────────
            assoc_m_count = 0
            for row in data.get('entry_machines', []):
                new_entry_id = entry_id_map.get(row[0])
                new_mach_id = mach_id_map.get(row[1])
                if new_entry_id and new_mach_id:
                    db.session.execute(
                        _text('INSERT INTO entry_machines (entry_id, machine_id) VALUES (:e, :m)'),
                        {'e': new_entry_id, 'm': new_mach_id}
                    )
                    assoc_m_count += 1
            counts['entry_machines'] = assoc_m_count

            # ── Hired Machines ─────────────────────────────────────────────
            hm_id_map = {}
            for h in data.get('hired_machine', []):
                new_proj_id = proj_id_map.get(h['project_id'])
                if not new_proj_id:
                    continue
                obj = HiredMachine(
                    project_id=new_proj_id,
                    machine_name=h['machine_name'],
                    machine_type=h.get('machine_type'),
                    hire_company=h.get('hire_company'),
                    hire_company_email=h.get('hire_company_email'),
                    hire_company_phone=h.get('hire_company_phone'),
                    delivery_date=_d(h.get('delivery_date')),
                    return_date=_d(h.get('return_date')),
                    cost_per_day=h.get('cost_per_day'),
                    cost_per_week=h.get('cost_per_week'),
                    count_saturdays=bool(h.get('count_saturdays', True)),
                    notes=h.get('notes'),
                    active=bool(h.get('active', True)),
                )
                db.session.add(obj)
                db.session.flush()
                hm_id_map[h['id']] = obj.id
            counts['hired_machines'] = len(hm_id_map)

            # ── Stand Downs ────────────────────────────────────────────────
            sd_count = 0
            for s in data.get('stand_down', []):
                new_hm_id = hm_id_map.get(s['hired_machine_id'])
                if not new_hm_id:
                    continue
                new_entry_id = entry_id_map.get(s.get('entry_id')) if s.get('entry_id') else None
                obj = StandDown(
                    hired_machine_id=new_hm_id,
                    entry_id=new_entry_id,
                    stand_down_date=_d(s['stand_down_date']),
                    reason=s.get('reason', ''),
                )
                db.session.add(obj)
                sd_count += 1
            counts['stand_downs'] = sd_count

            # ── Planned Data ───────────────────────────────────────────────
            pd_count = 0
            for p in data.get('planned_data', []):
                new_proj_id = proj_id_map.get(p['project_id'])
                if not new_proj_id:
                    continue
                obj = PlannedData(
                    project_id=new_proj_id,
                    lot=p.get('lot'),
                    location=p.get('location'),
                    material=p.get('material'),
                    day_number=p.get('day_number'),
                    planned_sqm=p.get('planned_sqm'),
                )
                db.session.add(obj)
                pd_count += 1
            counts['planned_data'] = pd_count

            # ── Project Non-Work Dates ─────────────────────────────────────
            nwd_count = 0
            for n in data.get('project_non_work_date', []):
                new_proj_id = proj_id_map.get(n['project_id'])
                if not new_proj_id:
                    continue
                db.session.add(ProjectNonWorkDate(
                    project_id=new_proj_id,
                    date=_d(n['date']),
                    reason=n.get('reason'),
                ))
                nwd_count += 1
            counts['non_work_dates'] = nwd_count

            # ── Project Budgeted Roles ─────────────────────────────────────
            br_count = 0
            for b in data.get('project_budgeted_role', []):
                new_proj_id = proj_id_map.get(b['project_id'])
                if not new_proj_id:
                    continue
                db.session.add(ProjectBudgetedRole(
                    project_id=new_proj_id,
                    role_name=b['role_name'],
                    budgeted_count=b.get('budgeted_count', 1),
                ))
                br_count += 1
            counts['budgeted_roles'] = br_count

            # ── Project Machines (own fleet) ───────────────────────────────
            pm_count = 0
            for p in data.get('project_machine', []):
                new_proj_id = proj_id_map.get(p['project_id'])
                new_mach_id = mach_id_map.get(p['machine_id'])
                if not new_proj_id or not new_mach_id:
                    continue
                db.session.add(ProjectMachine(
                    project_id=new_proj_id,
                    machine_id=new_mach_id,
                    assigned_date=_d(p.get('assigned_date')),
                    notes=p.get('notes'),
                ))
                pm_count += 1
            counts['project_machines'] = pm_count

            # ── Worked Sundays ─────────────────────────────────────────────
            ws_count = 0
            for w in data.get('project_worked_sunday', []):
                new_proj_id = proj_id_map.get(w['project_id'])
                if not new_proj_id:
                    continue
                db.session.add(ProjectWorkedSunday(
                    project_id=new_proj_id,
                    date=_d(w['date']),
                    reason=w.get('reason'),
                ))
                ws_count += 1
            counts['worked_sundays'] = ws_count

            db.session.commit()

            summary = ', '.join(f'{v} {k}' for k, v in counts.items() if v)
            flash(f'Import successful! {summary}. Note: uploaded photos/documents need to be re-uploaded manually.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {e}', 'danger')

        return redirect(url_for('admin_import_data'))

    # GET — show the import page
    entry_count = DailyEntry.query.count()
    project_count = Project.query.count()
    return render_template('admin/import_data.html',
                           entry_count=entry_count,
                           project_count=project_count)


# ---------------------------------------------------------------------------
# Admin — Projects / Employees / Machines / Roles / Settings
# ---------------------------------------------------------------------------

@app.route('/admin/projects', methods=['GET', 'POST'])
def admin_projects():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()
            if name:
                p = Project(name=name, description=description or None)
                start_str = request.form.get('start_date', '').strip()
                if start_str:
                    try:
                        p.start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                for attr, raw, cast in [
                    ('planned_crew', request.form.get('planned_crew', '').strip(), int),
                    ('hours_per_day', request.form.get('hours_per_day', '').strip(), float),
                    ('quoted_days', request.form.get('quoted_days', '').strip(), int),
                ]:
                    if raw:
                        try:
                            setattr(p, attr, cast(raw))
                        except ValueError:
                            pass
                db.session.add(p)
                db.session.commit()
                flash(f'Project "{name}" added.', 'success')
            else:
                flash('Project name is required.', 'danger')
        elif action == 'edit':
            project = Project.query.get_or_404(int(request.form.get('id')))
            project.name = request.form.get('name', '').strip()
            project.description = request.form.get('description', '').strip() or None
            start_str = request.form.get('start_date', '').strip()
            project.start_date = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else None
            planned_crew = request.form.get('planned_crew', '').strip()
            project.planned_crew = int(planned_crew) if planned_crew else None
            hours_per_day = request.form.get('hours_per_day', '').strip()
            project.hours_per_day = float(hours_per_day) if hours_per_day else None
            quoted_days = request.form.get('quoted_days', '').strip()
            project.quoted_days = int(quoted_days) if quoted_days else None
            db.session.commit()
            flash('Project updated.', 'success')
        elif action == 'toggle':
            project = Project.query.get_or_404(int(request.form.get('id')))
            project.active = not project.active
            db.session.commit()
            flash(f'Project "{project.name}" {"activated" if project.active else "deactivated"}.', 'info')
        elif action == 'delete':
            project = Project.query.get_or_404(int(request.form.get('id')))
            if project.entries:
                flash('Cannot delete — has existing entries. Deactivate instead.', 'danger')
            else:
                db.session.delete(project)
                db.session.commit()
                flash('Project deleted.', 'info')
        return redirect(url_for('admin_projects'))

    projects = Project.query.order_by(Project.name).all()
    return render_template('admin/projects.html', projects=projects)


@app.route('/admin/employees', methods=['GET', 'POST'])
def admin_employees():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip()
            role_ids = request.form.getlist('role_ids')
            delay_rate_raw = request.form.get('delay_rate', '').strip()
            if name:
                emp = Employee(name=name)
                db.session.add(emp)
                db.session.flush()  # get emp.id before setting m2m
                if role_ids:
                    role_objs = Role.query.filter(Role.id.in_([int(r) for r in role_ids])).all()
                    emp.roles = role_objs
                    emp.role = ', '.join(r.name for r in sorted(role_objs, key=lambda r: r.name))
                    emp.role_id = role_objs[0].id if len(role_objs) == 1 else None
                    if delay_rate_raw:
                        emp.delay_rate = float(delay_rate_raw)
                    else:
                        rates = [r.delay_rate for r in role_objs if r.delay_rate]
                        emp.delay_rate = max(rates) if rates else None
                else:
                    emp.delay_rate = float(delay_rate_raw) if delay_rate_raw else None
                db.session.commit()
                flash(f'Employee "{name}" added.', 'success')
            else:
                flash('Name is required.', 'danger')
        elif action == 'edit':
            emp = Employee.query.get_or_404(int(request.form.get('id')))
            emp.name = request.form.get('name', '').strip()
            role_ids = request.form.getlist('role_ids')
            delay_rate_raw = request.form.get('delay_rate', '').strip()
            if role_ids:
                role_objs = Role.query.filter(Role.id.in_([int(r) for r in role_ids])).all()
                emp.roles = role_objs
                emp.role = ', '.join(r.name for r in sorted(role_objs, key=lambda r: r.name))
                emp.role_id = role_objs[0].id if len(role_objs) == 1 else None
                if delay_rate_raw:
                    emp.delay_rate = float(delay_rate_raw)
                else:
                    rates = [r.delay_rate for r in role_objs if r.delay_rate]
                    emp.delay_rate = max(rates) if rates else None
            else:
                emp.roles = []
                emp.role_id = None
                emp.role = None
                emp.delay_rate = float(delay_rate_raw) if delay_rate_raw else None
            db.session.commit()
            flash('Employee updated.', 'success')
        elif action == 'toggle':
            emp = Employee.query.get_or_404(int(request.form.get('id')))
            emp.active = not emp.active
            db.session.commit()
            flash(f'"{emp.name}" {"activated" if emp.active else "deactivated"}.', 'info')
        elif action == 'delete':
            emp = Employee.query.get_or_404(int(request.form.get('id')))
            if emp.entries:
                flash('Cannot delete — has entries. Deactivate instead.', 'danger')
            else:
                db.session.delete(emp)
                db.session.commit()
                flash('Employee deleted.', 'info')
        return redirect(url_for('admin_employees'))

    employees = Employee.query.order_by(Employee.name).all()
    roles = Role.query.order_by(Role.name).all()
    return render_template('admin/employees.html', employees=employees, roles=roles)


@app.route('/admin/machines', methods=['GET', 'POST'])
def admin_machines():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip()
            machine_type = request.form.get('machine_type', '').strip()
            plant_id = request.form.get('plant_id', '').strip()
            description = request.form.get('description', '').strip()
            delay_rate = request.form.get('delay_rate', '').strip()
            if name:
                db.session.add(Machine(name=name, plant_id=plant_id or None,
                                       machine_type=machine_type or None,
                                       description=description or None,
                                       delay_rate=float(delay_rate) if delay_rate else None))
                db.session.commit()
                flash(f'Machine "{name}" added.', 'success')
            else:
                flash('Name is required.', 'danger')
        elif action == 'edit':
            machine = Machine.query.get_or_404(int(request.form.get('id')))
            machine.name = request.form.get('name', '').strip()
            machine.plant_id = request.form.get('plant_id', '').strip() or None
            machine.machine_type = request.form.get('machine_type', '').strip() or None
            machine.description = request.form.get('description', '').strip() or None
            delay_rate = request.form.get('delay_rate', '').strip()
            machine.delay_rate = float(delay_rate) if delay_rate else None
            db.session.commit()
            flash('Machine updated.', 'success')
        elif action == 'toggle':
            machine = Machine.query.get_or_404(int(request.form.get('id')))
            machine.active = not machine.active
            db.session.commit()
            flash(f'"{machine.name}" {"activated" if machine.active else "deactivated"}.', 'info')
        elif action == 'delete':
            machine = Machine.query.get_or_404(int(request.form.get('id')))
            if machine.entries:
                flash('Cannot delete — has entries. Deactivate instead.', 'danger')
            else:
                db.session.delete(machine)
                db.session.commit()
                flash('Machine deleted.', 'info')
        elif action == 'assign':
            machine_id = int(request.form.get('machine_id'))
            project_id = int(request.form.get('project_id'))
            if not ProjectMachine.query.filter_by(project_id=project_id, machine_id=machine_id).first():
                db.session.add(ProjectMachine(project_id=project_id, machine_id=machine_id))
                db.session.commit()
                flash('Machine assigned to project.', 'success')
            else:
                flash('Already assigned to that project.', 'warning')
        elif action == 'unassign':
            pm = ProjectMachine.query.get_or_404(int(request.form.get('pm_id')))
            db.session.delete(pm)
            db.session.commit()
            flash('Assignment removed.', 'info')
        return redirect(url_for('admin_machines'))

    machines = Machine.query.order_by(Machine.name).all()
    projects = Project.query.filter_by(active=True).order_by(Project.name).all()
    all_assignments = ProjectMachine.query.all()
    assignments_by_machine = {}
    for pm in all_assignments:
        assignments_by_machine.setdefault(pm.machine_id, []).append(pm)
    return render_template('admin/machines.html', machines=machines, projects=projects,
                           assignments_by_machine=assignments_by_machine)


@app.route('/admin/roles', methods=['GET', 'POST'])
def admin_roles():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip()
            delay_rate = request.form.get('delay_rate', '').strip()
            if name:
                db.session.add(Role(name=name, delay_rate=float(delay_rate) if delay_rate else None))
                db.session.commit()
                flash(f'Role "{name}" added.', 'success')
            else:
                flash('Role name is required.', 'danger')
        elif action == 'edit':
            role = Role.query.get_or_404(int(request.form.get('id')))
            role.name = request.form.get('name', '').strip()
            delay_rate = request.form.get('delay_rate', '').strip()
            role.delay_rate = float(delay_rate) if delay_rate else None
            group_name = request.form.get('group_name', '').strip()
            role.group_name = group_name or None
            db.session.commit()
            flash('Role updated.', 'success')
        elif action == 'delete':
            role = Role.query.get_or_404(int(request.form.get('id')))
            if role.employees:
                flash('Cannot delete — employees assigned to this role. Reassign first.', 'danger')
            else:
                db.session.delete(role)
                db.session.commit()
                flash('Role deleted.', 'info')
        return redirect(url_for('admin_roles'))

    roles = Role.query.order_by(Role.name).all()
    return render_template('admin/roles.html', roles=roles)


@app.route('/admin/holidays', methods=['GET', 'POST'])
@login_required
def admin_holidays():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            selected_states = request.form.getlist('states')
            date_str = request.form.get('date', '').strip()
            name = request.form.get('name', '').strip()
            if selected_states and date_str and name:
                try:
                    d = datetime.strptime(date_str, '%Y-%m-%d').date()
                    states_str = ','.join(selected_states)
                    db.session.add(PublicHoliday(states=states_str, date=d, name=name))
                    db.session.commit()
                    flash(f'Added: {name} ({states_str} — {d.strftime("%d/%m/%Y")}).', 'success')
                except ValueError:
                    flash('Invalid date.', 'danger')
            else:
                flash('Select at least one state, a date, and a name.', 'danger')
        elif action == 'delete':
            h = PublicHoliday.query.get(int(request.form.get('id', 0)))
            if h:
                db.session.delete(h)
                db.session.commit()
                flash('Holiday deleted.', 'success')
        return redirect(url_for('admin_holidays'))
    holidays = PublicHoliday.query.order_by(PublicHoliday.date).all()
    return render_template('admin/holidays.html', holidays=holidays, states=AUSTRALIAN_STATES)


@app.route('/admin/cfmeu', methods=['GET', 'POST'])
@login_required
def admin_cfmeu():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            date_str = request.form.get('date', '').strip()
            name = request.form.get('name', '').strip()
            selected_states = request.form.getlist('states')
            if date_str and name and selected_states:
                try:
                    d = datetime.strptime(date_str, '%Y-%m-%d').date()
                    states_str = ','.join(selected_states)
                    db.session.add(CFMEUDate(states=states_str, date=d, name=name))
                    db.session.commit()
                    flash(f'Added: {name} ({states_str} — {d.strftime("%d/%m/%Y")}).', 'success')
                except ValueError:
                    flash('Invalid date.', 'danger')
            else:
                flash('Select at least one state, a date, and a name.', 'danger')
        elif action == 'delete':
            c = CFMEUDate.query.get(int(request.form.get('id', 0)))
            if c:
                db.session.delete(c)
                db.session.commit()
                flash('CFMEU date deleted.', 'success')
        return redirect(url_for('admin_cfmeu'))
    cfmeu_dates = CFMEUDate.query.order_by(CFMEUDate.date).all()
    return render_template('admin/cfmeu.html', cfmeu_dates=cfmeu_dates, states=AUSTRALIAN_STATES)


@app.route('/admin/backup/download')
@login_required
def admin_backup_download():
    """Download a database backup. Admin only."""
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))

    db_url = app.config['SQLALCHEMY_DATABASE_URI']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if db_url.startswith('postgresql'):
        # Postgres — use pg_dump
        import subprocess, tempfile
        from urllib.parse import urlparse
        parsed = urlparse(db_url)
        env = os.environ.copy()
        env['PGPASSWORD'] = parsed.password or ''
        dump_file = os.path.join(tempfile.gettempdir(), f'plytrack_backup_{timestamp}.sql')
        cmd = [
            'pg_dump',
            '-h', parsed.hostname,
            '-p', str(parsed.port or 5432),
            '-U', parsed.username,
            '-d', parsed.path.lstrip('/'),
            '-f', dump_file,
            '--no-password',
        ]
        result = subprocess.run(cmd, env=env, capture_output=True, text=True)
        if result.returncode != 0:
            flash(f'pg_dump failed: {result.stderr[:200]}', 'danger')
            return redirect(url_for('admin_settings'))
        return send_file(
            dump_file,
            as_attachment=True,
            download_name=f'plytrack_backup_{timestamp}.sql',
            mimetype='application/sql',
        )
    else:
        # SQLite — send the .db file directly
        db_path = db_url.replace('sqlite:///', '')
        if not os.path.isabs(db_path):
            db_path = os.path.join(os.path.dirname(__file__), 'instance', 'tracking.db')
        if not os.path.exists(db_path):
            flash('Database file not found.', 'danger')
            return redirect(url_for('admin_settings'))
        return send_file(
            db_path,
            as_attachment=True,
            download_name=f'plytrack_backup_{timestamp}.db',
            mimetype='application/octet-stream',
        )


@app.route('/admin/settings', methods=['GET', 'POST'])
def admin_settings():
    settings = load_settings()
    if request.method == 'POST':
        settings['company_name'] = request.form.get('company_name', '').strip()
        settings['smtp_server'] = request.form.get('smtp_server', '').strip()
        settings['smtp_port'] = int(request.form.get('smtp_port') or 587)
        settings['smtp_username'] = request.form.get('smtp_username', '').strip()
        new_pw = request.form.get('smtp_password', '').strip()
        if new_pw:
            settings['smtp_password'] = new_pw
        settings['from_name'] = request.form.get('from_name', '').strip()
        settings['from_email'] = request.form.get('from_email', '').strip()
        save_settings(settings)
        flash('Settings saved.', 'success')
        return redirect(url_for('admin_settings'))
    return render_template('admin/settings.html', settings=settings)


# ---------------------------------------------------------------------------
# Delay Report — preview, PDF, email
# ---------------------------------------------------------------------------

@app.route('/delay-report')
def delay_report():
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    date_from_str = request.args.get('date_from', week_start.strftime('%Y-%m-%d'))
    date_to_str = request.args.get('date_to', week_end.strftime('%Y-%m-%d'))
    project_id = request.args.get('project_id', '')
    billable_filter = request.args.get('billable', 'all')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        date_from, date_to = week_start, week_end

    rows, summary = build_delay_report(project_id, date_from, date_to, billable_filter)
    projects = Project.query.order_by(Project.name).all()
    settings = load_settings()

    project_name = ''
    if project_id:
        p = Project.query.get(int(project_id))
        project_name = p.name if p else ''

    return render_template('delay_report.html',
                           rows=rows, summary=summary,
                           date_from=date_from, date_to=date_to,
                           date_from_str=date_from_str, date_to_str=date_to_str,
                           project_id=project_id, project_name=project_name,
                           projects=projects, settings=settings,
                           billable_filter=billable_filter)


@app.route('/delay-report/pdf')
def delay_report_pdf():
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    project_id = request.args.get('project_id', '')
    billable_filter = request.args.get('billable', 'all')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        today = date.today()
        date_from = today - timedelta(days=today.weekday())
        date_to = date_from + timedelta(days=6)

    rows, summary = build_delay_report(project_id, date_from, date_to, billable_filter)
    settings = load_settings()

    project_name = ''
    if project_id:
        p = Project.query.get(int(project_id))
        project_name = p.name if p else ''

    pdf_bytes = generate_delay_pdf(rows, summary, date_from, date_to, project_name, settings)
    filename = f"delay_report_{date_from_str}_to_{date_to_str}.pdf"
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})




# ---------------------------------------------------------------------------
# Scheduling helpers
# ---------------------------------------------------------------------------

def build_schedule_grid(employees, date_list):
    """
    Build a status grid for the given employees over the given dates.
    Returns: grid[employee_id][date_iso_str] = {status, label, project_name}
    Status priority: leave > rdo > assigned > available
    """
    from collections import defaultdict

    if not employees or not date_list:
        return {}

    emp_ids = [e.id for e in employees]
    min_date = min(date_list)
    max_date = max(date_list)

    # Bulk load all relevant data
    assignments = ProjectAssignment.query.filter(
        ProjectAssignment.employee_id.in_(emp_ids),
        ProjectAssignment.date_from <= max_date,
        db.or_(ProjectAssignment.date_to.is_(None), ProjectAssignment.date_to >= min_date)
    ).all()

    leaves = EmployeeLeave.query.filter(
        EmployeeLeave.employee_id.in_(emp_ids),
        EmployeeLeave.date_from <= max_date,
        EmployeeLeave.date_to >= min_date
    ).all()

    swings = EmployeeSwing.query.filter(
        EmployeeSwing.employee_id.in_(emp_ids),
        EmployeeSwing.start_date <= max_date
    ).order_by(EmployeeSwing.employee_id, EmployeeSwing.start_date).all()

    # Group by employee for O(1) lookup per employee
    assign_by_emp = defaultdict(list)
    for a in assignments:
        assign_by_emp[a.employee_id].append(a)

    leave_by_emp = defaultdict(list)
    for lv in leaves:
        leave_by_emp[lv.employee_id].append(lv)

    swing_by_emp = defaultdict(list)
    for s in swings:
        swing_by_emp[s.employee_id].append(s)

    overrides = ScheduleDayOverride.query.filter(
        ScheduleDayOverride.employee_id.in_(emp_ids),
        ScheduleDayOverride.date >= min_date,
        ScheduleDayOverride.date <= max_date
    ).all()
    override_by_emp = defaultdict(dict)
    for o in overrides:
        override_by_emp[o.employee_id][o.date] = o

    # Load project info (state + CFMEU flag) for all assigned projects
    proj_ids = {a.project_id for a in assignments}
    project_info = {}
    if proj_ids:
        for p in Project.query.filter(Project.id.in_(proj_ids)).all():
            project_info[p.id] = p

    # Load public holidays and CFMEU dates within the view window
    holidays_by_state = defaultdict(set)
    for h in PublicHoliday.query.filter(
            PublicHoliday.date >= min_date, PublicHoliday.date <= max_date).all():
        for s in h.states_list():
            holidays_by_state[s].add(h.date)
    cfmeu_by_state = defaultdict(set)
    for c in CFMEUDate.query.filter(
            CFMEUDate.date >= min_date, CFMEUDate.date <= max_date).all():
        for s in c.states_list():
            cfmeu_by_state[s].add(c.date)

    grid = {}
    for emp in employees:
        grid[emp.id] = {}
        emp_swings = swing_by_emp.get(emp.id, [])   # sorted by start_date ASC
        emp_leaves = leave_by_emp.get(emp.id, [])
        emp_assigns = assign_by_emp.get(emp.id, [])

        for d in date_list:
            date_str = d.isoformat()

            # Priority 1: Single-day override (explicit manual entry)
            override = override_by_emp.get(emp.id, {}).get(d)
            if override:
                _OV_LABELS = {
                    'available': 'Available', 'annual': 'Annual Leave', 'sick': 'Sick',
                    'personal': 'Personal', 'r_and_r': 'R&R', 'travel': 'Travel',
                    'rdo': 'RDO', 'other': 'Leave',
                }
                if override.status == 'project' and override.project:
                    grid[emp.id][date_str] = {
                        'status': 'assigned',
                        'label': override.project.name[:16],
                        'project_name': override.project.name,
                        'override_id': override.id,
                        'override_status': override.status,
                        'override_project_id': override.project_id or '',
                    }
                else:
                    css = override.status if override.status in (
                        'r_and_r', 'travel', 'rdo', 'available') else 'leave'
                    grid[emp.id][date_str] = {
                        'status': css,
                        'label': _OV_LABELS.get(override.status, override.status.replace('_', ' ').title()),
                        'project_name': '',
                        'override_id': override.id,
                        'override_status': override.status,
                        'override_project_id': '',
                    }
                continue

            # Priority 2: Leave (r_and_r and travel get distinct visual status)
            _LEAVE_LABELS = {
                'annual': 'Annual Leave', 'sick': 'Sick', 'personal': 'Personal',
                'r_and_r': 'R&R', 'travel': 'Travel', 'other': 'Leave'
            }
            leave = next((lv for lv in emp_leaves if lv.date_from <= d <= lv.date_to), None)
            if leave:
                lt = leave.leave_type
                grid[emp.id][date_str] = {
                    'status': lt if lt in ('r_and_r', 'travel') else 'leave',
                    'label': _LEAVE_LABELS.get(lt, lt.title()),
                    'project_name': ''
                }
                continue

            # Priority 3: RDO — find the most recent swing whose start_date <= d
            applicable_swing = None
            for s in emp_swings:
                if s.start_date <= d:
                    applicable_swing = s
                # swings are ordered ASC so keep going to find latest
            if applicable_swing and applicable_swing.is_rdo(d):
                grid[emp.id][date_str] = {
                    'status': 'r_and_r',
                    'label': 'R&R',
                    'project_name': ''
                }
                continue

            # Priority 4: Project-based non-work day (public holiday or CFMEU date)
            active_assign = next(
                (a for a in emp_assigns
                 if a.date_from <= d and (a.date_to is None or a.date_to >= d)),
                None
            )
            if active_assign:
                proj = project_info.get(active_assign.project_id)
                if proj:
                    is_pub_holiday = bool(proj.state and d in holidays_by_state.get(proj.state, set()))
                    is_cfmeu_day = bool(proj.is_cfmeu and (
                        d in cfmeu_by_state.get('ALL', set()) or
                        (proj.state and d in cfmeu_by_state.get(proj.state, set()))))
                    if is_pub_holiday or is_cfmeu_day:
                        grid[emp.id][date_str] = {
                            'status': 'rdo',
                            'label': 'PH' if is_pub_holiday else 'CFMEU',
                            'project_name': '',
                            'nwd_reason': 'Public Holiday' if is_pub_holiday else 'CFMEU RDO',
                        }
                        continue

            # Priority 5: Project assignment
            if active_assign:
                grid[emp.id][date_str] = {
                    'status': 'assigned',
                    'label': active_assign.project.name[:16],
                    'project_name': active_assign.project.name,
                    'project_id': active_assign.project_id,
                }
                continue

            # Default: sunday rest or available
            if d.weekday() == 6:
                grid[emp.id][date_str] = {
                    'status': 'sunday',
                    'label': 'Sun',
                    'project_name': ''
                }
            else:
                grid[emp.id][date_str] = {
                    'status': 'available',
                    'label': 'Available',
                    'project_name': ''
                }

    return grid


# ---------------------------------------------------------------------------
# Scheduling routes — overview
# ---------------------------------------------------------------------------

@app.route('/equipment/machine/save', methods=['POST'])
@login_required
def equipment_machine_save():
    action = request.form.get('action')
    if action == 'add':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Machine name is required.', 'danger')
            return redirect(url_for('equipment_overview') + '#tab-own')
        m = Machine(
            name=name,
            plant_id=request.form.get('plant_id', '').strip() or None,
            machine_type=request.form.get('machine_type', '').strip() or None,
            description=request.form.get('description', '').strip() or None,
            delay_rate=float(request.form.get('delay_rate')) if request.form.get('delay_rate') else None,
        )
        db.session.add(m)
        db.session.commit()
        flash(f'Machine "{name}" added.', 'success')
    elif action == 'edit':
        m = Machine.query.get_or_404(request.form.get('machine_id', type=int))
        m.name = request.form.get('name', '').strip()
        m.plant_id = request.form.get('plant_id', '').strip() or None
        m.machine_type = request.form.get('machine_type', '').strip() or None
        m.description = request.form.get('description', '').strip() or None
        m.delay_rate = float(request.form.get('delay_rate')) if request.form.get('delay_rate') else None
        db.session.commit()
        flash('Machine updated.', 'success')
    elif action == 'toggle':
        m = Machine.query.get_or_404(request.form.get('machine_id', type=int))
        m.active = not m.active
        db.session.commit()
        flash(f'"{m.name}" {"activated" if m.active else "deactivated"}.', 'info')
    elif action == 'delete':
        m = Machine.query.get_or_404(request.form.get('machine_id', type=int))
        if m.entries:
            flash('Cannot delete — machine has entries. Deactivate instead.', 'danger')
        else:
            db.session.delete(m)
            db.session.commit()
            flash('Machine deleted.', 'info')
    return redirect(url_for('equipment_overview') + '#tab-own')


@app.route('/equipment')
@login_required
def equipment_overview():
    own_machines = Machine.query.order_by(Machine.name).all()
    hired_machines = HiredMachine.query.order_by(HiredMachine.machine_name).all()
    projects = Project.query.filter_by(active=True).order_by(Project.name).all()

    # Active breakdown lookup: machine_id → breakdown, hired_machine_id → breakdown
    own_breakdowns = {b.machine_id: b for b in
                      MachineBreakdown.query.filter(
                          MachineBreakdown.machine_id.isnot(None),
                          MachineBreakdown.repair_status != 'completed'
                      ).all()}
    hired_breakdowns = {b.hired_machine_id: b for b in
                        MachineBreakdown.query.filter(
                            MachineBreakdown.hired_machine_id.isnot(None),
                            MachineBreakdown.repair_status != 'completed'
                        ).all()}

    # Current project assignment per machine (via ProjectEquipmentAssignment)
    all_assignments = (ProjectEquipmentAssignment.query
                       .join(ProjectEquipmentRequirement)
                       .join(Project)
                       .filter(Project.active == True)
                       .all())
    own_machine_projects = {}
    hired_machine_projects = {}
    for a in all_assignments:
        entry = (a.requirement.project, a.requirement.label)
        if a.machine_id:
            own_machine_projects.setdefault(a.machine_id, []).append(entry)
        if a.hired_machine_id:
            hired_machine_projects.setdefault(a.hired_machine_id, []).append(entry)

    # Full breakdown history (all statuses) per machine for history log
    from collections import defaultdict as _dd
    own_bd_history = _dd(list)
    for b in MachineBreakdown.query.filter(
            MachineBreakdown.machine_id.isnot(None)
    ).order_by(MachineBreakdown.incident_date.desc()).all():
        own_bd_history[b.machine_id].append(b)
    hired_bd_history = _dd(list)
    for b in MachineBreakdown.query.filter(
            MachineBreakdown.hired_machine_id.isnot(None)
    ).order_by(MachineBreakdown.incident_date.desc()).all():
        hired_bd_history[b.hired_machine_id].append(b)

    return render_template('equipment/index.html',
                           own_machines=own_machines,
                           hired_machines=hired_machines,
                           projects=projects,
                           own_breakdowns=own_breakdowns,
                           hired_breakdowns=hired_breakdowns,
                           own_machine_projects=own_machine_projects,
                           hired_machine_projects=hired_machine_projects,
                           own_bd_history=own_bd_history,
                           hired_bd_history=hired_bd_history,
                           today=date.today())


@app.route('/equipment/breakdown/add', methods=['POST'])
@login_required
def breakdown_add():
    machine_id = request.form.get('machine_id') or None
    hired_machine_id = request.form.get('hired_machine_id') or None
    incident_date_str = request.form.get('incident_date', '').strip()
    try:
        incident_date = datetime.strptime(incident_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('equipment_overview'))
    bd = MachineBreakdown(
        machine_id=int(machine_id) if machine_id else None,
        hired_machine_id=int(hired_machine_id) if hired_machine_id else None,
        incident_date=incident_date,
        incident_time=request.form.get('incident_time', '').strip() or None,
        description=request.form.get('description', '').strip() or None,
        repairing_by=request.form.get('repairing_by', '').strip() or None,
        repair_status=request.form.get('repair_status', 'pending'),
        anticipated_return=datetime.strptime(request.form['anticipated_return'], '%Y-%m-%d').date()
            if request.form.get('anticipated_return') else None,
    )
    db.session.add(bd)
    db.session.flush()
    # Handle photos
    photos = request.files.getlist('photos')
    for photo in photos:
        if photo and photo.filename:
            ext = os.path.splitext(photo.filename)[1].lower()
            stored = f"{uuid.uuid4()}{ext}"
            local_path = os.path.join(UPLOAD_FOLDER, 'breakdowns', stored)
            storage.upload_file(photo, f'breakdowns/{stored}', local_path)
            db.session.add(BreakdownPhoto(breakdown_id=bd.id, filename=stored, original_name=photo.filename))
    db.session.commit()
    flash('Breakdown recorded.', 'warning')
    return redirect(url_for('equipment_overview') + '#tab-own' if not hired_machine_id else '#tab-hired')


@app.route('/equipment/breakdown/<int:bd_id>/update', methods=['POST'])
@login_required
def breakdown_update(bd_id):
    bd = MachineBreakdown.query.get_or_404(bd_id)
    bd.repair_status = request.form.get('repair_status', bd.repair_status)
    bd.repairing_by = request.form.get('repairing_by', '').strip() or bd.repairing_by
    ret = request.form.get('anticipated_return', '').strip()
    if ret:
        try:
            bd.anticipated_return = datetime.strptime(ret, '%Y-%m-%d').date()
        except ValueError:
            pass
    if bd.repair_status == 'completed' and not bd.resolved_date:
        bd.resolved_date = date.today()
    db.session.commit()
    flash('Breakdown updated.', 'success')
    return redirect(url_for('equipment_overview'))


@app.route('/equipment/breakdown/<int:bd_id>/delete', methods=['POST'])
@login_required
def breakdown_delete(bd_id):
    bd = MachineBreakdown.query.get_or_404(bd_id)
    db.session.delete(bd)
    db.session.commit()
    flash('Breakdown record deleted.', 'success')
    return redirect(url_for('equipment_overview'))


@app.route('/scheduling')
def scheduling_overview():
    # Navigation: ?week=YYYY-MM-DD (Monday of start week), jumps in 4-week increments
    week_str = request.args.get('week')
    try:
        week_start = datetime.strptime(week_str, '%Y-%m-%d').date()
        week_start = week_start - timedelta(days=week_start.weekday())
    except (ValueError, TypeError):
        today = date.today()
        week_start = today - timedelta(days=today.weekday())

    # Show 12 weeks (84 days) — one full scroll covers 3 swing cycles
    date_list = [week_start + timedelta(days=i) for i in range(84)]
    week_end = date_list[-1]
    prev_period = (week_start - timedelta(weeks=4)).isoformat()
    next_period = (week_start + timedelta(weeks=4)).isoformat()

    employees = Employee.query.filter_by(active=True).order_by(Employee.role, Employee.name).all()
    grid = build_schedule_grid(employees, date_list)

    # Group employees by role group (if set) then by role name
    from itertools import groupby
    roles_db = Role.query.all()
    role_group_map = {r.name: (r.group_name or r.name) for r in roles_db}

    # Sort employees so they group by their display group name, then role, then name
    def emp_sort_key(e):
        grp = role_group_map.get(e.role, e.role or 'ZZZ')
        return (grp, e.role or '', e.name)
    employees_sorted = sorted(employees, key=emp_sort_key)

    roles_grouped = []
    for role_name, emp_iter in groupby(employees_sorted, key=lambda e: e.role or 'No Role'):
        roles_grouped.append((role_name, list(emp_iter)))

    projects = Project.query.filter_by(active=True).order_by(Project.name).all()

    all_proj_ordered = Project.query.order_by(Project.id).all()
    project_colour_map = {
        p.id: SCHED_PALETTE[i % len(SCHED_PALETTE)]
        for i, p in enumerate(all_proj_ordered)
    }

    return render_template(
        'scheduling/overview.html',
        date_list=date_list,
        week_start=week_start,
        week_end=week_end,
        prev_period=prev_period,
        next_period=next_period,
        employees=employees_sorted,
        roles_grouped=roles_grouped,
        grid=grid,
        projects=projects,
        project_colour_map=project_colour_map,
        today=date.today()
    )


@app.route('/scheduling/project/<int:project_id>')
def scheduling_project(project_id):
    project = Project.query.get_or_404(project_id)

    week_str = request.args.get('week')
    try:
        week_start = datetime.strptime(week_str, '%Y-%m-%d').date()
        week_start = week_start - timedelta(days=week_start.weekday())
    except (ValueError, TypeError):
        today = date.today()
        week_start = today - timedelta(days=today.weekday())

    # Show 12 weeks, navigate in 4-week increments
    date_list = [week_start + timedelta(days=i) for i in range(84)]
    week_end = date_list[-1]
    prev_period = (week_start - timedelta(weeks=4)).isoformat()
    next_period = (week_start + timedelta(weeks=4)).isoformat()

    # Employees assigned to this project whose assignment overlaps the view window
    assignments = ProjectAssignment.query.filter(
        ProjectAssignment.project_id == project_id,
        ProjectAssignment.date_from <= week_end,
        db.or_(ProjectAssignment.date_to.is_(None), ProjectAssignment.date_to >= week_start)
    ).order_by(ProjectAssignment.date_from).all()

    assigned_emp_ids = list({a.employee_id for a in assignments})
    assigned_employees = Employee.query.filter(Employee.id.in_(assigned_emp_ids)).order_by(Employee.role, Employee.name).all() if assigned_emp_ids else []

    grid = build_schedule_grid(assigned_employees, date_list)

    # Build role group mapping for coverage (roles in the same group count together)
    roles_db = Role.query.all()
    role_group_map = {r.name: (r.group_name or r.name) for r in roles_db}

    # Budgeted by role group name (ProjectBudgetedRole.role_name should match group names)
    budgeted = {br.role_name: br.budgeted_count for br in project.budgeted_roles}

    # Map employee → their scheduled role name for this project (from active assignments)
    emp_scheduled_role_name = {}
    for a in assignments:
        if a.scheduled_role:
            emp_scheduled_role_name[a.employee_id] = a.scheduled_role.name
        elif a.employee.role:
            emp_scheduled_role_name.setdefault(a.employee_id, a.employee.role)

    # Count employees on site (status='assigned') grouped by their role group per date
    role_coverage = {}   # group_name -> {date_str -> count}
    for emp in assigned_employees:
        role_name = emp_scheduled_role_name.get(emp.id) or emp.role or ''
        group = role_group_map.get(role_name, role_name or 'No Role')
        if group not in role_coverage:
            role_coverage[group] = {d.isoformat(): 0 for d in date_list}
        for d in date_list:
            ds = d.isoformat()
            cell = grid.get(emp.id, {}).get(ds, {})
            if cell.get('status') == 'assigned':
                role_coverage[group][ds] += 1

    all_employees = Employee.query.filter_by(active=True).order_by(Employee.role, Employee.name).all()
    # Build per-employee roles JSON for the assignment form's dynamic role dropdown
    emp_roles_data = {
        emp.id: [{'id': r.id, 'name': r.name, 'delay_rate': r.delay_rate} for r in emp.roles]
        for emp in all_employees
    }

    # Non-work dates including public holidays
    project_nwd = {nwd.date for nwd in ProjectNonWorkDate.query.filter_by(project_id=project_id).all()}
    if project.state:
        for h in PublicHoliday.query.all():
            if project.state in h.states_list():
                project_nwd.add(h.date)
        if project.is_cfmeu:
            for c in CFMEUDate.query.all():
                if 'ALL' in c.states_list() or project.state in c.states_list():
                    project_nwd.add(c.date)

    # Equipment requirements with coverage
    equip_reqs = (ProjectEquipmentRequirement.query
                  .filter_by(project_id=project_id)
                  .order_by(ProjectEquipmentRequirement.label)
                  .all())
    # Active breakdowns by machine id
    own_breakdowns = {b.machine_id: b for b in
                      MachineBreakdown.query.filter(
                          MachineBreakdown.machine_id.isnot(None),
                          MachineBreakdown.repair_status != 'completed').all()}
    hired_breakdowns = {b.hired_machine_id: b for b in
                        MachineBreakdown.query.filter(
                            MachineBreakdown.hired_machine_id.isnot(None),
                            MachineBreakdown.repair_status != 'completed').all()}
    equip_coverage = []
    for er in equip_reqs:
        assignments_for_req = er.assignments
        assigned = len(assignments_for_req)
        gap = max(0, er.required_count - assigned)
        machines = []
        for a in assignments_for_req:
            broken = (own_breakdowns.get(a.machine_id) if a.machine_id else
                      hired_breakdowns.get(a.hired_machine_id))
            machines.append({'assignment': a, 'broken': broken})
        equip_coverage.append({
            'req': er,
            'label': er.label,
            'required': er.required_count,
            'machines': machines,
            'assigned': assigned,
            'gap': gap,
        })

    all_projects = Project.query.filter_by(active=True).order_by(Project.name).all()

    all_proj_ordered = Project.query.order_by(Project.id).all()
    project_colour_map = {
        p.id: SCHED_PALETTE[i % len(SCHED_PALETTE)]
        for i, p in enumerate(all_proj_ordered)
    }

    return render_template(
        'scheduling/project.html',
        project=project,
        date_list=date_list,
        week_start=week_start,
        week_end=week_end,
        prev_period=prev_period,
        next_period=next_period,
        assignments=assignments,
        assigned_employees=assigned_employees,
        emp_scheduled_role_name=emp_scheduled_role_name,
        grid=grid,
        budgeted=budgeted,
        role_coverage=role_coverage,
        all_employees=all_employees,
        emp_roles_data=emp_roles_data,
        all_projects=all_projects,
        project_nwd=project_nwd,
        equip_coverage=equip_coverage,
        project_colour_map=project_colour_map,
        today=date.today()
    )


# ---------------------------------------------------------------------------
# Scheduling — assignment management
# ---------------------------------------------------------------------------

@app.route('/scheduling/assign/add', methods=['POST'])
def scheduling_assign_add():
    employee_id = request.form.get('employee_id', type=int)
    project_id = request.form.get('project_id', type=int)
    date_from_str = request.form.get('date_from', '').strip()
    date_to_str = request.form.get('date_to', '').strip()
    notes = request.form.get('notes', '').strip()
    redirect_to = request.form.get('redirect_to', 'scheduling_overview')
    redirect_project = request.form.get('redirect_project_id', type=int)

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        flash('Invalid date format.', 'danger')
        return redirect(url_for(redirect_to))

    if not employee_id or not project_id:
        flash('Employee and project are required.', 'danger')
        return redirect(url_for(redirect_to))

    scheduled_role_id = request.form.get('scheduled_role_id', type=int) or None
    pa = ProjectAssignment(
        employee_id=employee_id,
        project_id=project_id,
        date_from=date_from,
        date_to=date_to,
        notes=notes or None,
        scheduled_role_id=scheduled_role_id,
    )
    db.session.add(pa)
    db.session.commit()
    flash('Assignment added.', 'success')

    if redirect_to == 'scheduling_project' and redirect_project:
        return redirect(url_for('scheduling_project', project_id=redirect_project))
    return redirect(url_for('scheduling_overview'))


@app.route('/scheduling/assign/<int:pa_id>/delete', methods=['POST'])
def scheduling_assign_delete(pa_id):
    pa = ProjectAssignment.query.get_or_404(pa_id)
    project_id = pa.project_id
    redirect_to = request.form.get('redirect_to', 'scheduling_overview')
    db.session.delete(pa)
    db.session.commit()
    flash('Assignment removed.', 'success')
    if redirect_to == 'scheduling_project':
        return redirect(url_for('scheduling_project', project_id=project_id))
    return redirect(url_for('scheduling_overview'))


# ---------------------------------------------------------------------------
# Scheduling — leave management
# ---------------------------------------------------------------------------

@app.route('/scheduling/leave/add', methods=['POST'])
def scheduling_leave_add():
    employee_id = request.form.get('employee_id', type=int)
    date_from_str = request.form.get('date_from', '').strip()
    date_to_str = request.form.get('date_to', '').strip()
    leave_type = request.form.get('leave_type', 'annual').strip()
    notes = request.form.get('notes', '').strip()
    redirect_project = request.form.get('redirect_project_id', type=int)

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format.', 'danger')
        return redirect(url_for('scheduling_overview'))

    if not employee_id:
        flash('Employee is required.', 'danger')
        return redirect(url_for('scheduling_overview'))

    lv = EmployeeLeave(
        employee_id=employee_id,
        date_from=date_from,
        date_to=date_to,
        leave_type=leave_type,
        notes=notes or None
    )
    db.session.add(lv)
    db.session.commit()
    flash('Leave recorded.', 'success')

    if redirect_project:
        return redirect(url_for('scheduling_project', project_id=redirect_project))
    return redirect(url_for('scheduling_overview'))


@app.route('/scheduling/leave/<int:leave_id>/delete', methods=['POST'])
def scheduling_leave_delete(leave_id):
    lv = EmployeeLeave.query.get_or_404(leave_id)
    redirect_project = request.form.get('redirect_project_id', type=int)
    db.session.delete(lv)
    db.session.commit()
    flash('Leave removed.', 'success')
    if redirect_project:
        return redirect(url_for('scheduling_project', project_id=redirect_project))
    return redirect(url_for('scheduling_overview'))


# ---------------------------------------------------------------------------
# Scheduling — single-day override
# ---------------------------------------------------------------------------

@app.route('/scheduling/override', methods=['POST'])
@login_required
def schedule_override():
    employee_id = request.form.get('employee_id', type=int)
    date_str = request.form.get('date', '').strip()
    action = request.form.get('action', 'set')
    week = request.form.get('week', '')
    redirect_project = request.form.get('redirect_project_id', type=int)

    try:
        override_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date.', 'danger')
        if redirect_project:
            return redirect(url_for('scheduling_project', project_id=redirect_project, week=week))
        return redirect(url_for('scheduling_overview', week=week))

    existing = ScheduleDayOverride.query.filter_by(
        employee_id=employee_id, date=override_date
    ).first()

    if action == 'clear':
        if existing:
            db.session.delete(existing)
            db.session.commit()
    else:
        status = request.form.get('status', 'available')
        project_id = request.form.get('project_id', type=int)
        notes = request.form.get('notes', '').strip() or None
        if existing:
            existing.status = status
            existing.project_id = project_id if status == 'project' else None
            existing.notes = notes
        else:
            db.session.add(ScheduleDayOverride(
                employee_id=employee_id,
                date=override_date,
                status=status,
                project_id=project_id if status == 'project' else None,
                notes=notes,
            ))
        db.session.commit()

    if redirect_project:
        return redirect(url_for('scheduling_project', project_id=redirect_project, week=week))
    return redirect(url_for('scheduling_overview', week=week))


# ---------------------------------------------------------------------------
# Admin — swing patterns
# ---------------------------------------------------------------------------

@app.route('/admin/swings', methods=['GET', 'POST'])
def admin_swings():
    if not current_user.is_admin:
        flash('Admin access required.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_pattern':
            name = request.form.get('name', '').strip()
            work_weeks = request.form.get('work_weeks', type=int)
            off_days = request.form.get('off_days', type=int)
            description = request.form.get('description', '').strip()
            if not name or not work_weeks or not off_days:
                flash('Name, work weeks, and off days are required.', 'danger')
            else:
                sp = SwingPattern(name=name, work_weeks=work_weeks, off_days=off_days,
                                  description=description or None)
                db.session.add(sp)
                db.session.commit()
                flash(f'Pattern "{name}" added.', 'success')

        elif action == 'delete_pattern':
            pattern_id = request.form.get('pattern_id', type=int)
            sp = SwingPattern.query.get(pattern_id)
            if sp:
                db.session.delete(sp)
                db.session.commit()
                flash('Pattern deleted.', 'success')

        elif action == 'assign_swing':
            employee_id = request.form.get('employee_id', type=int)
            pattern_id = request.form.get('pattern_id', type=int)
            start_date_str = request.form.get('start_date', '').strip()
            day_offset = request.form.get('day_offset', 0, type=int)
            notes = request.form.get('notes', '').strip()
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid start date.', 'danger')
                return redirect(url_for('admin_swings'))
            if not employee_id or not pattern_id:
                flash('Employee and pattern are required.', 'danger')
            else:
                es = EmployeeSwing(employee_id=employee_id, pattern_id=pattern_id,
                                   start_date=start_date, day_offset=day_offset,
                                   notes=notes or None)
                db.session.add(es)
                db.session.commit()
                flash('Swing pattern assigned.', 'success')

        elif action == 'delete_swing':
            swing_id = request.form.get('swing_id', type=int)
            es = EmployeeSwing.query.get(swing_id)
            if es:
                db.session.delete(es)
                db.session.commit()
                flash('Swing assignment removed.', 'success')

        return redirect(url_for('admin_swings'))

    patterns = SwingPattern.query.order_by(SwingPattern.name).all()
    employees = Employee.query.filter_by(active=True).order_by(Employee.role, Employee.name).all()
    # Load all swing assignments with employee + pattern eager
    swing_assignments = EmployeeSwing.query.order_by(EmployeeSwing.employee_id, EmployeeSwing.start_date).all()

    return render_template('admin/swings.html',
                           patterns=patterns,
                           employees=employees,
                           swing_assignments=swing_assignments)


@app.route('/project/<int:project_id>/budgeted-crew/save-all', methods=['POST'])
def budgeted_crew_save_all(project_id):
    project = Project.query.get_or_404(project_id)
    # Delete all existing budgeted roles for this project
    ProjectBudgetedRole.query.filter_by(project_id=project_id).delete()
    # Re-add from form (group_name -> count)
    for key, val in request.form.items():
        if key.startswith('group_'):
            group_name = key[6:]  # strip 'group_' prefix
            try:
                count = int(val)
            except (ValueError, TypeError):
                count = 0
            if count > 0:
                db.session.add(ProjectBudgetedRole(
                    project_id=project_id,
                    role_name=group_name,
                    budgeted_count=count
                ))
    db.session.commit()
    flash('Crew requirements saved.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')


@app.route('/project/<int:project_id>/equipment-requirements/add', methods=['POST'])
def equipment_req_add(project_id):
    Project.query.get_or_404(project_id)
    label = request.form.get('label', '').strip()
    try:
        required_count = max(1, int(request.form.get('required_count', 1)))
    except (ValueError, TypeError):
        required_count = 1
    if not label:
        flash('Equipment name is required.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')
    db.session.add(ProjectEquipmentRequirement(
        project_id=project_id, label=label, required_count=required_count))
    db.session.commit()
    flash(f'Added requirement: {label}.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')


@app.route('/project/<int:project_id>/equipment-requirements/<int:req_id>/update', methods=['POST'])
def equipment_req_update(project_id, req_id):
    req = ProjectEquipmentRequirement.query.filter_by(id=req_id, project_id=project_id).first_or_404()
    label = request.form.get('label', '').strip()
    try:
        required_count = max(1, int(request.form.get('required_count', 1)))
    except (ValueError, TypeError):
        required_count = 1
    if label:
        req.label = label
    req.required_count = required_count
    db.session.commit()
    flash('Requirement updated.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')


@app.route('/project/<int:project_id>/equipment-requirements/<int:req_id>/delete', methods=['POST'])
def equipment_req_delete(project_id, req_id):
    req = ProjectEquipmentRequirement.query.filter_by(id=req_id, project_id=project_id).first_or_404()
    # cascade delete assignments
    ProjectEquipmentAssignment.query.filter_by(requirement_id=req_id).delete()
    db.session.delete(req)
    db.session.commit()
    flash('Requirement removed.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')


@app.route('/project/<int:project_id>/equipment-requirements/<int:req_id>/assign', methods=['POST'])
def equipment_req_assign(project_id, req_id):
    req = ProjectEquipmentRequirement.query.filter_by(id=req_id, project_id=project_id).first_or_404()
    machine_id = request.form.get('machine_id', '').strip()
    hired_machine_id = request.form.get('hired_machine_id', '').strip()
    if not machine_id and not hired_machine_id:
        flash('Select a machine to assign.', 'danger')
        return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')
    assignment = ProjectEquipmentAssignment(requirement_id=req.id)
    if machine_id:
        assignment.machine_id = int(machine_id)
    else:
        assignment.hired_machine_id = int(hired_machine_id)
    db.session.add(assignment)
    db.session.commit()
    flash('Machine assigned.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')


@app.route('/project/<int:project_id>/equipment-assignments/<int:assign_id>/remove', methods=['POST'])
def equipment_req_unassign(project_id, assign_id):
    a = ProjectEquipmentAssignment.query.filter_by(id=assign_id).first_or_404()
    # verify it belongs to this project
    if a.requirement.project_id != project_id:
        return 'Forbidden', 403
    db.session.delete(a)
    db.session.commit()
    flash('Machine removed from requirement.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-scheduling')


@app.route('/project/<int:project_id>/settings/save', methods=['POST'])
def project_settings_save(project_id):
    project = Project.query.get_or_404(project_id)
    project.name = request.form.get('name', '').strip() or project.name
    project.description = request.form.get('description', '').strip() or None
    start_date_str = request.form.get('start_date', '').strip()
    project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
    planned_crew = request.form.get('planned_crew', '').strip()
    project.planned_crew = int(planned_crew) if planned_crew else None
    hours_per_day = request.form.get('hours_per_day', '').strip()
    project.hours_per_day = float(hours_per_day) if hours_per_day else None
    quoted_days = request.form.get('quoted_days', '').strip()
    project.quoted_days = int(quoted_days) if quoted_days else None
    project.state = request.form.get('state', '').strip() or None
    project.is_cfmeu = bool(request.form.get('is_cfmeu'))
    db.session.commit()
    flash('Project settings saved.', 'success')
    return redirect(url_for('project_dashboard', project_id=project_id) + '#tab-settings')


# ---------------------------------------------------------------------------
# utils/ re-exports — keeps existing routes working during refactor
# ---------------------------------------------------------------------------
from utils.files import safe, allowed_file, allowed_photo, allowed_doc
from utils.settings import load_settings, save_settings
from utils.helpers import _natural_key
from utils.reports import (generate_pdf, generate_delay_pdf,
                            generate_project_report_pdf,
                            generate_weekly_report_pdf)
from utils.gantt import compute_gantt_data
from utils.progress import (compute_project_progress,
                             compute_delay_summary, build_delay_report)
from utils.schedule import build_schedule_grid, build_day_summary


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
